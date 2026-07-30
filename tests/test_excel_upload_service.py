import io
import re
import tempfile
import unittest
import zipfile
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import HTTPException
from openpyxl import Workbook

from classroom_app.services.excel_upload_service import (
    XLS_OLE_MAGIC,
    load_upload_workbook_bytes,
    open_upload_workbook_pair,
    sniff_excel_kind,
)


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "考核登分表"
    worksheet["A1"] = "序号"
    worksheet["B1"] = "学号"
    worksheet["C1"] = "姓名"
    worksheet["D1"] = "总分"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _without_worksheet_dimension(content: bytes) -> bytes:
    source = io.BytesIO(content)
    output = io.BytesIO()
    with zipfile.ZipFile(source) as input_archive, zipfile.ZipFile(output, "w") as output_archive:
        for info in input_archive.infolist():
            data = input_archive.read(info.filename)
            if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                data = re.sub(br"<dimension\b[^>]*/>", b"", data, count=1)
            output_archive.writestr(info, data)
    return output.getvalue()


class ExcelUploadServiceTests(unittest.TestCase):
    def test_content_hash_storage_path_loads_without_extension(self):
        content = _xlsx_bytes()
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ("a" * 64)
            path.write_bytes(content)
            with open_upload_workbook_pair(
                path,
                "考核登分表.xlsx",
                material_label="考核登分表",
            ) as (formula_workbook, value_workbook):
                self.assertEqual(formula_workbook.sheetnames, ["考核登分表"])
                self.assertEqual(value_workbook.sheetnames, ["考核登分表"])
                self.assertEqual(formula_workbook.active["A1"].value, "序号")

    def test_missing_worksheet_dimension_is_inferred_like_real_school_file(self):
        content = _without_worksheet_dimension(_xlsx_bytes())
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ("e" * 64)
            path.write_bytes(content)
            with open_upload_workbook_pair(
                path,
                "考核登分表.xlsx",
                material_label="考核登分表",
            ) as (formula_workbook, _value_workbook):
                worksheet = formula_workbook.active
                self.assertEqual(worksheet.max_row, 1)
                self.assertEqual(worksheet.max_column, 4)

    def test_xlsx_signature_requires_real_ooxml_parts(self):
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as archive:
            archive.writestr("not-an-excel.txt", "hello")
        content = output.getvalue()
        self.assertEqual(sniff_excel_kind(content[:8]), "xlsx")
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ("b" * 64)
            path.write_bytes(content)
            with self.assertRaises(HTTPException) as caught:
                load_upload_workbook_bytes(
                    path,
                    "伪装工作簿.xlsx",
                    material_label="考核登分表",
                )
        self.assertEqual(caught.exception.status_code, 422)
        self.assertIn("缺少 Excel 必要结构", str(caught.exception.detail))

    def test_non_excel_content_is_rejected_before_openpyxl(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ("c" * 64)
            path.write_bytes(b"student_number,score\n20240001,90\n")
            with self.assertRaises(HTTPException) as caught:
                load_upload_workbook_bytes(
                    path,
                    "改名文件.xlsx",
                    material_label="考核登分表",
                )
        self.assertEqual(caught.exception.status_code, 415)
        self.assertIn("不是有效的 Excel 工作簿", str(caught.exception.detail))

    def test_extensionless_legacy_xls_is_converted_with_xls_suffix(self):
        converted_content = _xlsx_bytes()
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ("d" * 64)
            path.write_bytes(XLS_OLE_MAGIC + b"legacy-excel")

            def fake_convert(source, output_format, *, timeout):
                self.assertEqual(Path(source).suffix, ".xls")
                self.assertEqual(output_format, "xlsx")
                self.assertEqual(timeout, 120)
                return SimpleNamespace(output_bytes=converted_content)

            with patch(
                "classroom_app.services.excel_upload_service.convert_office_file",
                side_effect=fake_convert,
            ):
                loaded = load_upload_workbook_bytes(
                    path,
                    "旧版考核登分表.xls",
                    material_label="考核登分表",
                )
        self.assertEqual(loaded, converted_content)


if __name__ == "__main__":
    unittest.main()
