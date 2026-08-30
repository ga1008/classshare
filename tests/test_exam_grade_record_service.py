import asyncio
import io
import json
import sqlite3

from classroom_app.db.schema_offering_class_links import ensure_offering_class_links_schema
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from classroom_app.services.exam_grade_record_service import (
    EXAM_GRADE_RECORD_TYPE,
    EXAM_GRADE_RECORD_TABLE_MODE,
    _load_exam_submissions,
    build_exam_grade_record_payload,
    build_exam_grade_record_xlsx,
    list_exam_grade_record_candidates,
    parse_exam_grade_record_file,
)
from classroom_app.services.material_ai_import_service import parse_material_document
from classroom_app.services.material_export_template_service import (
    XLSX_MEDIA_TYPE,
    build_material_export_artifact,
)


class ExamGradeRecordServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        # 每个用例都是全新内存库：重置模块级 _SCHEMA_READY，确保重修表按需重建。
        import classroom_app.db.schema_retake as schema_retake

        schema_retake._SCHEMA_READY = False
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self._create_schema()
        self._seed_data()
        ensure_offering_class_links_schema(self.conn, force=True, engine="sqlite")

    def tearDown(self) -> None:
        self.conn.close()

    def _create_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE teachers (
                id INTEGER PRIMARY KEY,
                name TEXT,
                college TEXT,
                department TEXT
            );
            CREATE TABLE courses (
                id INTEGER PRIMARY KEY,
                name TEXT,
                total_hours INTEGER,
                credits REAL,
                college TEXT,
                department TEXT
            );
            CREATE TABLE classes (
                id INTEGER PRIMARY KEY,
                name TEXT,
                college TEXT,
                department TEXT
            );
            CREATE TABLE class_offerings (
                id INTEGER PRIMARY KEY,
                class_id INTEGER,
                course_id INTEGER,
                teacher_id INTEGER,
                semester TEXT
            );
            CREATE TABLE students (
                id INTEGER PRIMARY KEY,
                class_id INTEGER,
                student_id_number TEXT,
                name TEXT,
                enrollment_status TEXT
            );
            CREATE TABLE exam_papers (
                id TEXT PRIMARY KEY,
                teacher_id INTEGER,
                title TEXT,
                description TEXT,
                questions_json TEXT,
                exam_config_json TEXT,
                status TEXT
            );
            CREATE TABLE assignments (
                id INTEGER PRIMARY KEY,
                title TEXT,
                status TEXT,
                exam_paper_id TEXT,
                created_at TEXT,
                due_at TEXT,
                grading_mode TEXT,
                class_offering_id INTEGER
            );
            CREATE TABLE submissions (
                id INTEGER PRIMARY KEY,
                assignment_id TEXT,
                student_pk_id INTEGER,
                student_name TEXT,
                status TEXT,
                score REAL,
                feedback_md TEXT,
                score_before_late_penalty REAL,
                late_penalty_points REAL,
                is_late_submission INTEGER,
                late_by_seconds INTEGER,
                late_score_cap_applied INTEGER
            );
            CREATE TABLE group_assignment_member_results (
                id INTEGER PRIMARY KEY,
                assignment_id TEXT,
                class_offering_id INTEGER,
                group_id INTEGER,
                student_pk_id INTEGER,
                submission_id INTEGER,
                work_score REAL,
                peer_avg REAL,
                peer_review_count INTEGER,
                final_score REAL,
                revealed INTEGER,
                finalized_at TEXT
            );
            """
        )

    def _seed_data(self) -> None:
        self.conn.execute("INSERT INTO teachers VALUES (1, '张海林', '数字科技学院', '软件工程系')")
        self.conn.execute("INSERT INTO courses VALUES (10, '服务器配置与管理', 48, 3.0, '数字科技学院', '软件工程系')")
        self.conn.execute("INSERT INTO classes VALUES (20, '软工2406班（专升本）', '数字科技学院', '软件工程系')")
        self.conn.execute("INSERT INTO class_offerings VALUES (30, 20, 10, 1, '2025-2026-1')")
        self.conn.executemany(
            "INSERT INTO students VALUES (?, ?, ?, ?, ?)",
            [
                (101, 20, "20240101", "学生一", "active"),
                (102, 20, "20240102", "学生二", "active"),
                (103, 20, "20240103", "学生三", "active"),
            ],
        )
        paper = {
            "grading": {"total_score": 100},
            "pages": [
                {"name": "第一部分", "questions": [{"id": "p1_q1", "type": "textarea", "text": "一", "answer": "A", "points": 30, "grading_guidance": "按步骤", "deduction_points": "缺步骤扣分"}]},
                {"name": "第二部分", "questions": [{"id": "p2_q1", "type": "textarea", "text": "二", "answer": "B", "points": 30, "grading_guidance": "按步骤", "deduction_points": "缺步骤扣分"}]},
                {"name": "第三部分", "questions": [{"id": "p3_q1", "type": "textarea", "text": "三", "answer": "C", "points": 40, "grading_guidance": "按步骤", "deduction_points": "缺步骤扣分"}]},
            ],
        }
        self.conn.execute(
            "INSERT INTO exam_papers VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("paper-1", 1, "服务器配置与管理期末机试", "", json.dumps(paper, ensure_ascii=False), "", "ready"),
        )
        self.conn.execute(
            "INSERT INTO assignments VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (301, "期末机试", "published", "paper-1", "2025-12-20", "2025-12-30", "ai", 30),
        )
        feedback_one = """
## 逐题反馈
### 第 p1_q1 题
- 本题得分：30/30
### 第 p2_q1 题
- 本题得分：30/30
### 第 p3_q1 题
- 本题得分：31/40
"""
        feedback_two = """
## 逐题反馈
### 第 p1_q1 题
- 本题得分：30/30
### 第 p2_q1 题
- 本题得分：30/30
### 第 p3_q1 题
- 本题得分：30/40
"""
        self.conn.executemany(
            "INSERT INTO submissions VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                (1, "301", 101, "学生一", "graded", 87, feedback_one, 91, 4, 1, 7200, 0),
                (2, "301", 102, "学生二", "graded", 82, feedback_two, None, 0, 0, 0, 0),
            ],
        )
        self.conn.execute(
            "INSERT INTO group_assignment_member_results VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (1, "301", 30, 1, 102, 2, 90, 10, 2, 82, 1, "2025-12-30"),
        )
        self.conn.commit()

    def test_candidates_and_payload_distribute_integer_deductions(self):
        candidates = list_exam_grade_record_candidates(self.conn, class_offering_id=30, teacher_id=1)
        self.assertEqual([item["id"] for item in candidates], [301])
        self.assertEqual(candidates[0]["section_count"], 3)
        self.assertEqual(candidates[0]["total_score"], 100)
        self.assertEqual(candidates[0]["graded_count"], 2)
        self.assertEqual(candidates[0]["roster_count"], 3)
        self.assertEqual(candidates[0]["missing_grade_count"], 1)
        self.assertEqual(candidates[0]["coverage_percent"], 66.7)
        self.assertTrue(candidates[0]["eligible"])
        self.assertEqual(candidates[0]["blocking_reason"], "")

        payload = build_exam_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            exam_assignment_id=301,
        )

        self.assertEqual(payload["document_type"], EXAM_GRADE_RECORD_TYPE)
        self.assertEqual(payload["structured"]["table_mode"], EXAM_GRADE_RECORD_TABLE_MODE)
        self.assertEqual(payload["structured"]["ordering_source"], "active_class_roster.student_number_then_id")
        self.assertEqual(len(payload["tables"]), 1)
        self.assertEqual(len(payload["tables"][0]["rows"]), 5)
        self.assertEqual([item["label"] for item in payload["structured"]["sections"]], ["一", "二", "三"])
        students = payload["structured"]["students"]
        self.assertEqual([item["row_order"] for item in students], [1, 2, 3])
        self.assertEqual(students[0]["raw_section_scores"], [30, 30, 31])
        self.assertEqual(students[0]["section_scores"], [29, 29, 29])
        self.assertEqual(students[0]["total_score"], 87)
        self.assertEqual(sum(students[0]["section_scores"]), students[0]["total_score"])
        self.assertIn("迟交扣 4 分", students[0]["score_adjustment_reason"])
        self.assertEqual(students[1]["total_score"], 82)
        self.assertEqual(sum(students[1]["section_scores"]), 82)
        self.assertIn("小组互评折算扣", students[1]["score_adjustment_reason"])
        self.assertTrue(any("学生三" in warning for warning in payload["structured"]["warnings"]))

    def test_roster_confirmed_retake_student_defaults_when_absent_from_exam(self):
        import classroom_app.db.schema_retake as schema_retake

        schema_retake._SCHEMA_READY = False
        schema_retake.ensure_retake_schema(self.conn)
        self.conn.execute(
            """
            INSERT INTO classroom_retake_students
                (class_offering_id, student_id, student_number, student_name, status,
                 default_ordinary_score, created_at, updated_at)
            VALUES (30, 103, '20240103', '学生三', 'confirmed', 70, '2026-07-30T10:00:00', '2026-07-30T10:00:00')
            """
        )
        payload = build_exam_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            exam_assignment_id=301,
        )
        students = payload["structured"]["students"]
        retake_row = next(item for item in students if item["student_number"] == "20240103")
        self.assertTrue(retake_row["is_retake"])
        self.assertEqual(retake_row["total_score"], 70)
        self.assertEqual(sum(retake_row["section_scores"]), 70)
        self.assertIn("默认分记录", retake_row["score_adjustment_reason"])
        # 已提交的学生仍按真实批改分数入库。
        real_row = next(item for item in students if item["student_number"] == "20240101")
        self.assertEqual(real_row["total_score"], 87)
        self.assertTrue(
            any("学生三" in w and "重修" in w for w in payload["structured"]["warnings"])
        )

    def test_generation_blocks_empty_roster_or_exam_without_graded_scores(self):
        self.conn.execute("DELETE FROM submissions")
        self.conn.execute("DELETE FROM group_assignment_member_results")
        self.conn.commit()

        candidates = list_exam_grade_record_candidates(self.conn, class_offering_id=30, teacher_id=1)
        self.assertFalse(candidates[0]["eligible"])
        self.assertIn("尚无已评分成绩", candidates[0]["blocking_reason"])
        with self.assertRaisesRegex(Exception, "尚无已评分成绩"):
            build_exam_grade_record_payload(
                self.conn,
                class_offering_id=30,
                teacher_id=1,
                exam_assignment_id=301,
            )

        self.conn.execute("DELETE FROM students")
        self.conn.commit()
        with self.assertRaisesRegex(Exception, "没有在读学生"):
            build_exam_grade_record_payload(
                self.conn,
                class_offering_id=30,
                teacher_id=1,
                exam_assignment_id=301,
            )

    def test_postgres_queries_keep_assignment_ids_native(self):
        class StrictPostgresLikeConnection:
            def __init__(self):
                self.sql = ""
                self.params = ()

            def execute(self, sql, params):
                self.sql = " ".join(str(sql).split())
                self.params = tuple(params)
                if "CAST(a.id AS TEXT)" in self.sql:
                    raise AssertionError("PostgreSQL bigint assignment ids must not be compared to text")
                return self

            def fetchall(self):
                return []

        candidates_conn = StrictPostgresLikeConnection()
        candidates = list_exam_grade_record_candidates(candidates_conn, class_offering_id=30, teacher_id=1)
        self.assertEqual(candidates, [])
        self.assertIn("LEFT JOIN submissions s ON s.assignment_id = a.id", candidates_conn.sql)
        self.assertEqual(candidates_conn.params, (30, 1))

        submissions_conn = StrictPostgresLikeConnection()
        submissions = _load_exam_submissions(submissions_conn, assignment_id=301)
        self.assertEqual(submissions, {})
        self.assertIn("gr.assignment_id = CAST(s.assignment_id AS TEXT)", submissions_conn.sql)
        self.assertIn("WHERE s.assignment_id = ?", submissions_conn.sql)
        self.assertEqual(submissions_conn.params, (301,))

    def test_payload_preview_keeps_every_student_in_student_number_order(self):
        extra_students = [
            (1000 + index, 20, f"2025{index:04d}", f"扩展学生{index}", "active")
            for index in range(125, 0, -1)
        ]
        self.conn.executemany("INSERT INTO students VALUES (?, ?, ?, ?, ?)", extra_students)
        self.conn.commit()

        payload = build_exam_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            exam_assignment_id=301,
        )
        students = payload["structured"]["students"]

        self.assertEqual(len(students), 128)
        self.assertEqual([item["student_number"] for item in students], sorted(item["student_number"] for item in students))
        self.assertEqual([item["row_order"] for item in students], list(range(1, 129)))
        self.assertIn("扩展学生125", payload["content_markdown"])
        self.assertEqual(len(payload["tables"]), 1)
        self.assertEqual(len(payload["tables"][0]["rows"]), 130)
        self.assertEqual(payload["queryable_fields"]["student_count"], 128)
        self.assertEqual(payload["queryable_fields"]["table_mode"], EXAM_GRADE_RECORD_TABLE_MODE)

    def test_xlsx_export_uses_one_continuous_roster_sheet_with_a4_headers_and_formulas(self):
        payload = build_exam_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            exam_assignment_id=301,
        )
        students = payload["structured"]["students"]
        for index in range(4, 27):
            clone = json.loads(json.dumps(students[0], ensure_ascii=False))
            clone["index"] = index
            clone["student_id"] = 1000 + index
            clone["student_number"] = f"2024{index:04d}"
            clone["student_name"] = f"学生{index}"
            students.append(clone)
        content = build_exam_grade_record_xlsx(payload)
        wb = load_workbook(io.BytesIO(content), data_only=False)
        ws = wb.active

        self.assertEqual(ws["A1"].value, "广西外国语学院机试（作品设计）考核登分表")
        self.assertIn("课程：服务器配置与管理", ws["A2"].value)
        self.assertIn("A1:G1", [str(item) for item in ws.merged_cells.ranges])
        self.assertIn("A2:G2", [str(item) for item in ws.merged_cells.ranges])
        self.assertEqual(ws["D3"].value, "一")
        self.assertEqual(ws["E3"].value, "二")
        self.assertEqual(ws["F3"].value, "三")
        self.assertEqual(ws["D4"].value, 30)
        self.assertEqual(ws["D4"].fill.fgColor.rgb, "FF92D050")
        self.assertEqual(ws["G4"].value, 100)
        self.assertEqual(ws["G5"].value, "=SUM(D5:F5)")
        self.assertEqual(ws["D5"].value, 29)
        self.assertEqual(ws["E5"].value, 29)
        self.assertEqual(ws["F5"].value, 29)
        self.assertEqual(str(ws.page_setup.paperSize), "9")
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        self.assertEqual(ws.page_setup.fitToHeight, 0)
        self.assertEqual(ws.freeze_panes, "D5")
        self.assertEqual(ws.print_title_rows, "$1:$4")
        self.assertTrue(ws.print_options.horizontalCentered)
        self.assertEqual(ws.oddFooter.center.text, "第 &P 页 / 共 &N 页")
        self.assertEqual([item.id for item in ws.row_breaks.brk], [])
        self.assertEqual(len(ws.data_validations.dataValidation), 3)
        self.assertEqual(ws["A1"].font.name, "宋体")
        self.assertEqual(ws["A1"].font.charset, 134)
        self.assertLessEqual(ws.column_dimensions["D"].width, 16)
        self.assertEqual(wb.sheetnames, ["考核登分表"])
        self.assertEqual(ws.max_row, 30)
        self.assertEqual([ws.cell(row, 1).value for row in range(5, 31)], list(range(1, 27)))
        self.assertEqual(ws.print_area, "'考核登分表'!$A$1:$G$30")

    def test_parser_ai_import_and_export_artifact_force_xlsx(self):
        payload = build_exam_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            exam_assignment_id=301,
        )
        content = build_exam_grade_record_xlsx(payload)
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir) / ("a" * 64)
            temp_path.write_bytes(content)
            parsed = parse_exam_grade_record_file(temp_path, "考核登分表.xlsx")
            self.assertEqual(parsed.formula_count, 2)
            self.assertEqual(len(parsed.export_payload["structured"]["sections"]), 3)
            self.assertEqual(parsed.export_payload["structured"]["students"][0]["total_score"], 87)
            self.assertEqual(len(parsed.tables), 1)
            self.assertEqual([row[0] for row in parsed.tables[0]["rows"][2:]], [1, 2, 3])

            async def fail_ai_chat(*args, **kwargs):  # pragma: no cover - must not be called
                raise AssertionError("exam grade import must use the local Excel parser")

            result = asyncio.run(
                parse_material_document(
                    file_path=temp_path,
                    original_name="考核登分表.xlsx",
                    document_group="final_material",
                    document_type=EXAM_GRADE_RECORD_TYPE,
                    ai_chat=fail_ai_chat,
                )
            )
            self.assertFalse(result.ai_used)
            self.assertEqual(result.document_type, EXAM_GRADE_RECORD_TYPE)
            self.assertEqual(result.extraction_method, "exam_grade_excel_formula_parser")

            artifact = build_material_export_artifact(payload, fallback_filename="exam-grade", requested_format="docx")
            self.assertEqual(artifact.media_type, XLSX_MEDIA_TYPE)
            self.assertTrue(artifact.filename.endswith(".xlsx"))
            self.assertGreater(len(artifact.content), 5000)

    def test_parser_matches_real_source_shape_and_audits_score_integrity(self):
        workbook = Workbook()
        cover = workbook.active
        cover.title = "说明"
        cover["A1"] = "请勿删除本说明页"
        worksheet = workbook.create_sheet("考核登分表")
        worksheet.merge_cells("A1:J1")
        worksheet["A1"] = "广西外国语学院期末考试考核登分表"
        worksheet.merge_cells("A2:J2")
        worksheet["A2"] = (
            "课程：计算机网络    专业年级班级：软工2302班\n"
            "授课老师：张海林    学年学期：2025-2026学年第二学期"
        )
        worksheet.append(["序号", "学号", "姓名", "一", "二", "三", "四", "五", "六", "总分"])
        worksheet.append(["", "", "", 10, 10, 10, 20, 20, 30, 100])
        expected_total = 0
        for index in range(1, 50):
            scores = [8, 7, 9, 14, 18, 24]
            row = index + 4
            expected_total += sum(scores)
            worksheet.append(
                [
                    index,
                    f"2305301{index:04d}",
                    f"学生{index}",
                    *scores,
                    f'=IF(COUNT(D{row}:I{row})=0,"",SUM(D{row}:I{row}))',
                ]
            )
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ("b" * 64)
            path.write_bytes(output.getvalue())
            parsed = parse_exam_grade_record_file(
                path,
                "2025-2026-2《计算机网络》期末考试考核登分表-软工2302班.xlsx",
            )

        fields = parsed.export_payload["fields"]
        structured = parsed.export_payload["structured"]
        self.assertEqual(parsed.metadata["source_sheet"], "考核登分表")
        self.assertEqual(parsed.metadata["academic_year"], "2025-2026")
        self.assertEqual(parsed.metadata["semester"], "第二学期")
        self.assertEqual(fields["course_name"], "计算机网络")
        self.assertEqual(fields["class_name"], "软工2302班")
        self.assertEqual(fields["teacher_name"], "张海林")
        self.assertEqual(fields["total_score"], 100)
        self.assertEqual(len(structured["sections"]), 6)
        self.assertEqual(len(structured["students"]), 49)
        self.assertEqual(parsed.formula_count, 49)
        self.assertEqual(
            sum(int(item["total_score"]) for item in structured["students"]),
            expected_total,
        )
        self.assertEqual(parsed.warnings, [])

    def test_parser_surfaces_duplicate_student_formula_and_score_anomalies(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "考核登分表"
        worksheet["A1"] = "广西外国语学院期末考试考核登分表"
        worksheet["A2"] = (
            "课程：计算机网络    专业年级班级：软工2302班\n"
            "授课老师：张海林    学年学期：2025-2026学年第二学期"
        )
        worksheet.append(["序号", "学号", "姓名", "一", "二", "总分"])
        worksheet.append(["", "", "", 40, 60, 100])
        worksheet.append([1, "23050001", "学生一", 45, 50, "=SUM(D5:E5)"])
        worksheet.append([2, "23050001", "学生二", 30, 40, "=SUM(D5:E5)"])
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / ("c" * 64)
            path.write_bytes(output.getvalue())
            parsed = parse_exam_grade_record_file(path, "考核登分表.xlsx")

        warning_text = "\n".join(parsed.warnings)
        self.assertIn("超出 0 至 40 分", warning_text)
        self.assertIn("学号 23050001", warning_text)
        self.assertIn("总分公式未覆盖全部大题列", warning_text)


if __name__ == "__main__":
    unittest.main()
