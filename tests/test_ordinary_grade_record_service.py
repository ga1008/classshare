import asyncio
import io
import sqlite3
import tempfile
import unittest
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook

from classroom_app.services.material_ai_import_service import parse_material_document
from classroom_app.services.material_export_template_service import (
    XLSX_MEDIA_TYPE,
    build_material_export_artifact,
)
from classroom_app.services.ordinary_grade_record_service import (
    ORDINARY_GRADE_RECORD_TYPE,
    apply_ordinary_grade_score_floor,
    build_ordinary_grade_record_export_filename,
    build_ordinary_grade_record_payload,
    build_ordinary_grade_record_xlsx,
    calculate_ordinary_grade_score,
    classify_ordinary_grade_assignment,
    distribute_retake_ordinary_score,
    normalize_retake_students,
    list_ordinary_grade_assignment_candidates,
    normalize_ordinary_grade_kind_override,
    normalize_ordinary_grade_record_payload,
    parse_ordinary_grade_record_file,
    validate_ordinary_grade_sources,
)


class OrdinaryGradeRecordServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self._create_schema()
        self._seed_data()

    def tearDown(self) -> None:
        self.conn.close()

    def _create_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE teachers (
                id INTEGER PRIMARY KEY,
                name TEXT,
                college TEXT,
                department TEXT
            );
            CREATE TABLE courses (
                id INTEGER PRIMARY KEY,
                name TEXT,
                total_hours INTEGER,
                credits REAL,
                college TEXT,
                department TEXT
            );
            CREATE TABLE classes (
                id INTEGER PRIMARY KEY,
                name TEXT,
                college TEXT,
                department TEXT
            );
            CREATE TABLE class_offerings (
                id INTEGER PRIMARY KEY,
                class_id INTEGER,
                course_id INTEGER,
                teacher_id INTEGER,
                semester TEXT
            );
            CREATE TABLE students (
                id INTEGER PRIMARY KEY,
                class_id INTEGER,
                student_id_number TEXT,
                name TEXT,
                enrollment_status TEXT
            );
            CREATE TABLE assignments (
                id INTEGER PRIMARY KEY,
                title TEXT,
                status TEXT,
                exam_paper_id TEXT,
                created_at TEXT,
                due_at TEXT,
                grading_mode TEXT,
                class_offering_id INTEGER,
                ordinary_grade_kind_override TEXT,
                ordinary_grade_kind_updated_at TEXT,
                ordinary_grade_kind_updated_by_teacher_id INTEGER
            );
            CREATE TABLE submissions (
                id INTEGER PRIMARY KEY,
                assignment_id TEXT,
                student_pk_id INTEGER,
                score REAL
            );
            CREATE TABLE smart_classroom_checkin_sessions (
                id INTEGER PRIMARY KEY,
                teacher_id INTEGER,
                class_offering_id INTEGER,
                session_id INTEGER,
                checkin_time TEXT,
                synced_at TEXT
            );
            CREATE TABLE smart_classroom_checkin_students (
                id INTEGER PRIMARY KEY,
                checkin_session_id INTEGER,
                student_id INTEGER,
                status TEXT
            );
            """
        )

    def _seed_data(self) -> None:
        self.conn.execute("INSERT INTO teachers VALUES (1, '张海林', '数字科技学院', '软件工程系')")
        self.conn.execute("INSERT INTO courses VALUES (10, '服务器配置与管理', 48, 3.0, '数字科技学院', '软件工程系')")
        self.conn.execute("INSERT INTO classes VALUES (20, '软工2406班（专升本）', '数字科技学院', '软件工程系')")
        self.conn.execute("INSERT INTO class_offerings VALUES (30, 20, 10, 1, '2025-2026-1')")
        students = [
            (101, 20, "20240101", "学生一", "active"),
            (102, 20, "20240102", "学生二", "active"),
            (103, 20, "20240103", "学生三", "active"),
        ]
        self.conn.executemany("INSERT INTO students VALUES (?, ?, ?, ?, ?)", students)
        assignments = [
            (201, "第一次作业", "published", "", "2025-09-01", "2025-09-10", "manual", 30, None, None, None),
            (202, "第二次作业", "published", "", "2025-09-11", "2025-09-20", "manual", 30, None, None, None),
            (203, "第三次作业", "published", "", "2025-09-21", "2025-09-30", "manual", 30, None, None, None),
            (204, "阶段测评", "published", "9", "2025-10-01", "2025-10-10", "manual", 30, None, None, None),
        ]
        self.conn.executemany("INSERT INTO assignments VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", assignments)
        submissions = [
            (1, "201", 101, 91),
            (2, "202", 101, 92),
            (3, "203", 101, 93),
            (4, "204", 101, 88),
            (5, "201", 102, 81),
            (6, "202", 102, 82),
            (7, "203", 102, 83),
            (8, "204", 102, 78),
            (9, "201", 103, 71),
        ]
        self.conn.executemany("INSERT INTO submissions VALUES (?, ?, ?, ?)", submissions)
        self.conn.executemany(
            "INSERT INTO smart_classroom_checkin_sessions VALUES (?, ?, ?, ?, ?, ?)",
            [
                (301, 1, 30, 1, "2025-09-01 08:00:00", "2025-09-01 09:00:00"),
                (302, 1, 30, 2, "2025-09-08 08:00:00", "2025-09-08 09:00:00"),
            ],
        )
        self.conn.executemany(
            "INSERT INTO smart_classroom_checkin_students VALUES (?, ?, ?, ?)",
            [
                (401, 301, 101, "CHECKED"),
                (402, 301, 102, "CHECKED"),
                (403, 302, 101, "CHECKED"),
                (404, 302, 102, "UNCHECKED"),
            ],
        )
        self.conn.commit()

    def test_retake_distribution_hits_target_exactly_across_full_range(self):
        for tenth in range(0, 1001, 5):
            target = tenth / 10.0
            result = distribute_retake_ordinary_score(target, seed_parts=(30, 101, 201, 202, 203, 204))
            achieved = calculate_ordinary_grade_score(
                result["attendance_score"],
                result["homework_scores"],
                result["assessment_score"],
            )
            self.assertAlmostEqual(achieved, round(target, 2), places=6, msg=f"target={target}")
            for value in (result["attendance_score"], result["assessment_score"], *result["homework_scores"]):
                self.assertGreaterEqual(value, 0.0, msg=f"target={target}")
                self.assertLessEqual(value, 100.0, msg=f"target={target}")
        # Deterministic: same inputs → identical distribution.
        first = distribute_retake_ordinary_score(60, seed_parts=(30, 101, 201, 202, 203, 204))
        second = distribute_retake_ordinary_score(60, seed_parts=(30, 101, 201, 202, 203, 204))
        self.assertEqual(first, second)

    def test_retake_students_receive_exact_teacher_set_score(self):
        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 203],
            assessment_assignment_id=204,
            retake_students=[{"student_number": "20240103", "ordinary_score": 72}],
        )
        students = payload["structured"]["students"]
        retake_row = next(row for row in students if row["student_number"] == "20240103")
        self.assertTrue(retake_row["is_retake"])
        self.assertEqual(retake_row["calculated_scores"]["ordinary_score"], 72)
        achieved = calculate_ordinary_grade_score(
            retake_row["attendance_raw_score"],
            retake_row["homework_scores"],
            retake_row["assessment_score"],
        )
        self.assertAlmostEqual(achieved, 72.0, places=6)
        self.assertEqual(retake_row["score_floor_adjustment"]["reason"], "retake_override")
        # 学生三 normally triggers missing-score warnings; as a retake student
        # the only warning about them must be the teacher-set score notice.
        warnings = payload["structured"]["warnings"]
        self.assertFalse(any("学生三" in warning and "按 0 分计入" in warning for warning in warnings))
        self.assertTrue(any("学生三" in warning and "重修" in warning for warning in warnings))
        policy = payload["structured"]["retake_policy"]
        self.assertEqual(policy["count"], 1)
        self.assertEqual(policy["students"][0]["student_number"], "20240103")
        self.assertEqual(policy["students"][0]["target_score"], 72)
        self.assertEqual(payload["fields"]["retake_student_count"], 1)
        # Other students keep the normal pipeline.
        normal_row = next(row for row in students if row["student_number"] == "20240101")
        self.assertNotIn("is_retake", normal_row)

    def test_retake_validation_rejects_unknown_and_invalid_entries(self):
        with self.assertRaises(HTTPException) as unknown:
            build_ordinary_grade_record_payload(
                self.conn,
                class_offering_id=30,
                teacher_id=1,
                homework_assignment_ids=[201, 202, 203],
                assessment_assignment_id=204,
                retake_students=[{"student_number": "99999999", "ordinary_score": 60}],
            )
        self.assertEqual(unknown.exception.status_code, 400)
        self.assertIn("不在本课堂名单", unknown.exception.detail)
        with self.assertRaises(HTTPException):
            normalize_retake_students([{"student_number": "20240101", "ordinary_score": 101}])
        with self.assertRaises(HTTPException):
            normalize_retake_students([{"student_number": "", "ordinary_score": 60}])
        with self.assertRaises(HTTPException):
            normalize_retake_students([
                {"student_number": "20240101", "ordinary_score": 60},
                {"student_number": "20240101", "ordinary_score": 70},
            ])

    def test_source_validation_rejects_wrong_count_and_overlap(self):
        with self.assertRaises(HTTPException):
            validate_ordinary_grade_sources(homework_assignment_ids=[201, 202], assessment_assignment_id=204)
        with self.assertRaises(HTTPException):
            validate_ordinary_grade_sources(homework_assignment_ids=[201, 202, 203], assessment_assignment_id=202)

    def test_candidates_and_payload_use_classroom_scores(self):
        candidates = list_ordinary_grade_assignment_candidates(self.conn, class_offering_id=30, teacher_id=1)
        self.assertEqual([item["id"] for item in candidates], [201, 202, 203, 204])
        self.assertEqual(candidates[0]["graded_count"], 3)
        self.assertEqual(candidates[3]["kind"], "exam")

        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 203],
            assessment_assignment_id=204,
            attendance_sync={"status": "cached", "cache_hit": True, "synced_at": "2026-07-28T10:00:00"},
            generation_requirements="课程组归档前复核异常分数。",
        )

        self.assertEqual(payload["document_type"], ORDINARY_GRADE_RECORD_TYPE)
        self.assertEqual(payload["fields"]["course_name"], "服务器配置与管理")
        self.assertEqual(payload["fields"]["class_size"], 3)
        students = payload["structured"]["students"]
        self.assertEqual(students[0]["attendance_raw_score"], 100.0)
        self.assertEqual(students[1]["attendance_raw_score"], 50.0)
        self.assertEqual(students[0]["homework_scores"], [91.0, 92.0, 93.0])
        self.assertEqual(students[0]["assessment_score"], 88.0)
        self.assertEqual(students[2]["attendance_raw_score"], 0.0)
        self.assertEqual(students[2]["homework_scores"], [71.0, 0.0, 0.0])
        self.assertEqual(students[2]["assessment_score"], 0.0)
        self.assertEqual(payload["structured"]["attendance_sync"]["status"], "cached")
        self.assertEqual(payload["structured"]["generation_requirements"], "课程组归档前复核异常分数。")
        self.assertTrue(any("学生三" in warning for warning in payload["structured"]["warnings"]))
        self.assertTrue(any("按 0 分计入" in warning for warning in payload["structured"]["warnings"]))

    def test_candidate_query_keeps_assignment_ids_native_for_postgres(self):
        class StrictPostgresLikeConnection:
            def __init__(self):
                self.sql = ""
                self.params = ()

            def execute(self, sql, params):
                self.sql = " ".join(str(sql).split())
                self.params = tuple(params)
                if "CAST(a.id AS TEXT)" in self.sql:
                    raise AssertionError("PostgreSQL bigint assignment ids must not be compared to text")
                return self

            def fetchall(self):
                return []

        conn = StrictPostgresLikeConnection()
        candidates = list_ordinary_grade_assignment_candidates(conn, class_offering_id=30, teacher_id=1)
        self.assertEqual(candidates, [])
        self.assertIn("LEFT JOIN submissions s ON s.assignment_id = a.id", conn.sql)
        self.assertEqual(conn.params, (30, 1))

    def test_linked_exam_paper_can_still_be_a_homework_by_classroom_purpose(self):
        self.assertEqual(
            classify_ordinary_grade_assignment(
                {"title": "动态 Web 作业 2 - 第十讲实战", "exam_paper_id": "paper-2"}
            ),
            "assignment",
        )
        self.assertEqual(
            classify_ordinary_grade_assignment(
                {"title": "期末综合实验验收", "exam_paper_id": "paper-final"}
            ),
            "exam",
        )
        self.assertEqual(
            classify_ordinary_grade_assignment(
                {"title": "阶段测评", "exam_paper_id": None}
            ),
            "exam",
        )
        self.assertEqual(
            classify_ordinary_grade_assignment(
                {
                    "title": "阶段测评",
                    "exam_paper_id": "paper-stage",
                    "ordinary_grade_kind_override": "assignment",
                }
            ),
            "assignment",
        )

    def test_manual_kind_override_is_visible_and_can_satisfy_three_homework_sources(self):
        self.conn.execute(
            """
            UPDATE assignments
            SET ordinary_grade_kind_override = 'exam'
            WHERE id = 203
            """
        )
        self.conn.execute(
            """
            UPDATE assignments
            SET ordinary_grade_kind_override = 'assignment',
                ordinary_grade_kind_updated_at = '2026-07-29T10:00:00',
                ordinary_grade_kind_updated_by_teacher_id = 1
            WHERE id = 204
            """
        )
        candidates = list_ordinary_grade_assignment_candidates(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )
        changed = next(item for item in candidates if item["id"] == 204)
        self.assertEqual("assignment", changed["kind"])
        self.assertEqual("exam", changed["ordinary_grade_auto_kind"])
        self.assertEqual("manual", changed["ordinary_grade_kind_source"])
        self.assertEqual(1, changed["ordinary_grade_kind_updated_by_teacher_id"])
        self.assertEqual(3, sum(item["kind"] == "assignment" for item in candidates))
        self.assertEqual(1, sum(item["kind"] == "exam" for item in candidates))

        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 204],
            assessment_assignment_id=203,
        )
        self.assertEqual([201, 202, 204], [
            item["id"] for item in payload["structured"]["source_assignments"]["homework_assignments"]
        ])
        self.assertEqual(
            203,
            payload["structured"]["source_assignments"]["assessment_assignment"]["id"],
        )

        self.assertEqual("", normalize_ordinary_grade_kind_override("auto"))
        self.assertEqual("assignment", normalize_ordinary_grade_kind_override("assignment"))
        with self.assertRaises(ValueError):
            normalize_ordinary_grade_kind_override("quiz")

    def test_generation_revalidates_effective_kind_after_selection(self):
        self.conn.execute(
            "UPDATE assignments SET ordinary_grade_kind_override = 'exam' WHERE id = 203"
        )
        with self.assertRaises(HTTPException) as ctx:
            build_ordinary_grade_record_payload(
                self.conn,
                class_offering_id=30,
                teacher_id=1,
                homework_assignment_ids=[201, 202, 203],
                assessment_assignment_id=204,
            )
        self.assertEqual(400, ctx.exception.status_code)
        self.assertIn("不能放入平时作业", str(ctx.exception.detail))

    def test_score_floor_is_deterministic_balanced_and_requires_seventy_percent_attendance(self):
        seed_parts = (30, 103, 201, 202, 203, 204)
        eligible = apply_ordinary_grade_score_floor(
            attendance_score=70,
            homework_scores=[0, 0, 0],
            assessment_score=0,
            enabled=True,
            minimum_score=60,
            seed_parts=seed_parts,
        )
        repeated = apply_ordinary_grade_score_floor(
            attendance_score=70,
            homework_scores=[0, 0, 0],
            assessment_score=0,
            enabled=True,
            minimum_score=60,
            seed_parts=seed_parts,
        )
        self.assertEqual(eligible, repeated)
        self.assertTrue(eligible["eligible"])
        self.assertTrue(eligible["applied"])
        self.assertGreaterEqual(eligible["achieved_score"], 60)
        self.assertLess(eligible["achieved_score"], 60.3)
        adjusted_values = [*eligible["homework_scores"], eligible["assessment_score"]]
        self.assertLessEqual(max(adjusted_values) - min(adjusted_values), 10)
        self.assertEqual(
            eligible["achieved_score"],
            round(calculate_ordinary_grade_score(70, eligible["homework_scores"], eligible["assessment_score"]), 4),
        )

        ineligible = apply_ordinary_grade_score_floor(
            attendance_score=69.99,
            homework_scores=[0, 0, 0],
            assessment_score=0,
            enabled=True,
            minimum_score=60,
            seed_parts=seed_parts,
        )
        self.assertFalse(ineligible["eligible"])
        self.assertFalse(ineligible["applied"])
        self.assertEqual([0.0, 0.0, 0.0], ineligible["homework_scores"])
        self.assertEqual(0.0, ineligible["assessment_score"])
        self.assertEqual("attendance_below_threshold", ineligible["reason"])

    def test_score_floor_never_reduces_real_scores_and_reports_unreachable_target(self):
        already_high = apply_ordinary_grade_score_floor(
            attendance_score=100,
            homework_scores=[90, 80, 95],
            assessment_score=88,
            enabled=True,
            minimum_score=60,
            seed_parts=(30, 101),
        )
        self.assertFalse(already_high["applied"])
        self.assertEqual([90.0, 80.0, 95.0], already_high["homework_scores"])
        self.assertEqual(88.0, already_high["assessment_score"])

        capped = apply_ordinary_grade_score_floor(
            attendance_score=70,
            homework_scores=[0, 0, 0],
            assessment_score=0,
            enabled=True,
            minimum_score=100,
            seed_parts=(30, 102),
        )
        self.assertTrue(capped["capped"])
        self.assertEqual("capped_by_attendance", capped["reason"])
        self.assertEqual(88.0, capped["achieved_score"])
        self.assertEqual([100.0, 100.0, 100.0], capped["homework_scores"])
        self.assertEqual(100.0, capped["assessment_score"])

    def test_attendance_denominator_includes_sessions_missing_a_student_row(self):
        self.conn.execute(
            "INSERT INTO smart_classroom_checkin_students VALUES (?, ?, ?, ?)",
            (405, 301, 103, "CHECKED"),
        )
        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 203],
            assessment_assignment_id=204,
        )
        student = payload["structured"]["students"][2]
        self.assertEqual(50.0, student["attendance_raw_score"])
        self.assertFalse(student["score_floor_adjustment"]["eligible"])
        self.assertFalse(student["score_floor_adjustment"]["applied"])

    def test_payload_adjusts_only_eligible_task_scores_and_exports_hidden_audit(self):
        self.conn.executemany(
            "INSERT INTO smart_classroom_checkin_students VALUES (?, ?, ?, ?)",
            [
                (405, 301, 103, "CHECKED"),
                (406, 302, 103, "CHECKED"),
            ],
        )
        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 203],
            assessment_assignment_id=204,
            minimum_ordinary_score_enabled=True,
            minimum_ordinary_score=60,
        )
        student = payload["structured"]["students"][2]
        self.assertEqual(100.0, student["attendance_raw_score"])
        self.assertEqual([71.0, 0.0, 0.0], student["source_homework_scores"])
        self.assertEqual(0.0, student["source_assessment_score"])
        self.assertTrue(student["score_floor_adjustment"]["applied"])
        self.assertGreaterEqual(student["score_floor_adjustment"]["achieved_score"], 60)
        policy = payload["structured"]["score_floor_policy"]
        self.assertEqual(2, policy["eligible_count"])
        self.assertEqual(1, policy["adjusted_count"])
        self.assertEqual(1, policy["ineligible_count"])

        content = build_ordinary_grade_record_xlsx(payload)
        wb = load_workbook(io.BytesIO(content), data_only=False)
        self.assertEqual("auto", wb.calculation.calcMode)
        self.assertTrue(wb.calculation.fullCalcOnLoad)
        self.assertTrue(wb.calculation.forceFullCalc)
        self.assertIn("最低分配平审计", wb.sheetnames)
        audit = wb["最低分配平审计"]
        self.assertEqual("hidden", audit.sheet_state)
        self.assertEqual("原始缺失项", audit["F1"].value)
        self.assertEqual("作业2、作业3、测评", audit["F4"].value)
        self.assertEqual(71.0, audit["G4"].value)
        self.assertEqual(0.0, audit["H4"].value)
        self.assertGreaterEqual(audit["P4"].value, 60)
        self.assertEqual("=I9*0.4+J9*0.3+K9*0.3", wb.active["L9"].value)

    def test_disabled_score_floor_preserves_zero_filled_source_scores(self):
        self.conn.executemany(
            "INSERT INTO smart_classroom_checkin_students VALUES (?, ?, ?, ?)",
            [
                (405, 301, 103, "CHECKED"),
                (406, 302, 103, "CHECKED"),
            ],
        )
        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 203],
            assessment_assignment_id=204,
            minimum_ordinary_score_enabled=False,
            minimum_ordinary_score=60,
        )
        student = payload["structured"]["students"][2]
        self.assertEqual([71.0, 0.0, 0.0], student["homework_scores"])
        self.assertEqual(0.0, student["assessment_score"])
        self.assertEqual("disabled", student["score_floor_adjustment"]["reason"])

    def test_xlsx_export_preserves_pages_headers_notes_and_formulas(self):
        students = []
        for index in range(1, 46):
            students.append(
                {
                    "index": index,
                    "student_number": f"2024{index:04d}",
                    "student_name": f"学生{index}",
                    "attendance_raw_score": 100,
                    "homework_scores": [80 + index % 5, 81 + index % 5, 82 + index % 5],
                    "assessment_score": 90,
                }
            )
        payload = normalize_ordinary_grade_record_payload(
            metadata={
                "college": "数字科技学院",
                "course_name": "服务器配置与管理",
                "course_hours": 48,
                "credits": 3.0,
                "teacher_name": "张海林",
                "class_name": "软工2406班（专升本）",
                "academic_year": "2025-2026",
                "semester": "第一学期",
                "class_size": 45,
            },
            content_markdown="",
            tables=[],
            export_payload={"structured": {"students": students}},
        )
        content = build_ordinary_grade_record_xlsx(payload)
        wb = load_workbook(io.BytesIO(content), data_only=False)
        ws = wb.active

        self.assertEqual(ws["A1"].value, "广西外国语学院学生平时成绩记录表")
        self.assertEqual(ws["A38"].value, "广西外国语学院学生平时成绩记录表")
        self.assertEqual(wb._fonts[0].name, "宋体")
        self.assertEqual(wb._fonts[0].sz, 12)
        self.assertIn("A1:L1", [str(item) for item in ws.merged_cells.ranges])
        self.assertEqual(ws["I7"].value, "=D7")
        self.assertEqual(ws["J7"].value, "=AVERAGE(E7:G7)")
        self.assertEqual(ws["K7"].value, "=H7")
        self.assertEqual(ws["L7"].value, "=I7*0.4+J7*0.3+K7*0.3")
        self.assertEqual(ws["H7"].number_format, '0_);[RED]\\(0\\)')
        self.assertIsNone(getattr(ws["A4"].border.top, "style", None))
        self.assertEqual(ws["A4"].border.bottom.style, "thin")
        self.assertEqual(ws["I44"].value, "=D44")
        self.assertIn("该表可为电子表格", str(ws["A69"].value))
        self.assertEqual(str(ws.page_setup.paperSize), "9")
        self.assertEqual(ws.page_setup.orientation, "portrait")
        self.assertEqual(ws.page_setup.scale, 100)
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        self.assertEqual(ws.page_setup.fitToHeight, 1)
        self.assertFalse(ws.sheet_properties.pageSetUpPr.fitToPage)
        self.assertAlmostEqual(ws.page_margins.left, 0.3541666667)
        self.assertAlmostEqual(ws.page_margins.right, 0.1576388889)
        self.assertAlmostEqual(ws.page_margins.top, 0.1965277778)
        self.assertAlmostEqual(ws.page_margins.bottom, 0.0)
        self.assertAlmostEqual(ws.page_margins.footer, 0.1181102362)
        self.assertEqual(ws.column_dimensions["E"].width, 5.49)
        self.assertEqual(ws.column_dimensions["F"].width, 5.49)
        self.assertEqual(ws.column_dimensions["G"].width, 5.49)
        self.assertEqual(ws.column_dimensions["H"].width, 5.49)
        self.assertEqual(ws.row_dimensions[36].height, 53.5)
        self.assertEqual(ws.row_dimensions[37].height, 40.0)
        self.assertEqual(len(ws.row_breaks.brk), 1)
        self.assertEqual(ws.row_breaks.brk[0].id, 36)
        self.assertEqual(sum(1 for row in ws.iter_rows() for cell in row if cell.data_type == "f"), 45 * 4)

    def test_export_filename_uses_period_course_class_and_xlsx_format(self):
        filename = build_ordinary_grade_record_export_filename(
            {
                "academic_year": "2025-2026",
                "semester": "第一学期",
                "course_name": "服务器配置与管理",
                "class_name": "软工2406班（专升本）",
            }
        )
        self.assertEqual(
            filename,
            "7. 2025-2026-1《服务器配置与管理》学生平时成绩记录表-软工2406班（专升本）.xlsx",
        )
        self.assertNotIn(".xls.xlsx", filename)

    def test_export_filename_does_not_mistake_unrelated_digits_for_semester(self):
        filename = build_ordinary_grade_record_export_filename(
            {
                "semester": "P03-2026",
                "course_name": "服务器配置与管理",
                "class_name": "软工2406班",
            }
        )
        self.assertEqual(
            filename,
            "7. 未设置学年《服务器配置与管理》学生平时成绩记录表-软工2406班.xlsx",
        )

    def test_parser_and_ai_import_path_recognize_excel_formulas_without_ai(self):
        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 203],
            assessment_assignment_id=204,
        )
        content = build_ordinary_grade_record_xlsx(payload)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp:
            temp.write(content)
            temp_path = Path(temp.name)
        try:
            parsed = parse_ordinary_grade_record_file(temp_path, "平时成绩记录表.xlsx")
            self.assertEqual(parsed.formula_count, 12)
            self.assertEqual(len(parsed.export_payload["structured"]["students"]), 3)

            async def fail_ai_chat(*args, **kwargs):  # pragma: no cover - must not be called
                raise AssertionError("ordinary grade import must use the local Excel parser")

            result = asyncio.run(
                parse_material_document(
                    file_path=temp_path,
                    original_name="平时成绩记录表.xlsx",
                    document_group="final_material",
                    document_type=ORDINARY_GRADE_RECORD_TYPE,
                    ai_chat=fail_ai_chat,
                )
            )
            self.assertFalse(result.ai_used)
            self.assertEqual(result.document_type, ORDINARY_GRADE_RECORD_TYPE)
            self.assertEqual(result.extraction_method, "ordinary_grade_excel_formula_parser")
        finally:
            temp_path.unlink(missing_ok=True)

    def test_export_artifact_forces_xlsx_even_if_docx_requested(self):
        payload = build_ordinary_grade_record_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            homework_assignment_ids=[201, 202, 203],
            assessment_assignment_id=204,
        )
        artifact = build_material_export_artifact(payload, fallback_filename="ordinary", requested_format="docx")
        self.assertEqual(artifact.media_type, XLSX_MEDIA_TYPE)
        self.assertEqual(
            artifact.filename,
            "7. 2025-2026-1《服务器配置与管理》学生平时成绩记录表-软工2406班（专升本）.xlsx",
        )
        self.assertGreater(len(artifact.content), 6000)


if __name__ == "__main__":
    unittest.main()
