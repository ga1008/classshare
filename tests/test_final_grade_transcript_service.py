import io
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook

from classroom_app.services.final_grade_transcript_service import (
    FINAL_GRADE_TRANSCRIPT_TYPE,
    build_final_grade_transcript_payload,
    build_final_grade_transcript_readiness,
    build_final_grade_transcript_xlsx,
    normalize_final_grade_semester,
    parse_final_grade_transcript_file,
)
from classroom_app.services.material_export_template_service import (
    XLSX_MEDIA_TYPE,
    build_material_export_artifact,
)


class FinalGradeTranscriptServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row
        self._create_schema()
        self._seed_data()

    def tearDown(self) -> None:
        self.conn.close()

    def _create_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE teachers (
                id INTEGER PRIMARY KEY,
                name TEXT,
                school_name TEXT,
                college TEXT,
                department TEXT
            );
            CREATE TABLE courses (
                id INTEGER PRIMARY KEY,
                name TEXT,
                academic_course_code TEXT,
                school_name TEXT,
                college TEXT,
                department TEXT
            );
            CREATE TABLE classes (
                id INTEGER PRIMARY KEY,
                name TEXT,
                school_name TEXT,
                college TEXT,
                department TEXT
            );
            CREATE TABLE class_offerings (
                id INTEGER PRIMARY KEY,
                class_id INTEGER,
                course_id INTEGER,
                teacher_id INTEGER,
                semester_id INTEGER,
                semester TEXT
            );
            CREATE TABLE teacher_academic_course_sync_items (
                id INTEGER PRIMARY KEY,
                teacher_id INTEGER,
                semester_id INTEGER,
                course_id INTEGER,
                course_name TEXT,
                course_nature TEXT,
                academic_year TEXT,
                academic_year_name TEXT,
                academic_term TEXT,
                academic_term_name TEXT,
                synced_at TEXT
            );
            CREATE TABLE teacher_academic_exam_roster_items (
                id INTEGER PRIMARY KEY,
                teacher_id INTEGER,
                class_offering_id INTEGER,
                sync_status TEXT,
                synced_at TEXT,
                course_name TEXT,
                exam_course_key TEXT
            );
            CREATE TABLE teacher_academic_exam_roster_students (
                id INTEGER PRIMARY KEY,
                exam_roster_item_id INTEGER,
                row_order INTEGER,
                student_number TEXT,
                student_name TEXT,
                admin_class_name TEXT
            );
            CREATE TABLE material_ai_import_records (
                id INTEGER PRIMARY KEY,
                teacher_id INTEGER,
                document_group TEXT,
                document_type TEXT,
                document_type_label TEXT,
                parse_status TEXT,
                package_material_id INTEGER,
                parsed_material_id INTEGER,
                source_material_id INTEGER,
                export_payload_json TEXT,
                updated_at TEXT
            );
            CREATE TABLE course_material_assignments (
                material_id INTEGER,
                class_offering_id INTEGER
            );
            """
        )

    def _seed_data(self) -> None:
        self.conn.execute(
            "INSERT INTO teachers VALUES (1, '张海林', '广西外国语学院', '信息工程学院', '软件工程系')"
        )
        self.conn.execute(
            "INSERT INTO courses VALUES (10, '服务器配置与管理', 'E020141B4', '广西外国语学院', '信息工程学院', '软件工程系')"
        )
        self.conn.execute(
            "INSERT INTO classes VALUES (20, '软工2406班（专升本）', '广西外国语学院', '信息工程学院', '软件工程系')"
        )
        self.conn.execute(
            "INSERT INTO class_offerings VALUES (30, 20, 10, 1, 40, '2025-2026-1')"
        )
        self.conn.execute(
            """
            INSERT INTO teacher_academic_course_sync_items
            VALUES (50, 1, 40, 10, '服务器配置与管理', '专业必修课',
                    '2025-2026', '2025-2026学年', '3', '第一学期', '2026-07-29 10:00:00')
            """
        )
        self.conn.execute(
            """
            INSERT INTO teacher_academic_exam_roster_items
            VALUES (60, 1, 30, 'active', '2026-07-29 10:10:00',
                    '服务器配置与管理', 'exam-course-1')
            """
        )
        self.conn.executemany(
            "INSERT INTO teacher_academic_exam_roster_students VALUES (?, 60, ?, ?, ?, ?)",
            [
                (601, 1, "20240102", "学生二", "软工2406班（专升本）"),
                (602, 2, "20240101", "学生一", "软工2406班（专升本）"),
            ],
        )
        fields = {
            "class_offering_id": 30,
            "academic_year": "2025-2026",
            "semester": "第一学期",
            "course_name": "服务器配置与管理",
            "class_name": "软工2406班（专升本）",
        }
        ordinary_payload = {
            "fields": fields,
            "structured": {
                "students": [
                    {
                        "student_number": "20240101",
                        "student_name": "学生一",
                        "calculated_scores": {"ordinary_score": 91.25},
                    },
                    {
                        "student_number": "20240102",
                        "student_name": "学生二",
                        "calculated_scores": {"ordinary_score": 82},
                    },
                ]
            },
        }
        exam_payload = {
            "fields": fields,
            "structured": {
                "students": [
                    {
                        "student_number": "20240101",
                        "student_name": "学生一",
                        "total_score": 88,
                    },
                    {
                        "student_number": "20240102",
                        "student_name": "学生二",
                        "total_score": 79.5,
                    },
                ]
            },
        }
        self.conn.executemany(
            """
            INSERT INTO material_ai_import_records
                (id, teacher_id, document_group, document_type, document_type_label,
                 parse_status, package_material_id, parsed_material_id, source_material_id,
                 export_payload_json, updated_at)
            VALUES (?, 1, 'final_material', ?, ?, 'completed', ?, NULL, NULL, ?, ?)
            """,
            [
                (
                    70,
                    "ordinary_grade_record",
                    "平时成绩表",
                    700,
                    json.dumps(ordinary_payload, ensure_ascii=False),
                    "2026-07-29 10:20:00",
                ),
                (
                    71,
                    "exam_grade_record",
                    "考核登分表",
                    710,
                    json.dumps(exam_payload, ensure_ascii=False),
                    "2026-07-29 10:21:00",
                ),
            ],
        )
        self.conn.executemany(
            "INSERT INTO course_material_assignments VALUES (?, 30)",
            [(700,), (710,)],
        )
        self.conn.commit()

    def test_readiness_and_payload_require_exact_context_and_preserve_roster_order(self):
        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )

        self.assertTrue(readiness["ready"])
        self.assertEqual(readiness["status"], "ready")
        self.assertEqual(
            [item["student_number"] for item in readiness["roster"]["preview"]],
            ["20240102", "20240101"],
        )
        self.assertEqual(
            [item["student_number"] for item in readiness["roster"]["students"]],
            ["20240102", "20240101"],
        )
        self.assertEqual(64, len(readiness["roster"]["signature"]))
        self.assertEqual(readiness["sources"]["ordinary_grade_record"]["matched_count"], 2)
        self.assertEqual(readiness["sources"]["exam_grade_record"]["matched_count"], 2)

        payload = build_final_grade_transcript_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
            expected_roster_synced_at="2026-07-29 10:10:00",
            expected_roster_signature=readiness["roster"]["signature"],
            expected_ordinary_record_id=70,
            expected_exam_record_id=71,
        )

        self.assertEqual(payload["document_type"], FINAL_GRADE_TRANSCRIPT_TYPE)
        students = payload["structured"]["students"]
        self.assertEqual([item["student_number"] for item in students], ["20240102", "20240101"])
        self.assertEqual([item["index"] for item in students], [1, 2])
        self.assertEqual([item["ordinary_score"] for item in students], [82, 91.25])
        self.assertEqual([item["final_score"] for item in students], [79.5, 88])
        self.assertEqual(
            payload["structured"]["source_lineage"]["exam_roster"]["order_field"],
            "row_order",
        )

    def test_readiness_blocks_same_number_with_different_name_and_stale_confirmation(self):
        record = self.conn.execute(
            "SELECT export_payload_json FROM material_ai_import_records WHERE id = 71"
        ).fetchone()
        payload = json.loads(record["export_payload_json"])
        payload["structured"]["students"][0]["student_name"] = "同号异名"
        self.conn.execute(
            "UPDATE material_ai_import_records SET export_payload_json = ? WHERE id = 71",
            (json.dumps(payload, ensure_ascii=False),),
        )
        self.conn.commit()

        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )
        self.assertFalse(readiness["ready"])
        self.assertEqual(readiness["status"], "exam_source_incomplete")
        self.assertEqual(readiness["sources"]["exam_grade_record"]["conflict_count"], 1)

        with self.assertRaises(HTTPException) as cm:
            build_final_grade_transcript_payload(
                self.conn,
                class_offering_id=30,
                teacher_id=1,
            )
        self.assertEqual(cm.exception.status_code, 409)

        payload["structured"]["students"][0]["student_name"] = "学生一"
        self.conn.execute(
            "UPDATE material_ai_import_records SET export_payload_json = ? WHERE id = 71",
            (json.dumps(payload, ensure_ascii=False),),
        )
        self.conn.commit()
        with self.assertRaises(HTTPException) as cm:
            build_final_grade_transcript_payload(
                self.conn,
                class_offering_id=30,
                teacher_id=1,
                expected_roster_synced_at="stale",
            )
        self.assertEqual(cm.exception.status_code, 409)
        self.assertIn("名单在确认后发生了变化", str(cm.exception.detail))

        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )
        self.conn.execute(
            """
            UPDATE teacher_academic_exam_roster_students
            SET row_order = 3
            WHERE id = 601
            """
        )
        self.conn.commit()
        with self.assertRaises(HTTPException) as cm:
            build_final_grade_transcript_payload(
                self.conn,
                class_offering_id=30,
                teacher_id=1,
                expected_roster_signature=readiness["roster"]["signature"],
            )
        self.assertEqual(cm.exception.status_code, 409)
        self.assertIn("名单内容在确认后发生了变化", str(cm.exception.detail))

    def test_explicit_offering_id_recovers_source_when_material_assignment_is_missing(self):
        self.conn.execute("DELETE FROM course_material_assignments WHERE material_id = 710")
        self.conn.commit()

        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )

        source = readiness["sources"]["exam_grade_record"]
        self.assertTrue(readiness["ready"])
        self.assertTrue(source["record_found"])
        self.assertEqual(source["record_id"], 71)

    def test_unscoped_unassigned_source_returns_direct_generation_link(self):
        record = self.conn.execute(
            "SELECT export_payload_json FROM material_ai_import_records WHERE id = 71"
        ).fetchone()
        payload = json.loads(record["export_payload_json"])
        payload["fields"].pop("class_offering_id")
        self.conn.execute(
            "UPDATE material_ai_import_records SET export_payload_json = ? WHERE id = 71",
            (json.dumps(payload, ensure_ascii=False),),
        )
        self.conn.execute("DELETE FROM course_material_assignments WHERE material_id = 710")
        self.conn.commit()

        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )

        self.assertFalse(readiness["ready"])
        source = readiness["sources"]["exam_grade_record"]
        self.assertFalse(source["record_found"])
        self.assertIn("/manage/teaching/exam-grade-records", source["generate_url"])
        self.assertIn("class_offering_id=30", source["generate_url"])

    def test_source_selection_prefers_complete_record_over_newer_incomplete_candidate(self):
        record = self.conn.execute(
            "SELECT export_payload_json FROM material_ai_import_records WHERE id = 71"
        ).fetchone()
        payload = json.loads(record["export_payload_json"])
        payload["structured"]["students"][0]["total_score"] = ""
        self.conn.execute(
            """
            INSERT INTO material_ai_import_records
                (id, teacher_id, document_group, document_type, document_type_label,
                 parse_status, package_material_id, parsed_material_id, source_material_id,
                 export_payload_json, updated_at)
            VALUES (72, 1, 'final_material', 'exam_grade_record', '考核登分表',
                    'completed', 720, NULL, NULL, ?, '2026-07-29 10:30:00')
            """,
            (json.dumps(payload, ensure_ascii=False),),
        )
        self.conn.execute(
            "INSERT INTO course_material_assignments VALUES (720, 30)"
        )
        self.conn.commit()

        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )

        self.assertTrue(readiness["ready"])
        self.assertEqual(readiness["sources"]["exam_grade_record"]["record_id"], 71)

    def test_duplicate_student_number_in_source_is_reported_and_blocked(self):
        record = self.conn.execute(
            "SELECT export_payload_json FROM material_ai_import_records WHERE id = 71"
        ).fetchone()
        payload = json.loads(record["export_payload_json"])
        payload["structured"]["students"].append(
            {
                "student_number": "20240101",
                "student_name": "学生一",
                "total_score": 90,
            }
        )
        self.conn.execute(
            "UPDATE material_ai_import_records SET export_payload_json = ? WHERE id = 71",
            (json.dumps(payload, ensure_ascii=False),),
        )
        self.conn.commit()

        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )

        source = readiness["sources"]["exam_grade_record"]
        self.assertFalse(readiness["ready"])
        self.assertEqual(source["duplicate_count"], 1)
        self.assertIn("重复学号", source["message"])

    def test_missing_students_carry_jump_targets_to_source_assignments(self):
        exam_record = self.conn.execute(
            "SELECT export_payload_json FROM material_ai_import_records WHERE id = 71"
        ).fetchone()
        exam_payload = json.loads(exam_record["export_payload_json"])
        exam_payload["structured"]["source_exam"] = {
            "assignment_id": 901,
            "assignment_title": "期末上机考核",
        }
        exam_payload["structured"]["students"][1]["total_score"] = ""
        self.conn.execute(
            "UPDATE material_ai_import_records SET export_payload_json = ? WHERE id = 71",
            (json.dumps(exam_payload, ensure_ascii=False),),
        )

        ordinary_record = self.conn.execute(
            "SELECT export_payload_json FROM material_ai_import_records WHERE id = 70"
        ).fetchone()
        ordinary_payload = json.loads(ordinary_record["export_payload_json"])
        ordinary_payload["structured"]["source_assignments"] = {
            "homework_assignment_ids": [801, 802],
            "homework_assignments": [
                {"id": 801, "title": "第一次作业"},
                {"id": 802, "title": "第二次作业"},
            ],
            "assessment_assignment_id": 803,
            "assessment_assignment": {"id": 803, "title": "阶段测评"},
        }
        # 学生二 disappears from the ordinary record entirely.
        ordinary_payload["structured"]["students"] = [
            student
            for student in ordinary_payload["structured"]["students"]
            if student["student_number"] != "20240102"
        ]
        self.conn.execute(
            "UPDATE material_ai_import_records SET export_payload_json = ? WHERE id = 70",
            (json.dumps(ordinary_payload, ensure_ascii=False),),
        )
        self.conn.commit()

        readiness = build_final_grade_transcript_readiness(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )

        self.assertFalse(readiness["ready"])
        exam_missing = readiness["sources"]["exam_grade_record"]["missing_students"]
        self.assertEqual(len(exam_missing), 1)
        exam_targets = exam_missing[0]["jump_targets"]
        self.assertEqual(len(exam_targets), 1)
        self.assertEqual(exam_targets[0]["assignment_id"], 901)
        self.assertEqual(exam_targets[0]["title"], "期末上机考核")
        self.assertEqual(exam_targets[0]["url"], "/assignment/901?locate=20240102")

        ordinary_missing = readiness["sources"]["ordinary_grade_record"]["missing_students"]
        self.assertEqual(len(ordinary_missing), 1)
        ordinary_targets = ordinary_missing[0]["jump_targets"]
        self.assertEqual(
            [target["assignment_id"] for target in ordinary_targets],
            [801, 802, 803],
        )
        self.assertTrue(all("locate=20240102" in target["url"] for target in ordinary_targets))

    def test_blank_ordinary_scores_target_only_blank_source_assignments(self):
        from classroom_app.services.final_grade_transcript_service import (
            _record_assignment_targets,
            _record_raw_students,
            _student_jump_targets,
        )

        record = self.conn.execute(
            "SELECT export_payload_json FROM material_ai_import_records WHERE id = 70"
        ).fetchone()
        payload = json.loads(record["export_payload_json"])
        payload["structured"]["source_assignments"] = {
            "homework_assignment_ids": [801, 802],
            "homework_assignments": [
                {"id": 801, "title": "第一次作业"},
                {"id": 802, "title": "第二次作业"},
            ],
            "assessment_assignment_id": 803,
            "assessment_assignment": {"id": 803, "title": "阶段测评"},
        }
        for student in payload["structured"]["students"]:
            if student["student_number"] == "20240102":
                student["homework_scores"] = [88, None]
                student["assessment_score"] = 90
        self.conn.execute(
            "UPDATE material_ai_import_records SET export_payload_json = ? WHERE id = 70",
            (json.dumps(payload, ensure_ascii=False),),
        )
        self.conn.commit()

        record_row = self.conn.execute(
            "SELECT * FROM material_ai_import_records WHERE id = 70"
        ).fetchone()
        targets = _record_assignment_targets(record_row, "ordinary_grade_record")
        raw_students = _record_raw_students(record_row)
        jump = _student_jump_targets(
            targets,
            raw_students.get("20240102"),
            document_type="ordinary_grade_record",
            student_number="20240102",
        )
        self.assertEqual([target["assignment_id"] for target in jump], [802])
        self.assertEqual(jump[0]["url"], "/assignment/802?locate=20240102")

    def test_xlsx_reproduces_template_styles_and_round_trips_every_column(self):
        payload = build_final_grade_transcript_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )
        content = build_final_grade_transcript_xlsx(payload)
        workbook = load_workbook(io.BytesIO(content))
        worksheet = workbook.active

        self.assertEqual(worksheet.title, "学生成绩录入模板")
        self.assertEqual(
            [worksheet.column_dimensions[column].width for column in "ABCDEFG"],
            [20.0] * 7,
        )
        self.assertEqual(workbook._named_styles[0].font.name, "等线")
        self.assertEqual(workbook._named_styles[0].font.sz, 11)
        self.assertIn(b'Aptos Narrow', workbook.loaded_theme)
        self.assertIn("等线".encode("utf-8"), workbook.loaded_theme)
        self.assertEqual(worksheet.row_dimensions[1].height, 25.0)
        self.assertEqual(worksheet.row_dimensions[2].height, 20.0)
        self.assertEqual(
            [worksheet.cell(1, column).value for column in range(1, 8)],
            ["序号", "班级", "学号", "姓名", "平时(必填)", "期末(必填)", "备注"],
        )
        self.assertEqual(worksheet["A1"].font.name, "宋体")
        self.assertEqual(worksheet["A1"].font.sz, 23)
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertEqual(worksheet["A1"].font.color.indexed, 8)
        self.assertEqual(worksheet["E1"].font.name, "黑体")
        self.assertEqual(worksheet["E1"].font.sz, 15)
        self.assertEqual(worksheet["E1"].font.color.indexed, 10)
        self.assertEqual(worksheet["A1"].fill.fgColor.indexed, 22)
        self.assertEqual(worksheet["A2"].font.name, "宋体")
        self.assertEqual(worksheet["A2"].font.sz, 13)
        self.assertEqual(worksheet["A2"].number_format, "General")
        self.assertEqual(worksheet["E2"].number_format, "0.00_ ")
        self.assertFalse(worksheet["A2"].protection.locked)
        self.assertEqual(worksheet["A2"].border.left.style, "thin")
        self.assertTrue(worksheet["A2"].border.left.color.auto)
        self.assertIsNotNone(worksheet["A2"].border.diagonal)
        self.assertEqual(worksheet["A2"].alignment.horizontal, "center")
        self.assertIsNotNone(worksheet["E1"].comment)
        self.assertIsNotNone(worksheet["F1"].comment)
        self.assertIsNotNone(worksheet["G1"].comment)
        self.assertEqual(
            worksheet["E1"].comment.text,
            "该分项或者阶段成绩录入级制为【百分制】,请输入 0 至 100 之间的数值!",
        )
        self.assertEqual(worksheet["E1"].comment.author, "None")
        validations = list(worksheet.data_validations.dataValidation)
        self.assertEqual(len(validations), 2)
        self.assertEqual(str(validations[0].sqref), "E2:F3")
        self.assertEqual(str(validations[1].sqref), "G2:G3")
        self.assertIsNone(validations[0].operator)
        self.assertTrue(validations[0].allowBlank)
        self.assertTrue(validations[1].showErrorMessage)
        self.assertEqual(worksheet.page_margins.left, 0.7)
        self.assertEqual(worksheet.page_margins.top, 0.75)

        metadata = {
            "academic_year": "2025-2026",
            "semester": "第一学期",
            "school": "广西外国语学院",
            "college": "信息工程学院",
            "department": "软件工程系",
            "class_name": "软工2406班（专升本）",
            "course_name": "服务器配置与管理",
            "course_nature": "专业必修课",
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "服务器配置与管理学生成绩录入模板[张海林].xlsx"
            path.write_bytes(content)
            parsed = parse_final_grade_transcript_file(path, path.name, metadata)

        students = parsed.export_payload["structured"]["students"]
        self.assertEqual(len(students), 2)
        self.assertEqual(students[0]["class_name"], "软工2406班（专升本）")
        self.assertEqual(students[0]["student_number"], "20240102")
        self.assertEqual(students[0]["student_name"], "学生二")
        self.assertEqual(students[0]["ordinary_score"], 82)
        self.assertEqual(students[0]["final_score"], 79.5)
        self.assertEqual(students[0]["remark"], "")
        self.assertTrue(
            parsed.export_payload["structured"]["import_contract"]["all_columns_preserved"]
        )

    def test_import_requires_academic_period_and_organization_metadata(self):
        payload = build_final_grade_transcript_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "期末成绩单.xlsx"
            path.write_bytes(build_final_grade_transcript_xlsx(payload))
            with self.assertRaises(HTTPException) as cm:
                parse_final_grade_transcript_file(path, path.name, {})
        self.assertEqual(cm.exception.status_code, 422)
        self.assertIn("学年", str(cm.exception.detail))
        self.assertIn("学期", str(cm.exception.detail))
        self.assertIn("学院", str(cm.exception.detail))

    def test_shared_export_pipeline_returns_xlsx(self):
        payload = build_final_grade_transcript_payload(
            self.conn,
            class_offering_id=30,
            teacher_id=1,
        )
        artifact = build_material_export_artifact(
            payload,
            fallback_filename="期末成绩单",
            requested_format="xlsx",
        )
        self.assertEqual(artifact.media_type, XLSX_MEDIA_TYPE)
        self.assertTrue(artifact.filename.endswith(".xlsx"))
        self.assertGreater(len(artifact.content), 1000)

    def test_semester_normalization_accepts_platform_and_jwxt_period_codes(self):
        self.assertEqual("第一学期", normalize_final_grade_semester("2025-2026-1"))
        self.assertEqual("第二学期", normalize_final_grade_semester("2025-2026-2"))
        self.assertEqual("第一学期", normalize_final_grade_semester("3"))
        self.assertEqual("第二学期", normalize_final_grade_semester("12"))


if __name__ == "__main__":
    unittest.main()
