from __future__ import annotations

import json
import re
import sqlite3
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import ROSTER_DIR
from ..database import get_db_connection
from .academic_calendar_sync_service import prepare_current_semester_from_academic_system
from .academic_classroom_sync_service import load_teacher_teaching_place_by_key
from .academic_integration_service import (
    load_teacher_academic_access_method,
    open_authenticated_academic_client,
)
from .academic_service import china_now, parse_date_input


ACADEMIC_EXAM_ROSTER_SOURCE = "gxufl_jwxt"
SCHOOL_CODE = "gxufl"

ZF_EXAM_COURSE_INDEX_PATH = (
    "/cjlrgl/jscjlr_cxJscjlrIndex.html?doType=details&gnmkdm=N302505&layout=default"
)
ZF_EXAM_COURSE_LIST_PATH = "/cjlrgl/jscjlr_cxJscjlrIndex.html?doType=query&gnmkdm=N302505"
ZF_EXAM_STUDENT_LIST_PATH = "/cjlrgl/jscjlr_cxZkcj.html"

EXAM_COURSE_PAGE_SIZE = 200
EXAM_ROSTER_TOTAL_ROWS = 47
EXAM_ROSTER_TABLE_START_ROW = 8
EXAM_ROSTER_FIRST_STUDENT_ROW = 9
EXAM_ROSTER_STUDENT_ROWS_PER_SIDE = EXAM_ROSTER_TOTAL_ROWS - EXAM_ROSTER_FIRST_STUDENT_ROW + 1
MAX_EXAM_ROSTER_STUDENTS = EXAM_ROSTER_STUDENT_ROWS_PER_SIDE * 2


@dataclass
class AcademicExamCourse:
    exam_course_key: str
    academic_year: str = ""
    academic_year_name: str = ""
    academic_term: str = ""
    academic_term_name: str = ""
    course_code: str = ""
    course_internal_id: str = ""
    course_name: str = ""
    teaching_class_id: str = ""
    teaching_class_name: str = ""
    class_composition: str = ""
    teacher_name: str = ""
    schedule_text: str = ""
    exam_method: str = ""
    grade_entry_status: str = ""
    credits: float | None = None
    declared_student_count: int = 0
    raw_json: dict[str, Any] = field(default_factory=dict)
    source_url: str = ""


@dataclass
class AcademicExamStudent:
    student_number: str
    student_name: str
    gender: str = ""
    admin_class_code: str = ""
    admin_class_name: str = ""
    college: str = ""
    grade: str = ""
    major: str = ""
    school_status: str = ""
    selection_type: str = ""
    raw_json: dict[str, Any] = field(default_factory=dict)


def _now_iso() -> str:
    return china_now().replace(tzinfo=None).isoformat(timespec="seconds")


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _json_loads(value: Any, fallback: Any) -> Any:
    if not value:
        return fallback
    try:
        return json.loads(str(value))
    except (TypeError, json.JSONDecodeError):
        return fallback


def _normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def _normalize_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def _field(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        for candidate in (key, key.lower(), key.upper()):
            if candidate in row and row.get(candidate) not in (None, ""):
                return _normalize_space(row.get(candidate))
    return ""


def _parse_int(value: Any) -> int:
    match = re.search(r"\d+", str(value or ""))
    if not match:
        return 0
    try:
        return int(match.group(0))
    except ValueError:
        return 0


def _parse_float(value: Any) -> float | None:
    match = re.search(r"\d+(?:\.\d+)?", str(value or ""))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _optional_int(value: Any) -> int | None:
    try:
        parsed = int(value or 0)
    except (TypeError, ValueError):
        return None
    return parsed or None


def _semester_year_start(semester: dict[str, Any]) -> int:
    name = str(semester.get("name") or semester.get("semester_name") or "")
    match = re.search(r"(20\d{2})\s*[-—至]\s*(20\d{2})", name)
    if match:
        return int(match.group(1))
    start_date = parse_date_input(semester.get("start_date"))
    if start_date:
        return start_date.year if start_date.month >= 8 else start_date.year - 1
    today = china_now().date()
    return today.year if today.month >= 8 else today.year - 1


def _semester_term_number(semester: dict[str, Any]) -> int:
    name = str(semester.get("name") or semester.get("semester_name") or "")
    if re.search(r"(第\s*)?(二|2)\s*学期", name):
        return 2
    if re.search(r"(第\s*)?(一|1)\s*学期", name):
        return 1
    start_date = parse_date_input(semester.get("start_date"))
    if start_date and 1 <= start_date.month <= 7:
        return 2
    return 1


def _term_param_candidates(semester: dict[str, Any]) -> list[dict[str, str]]:
    year_start = _semester_year_start(semester)
    term_number = _semester_term_number(semester)
    year_values = [str(year_start), f"{year_start}-{year_start + 1}"]
    term_values = ["12", "2"] if term_number == 2 else ["3", "1"]
    return [{"xnm": xnm, "xqm": xqm} for xnm in year_values for xqm in term_values]


def _decode_json_response(response: httpx.Response) -> Any:
    text = response.content.decode("utf-8", errors="replace")
    return json.loads(text)


def _ajax_headers(client: httpx.AsyncClient, *, referer: str = ZF_EXAM_COURSE_INDEX_PATH) -> dict[str, str]:
    base_url = str(client.base_url).rstrip("/")
    return {
        "Accept": "application/json,text/javascript,*/*;q=0.8",
        "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": base_url,
        "Referer": base_url + referer,
    }


def _jqgrid_form(
    *,
    term_params: dict[str, str],
    page: int = 1,
    show_count: int = EXAM_COURSE_PAGE_SIZE,
) -> dict[str, str]:
    return {
        **term_params,
        "lrzt": "",
        "kcmc": "",
        "_search": "false",
        "nd": str(int(time.time() * 1000)),
        "queryModel.showCount": str(show_count),
        "queryModel.currentPage": str(max(1, int(page or 1))),
        "queryModel.sortName": "",
        "queryModel.sortOrder": "asc",
        "time": str(max(0, int(page or 1) - 1)),
    }


def _extract_items(payload: Any) -> tuple[list[dict[str, Any]], int, int]:
    if isinstance(payload, list):
        return [dict(item) for item in payload if isinstance(item, dict)], len(payload), 1
    if not isinstance(payload, dict):
        return [], 0, 0
    for key in ("items", "rows", "data"):
        if isinstance(payload.get(key), list):
            rows = [dict(item) for item in payload[key] if isinstance(item, dict)]
            total = _parse_int(payload.get("totalCount") or payload.get("totalResult") or len(rows))
            total_page = max(1, _parse_int(payload.get("totalPage") or payload.get("total_page") or 1))
            return rows, total, total_page
    return [], _parse_int(payload.get("totalCount") or payload.get("totalResult")), 1


def _exam_course_from_row(row: dict[str, Any], *, source_url: str, term_params: dict[str, str]) -> AcademicExamCourse | None:
    teaching_class_id = _field(row, "jxb_id", "JXB_ID")
    course_code = _field(row, "kch", "KCH")
    course_internal_id = _field(row, "kch_id", "KCH_ID")
    course_name = _field(row, "kcmc", "KCMC")
    if not teaching_class_id or not course_name:
        return None
    return AcademicExamCourse(
        exam_course_key=teaching_class_id,
        academic_year=_field(row, "xnm", "XNM") or term_params.get("xnm", ""),
        academic_year_name=_field(row, "xnmmc", "XNMMC"),
        academic_term=_field(row, "xqm", "XQM") or term_params.get("xqm", ""),
        academic_term_name=_field(row, "xqmmc", "XQMMC"),
        course_code=course_code or course_internal_id,
        course_internal_id=course_internal_id or course_code,
        course_name=course_name,
        teaching_class_id=teaching_class_id,
        teaching_class_name=_field(row, "jxbmc", "JXBMC"),
        class_composition=_field(row, "jxbzc", "JXBZC"),
        teacher_name=_field(row, "jsxm", "JSXM"),
        schedule_text=_field(row, "sksj", "SKSJ"),
        exam_method=_field(row, "khfsmc", "KHFSMC"),
        grade_entry_status=_field(row, "lrzt", "lrztmc", "LRZT", "LRZTMC"),
        credits=_parse_float(_field(row, "xf", "XF")),
        declared_student_count=_parse_int(_field(row, "jxbrs", "rs", "JXBRS", "RS")),
        raw_json=dict(row),
        source_url=source_url,
    )


def _exam_student_from_row(row: dict[str, Any]) -> AcademicExamStudent | None:
    student_number = _field(row, "xh", "XH", "xh_id", "XH_ID")
    student_name = _field(row, "xm", "XM")
    if not student_number or not student_name:
        return None
    return AcademicExamStudent(
        student_number=student_number,
        student_name=student_name,
        gender=_field(row, "xbmc", "xb", "XBMC", "XB"),
        admin_class_code=_field(row, "bh", "bh_id", "BH", "BH_ID"),
        admin_class_name=_field(row, "bj", "BJ"),
        college=_field(row, "jgmc", "JGMC"),
        grade=_field(row, "njmc", "NJMC"),
        major=_field(row, "zymc", "ZYMC"),
        school_status=_field(row, "xjztmc", "XJZTMC"),
        selection_type=_field(row, "cjxzm", "CJXZM"),
        raw_json=dict(row),
    )


async def _fetch_exam_courses(
    client: httpx.AsyncClient,
    semester: dict[str, Any],
) -> tuple[list[AcademicExamCourse], list[dict[str, Any]], dict[str, str] | None]:
    sources: list[dict[str, Any]] = []
    try:
        response = await client.get(
            ZF_EXAM_COURSE_INDEX_PATH,
            headers={"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"},
        )
        sources.append(
            {
                "path": ZF_EXAM_COURSE_INDEX_PATH,
                "method": "GET",
                "status_code": response.status_code,
                "parser": "exam_course_index",
                "url": str(response.url),
            }
        )
    except httpx.HTTPError as exc:
        sources.append(
            {
                "path": ZF_EXAM_COURSE_INDEX_PATH,
                "method": "GET",
                "status": "failed",
                "message": str(exc)[:180],
            }
        )

    for term_params in _term_param_candidates(semester):
        courses: list[AcademicExamCourse] = []
        total_page = 1
        for page in range(1, 50):
            form = _jqgrid_form(term_params=term_params, page=page)
            response = await client.post(
                ZF_EXAM_COURSE_LIST_PATH,
                data=form,
                headers=_ajax_headers(client),
            )
            payload = _decode_json_response(response)
            rows, total_count, total_page = _extract_items(payload)
            sources.append(
                {
                    "path": ZF_EXAM_COURSE_LIST_PATH,
                    "method": "POST",
                    "params": {**term_params, "page": page, "showCount": EXAM_COURSE_PAGE_SIZE},
                    "status_code": response.status_code,
                    "parser": "exam_course_list",
                    "item_count": len(rows),
                    "total_count": total_count,
                    "total_page": total_page,
                    "url": str(response.url),
                }
            )
            for row in rows:
                course = _exam_course_from_row(row, source_url=str(response.url), term_params=term_params)
                if course:
                    courses.append(course)
            if page >= total_page:
                break
        if courses:
            return courses, sources, term_params
    return [], sources, None


async def _fetch_exam_students(
    client: httpx.AsyncClient,
    course: AcademicExamCourse,
    sources: list[dict[str, Any]],
) -> list[AcademicExamStudent]:
    form = {
        "doType": "query",
        "sfgly": "0",
        "jxb_id": course.teaching_class_id,
        "jgh_id": "",
        "kch_id": course.course_internal_id,
        "tjpccjlrbj": "0",
        "sfzdtskc": _field(course.raw_json, "sfzdtskc") or "0",
        "drfs": "",
    }
    response = await client.post(
        ZF_EXAM_STUDENT_LIST_PATH,
        data=form,
        headers=_ajax_headers(client, referer=ZF_EXAM_COURSE_INDEX_PATH),
    )
    payload = _decode_json_response(response)
    rows, total_count, total_page = _extract_items(payload)
    sources.append(
        {
            "path": ZF_EXAM_STUDENT_LIST_PATH,
            "method": "POST",
            "params": {
                "jxb_id": course.teaching_class_id,
                "kch_id": course.course_internal_id,
                "doType": "query",
            },
            "status_code": response.status_code,
            "parser": "exam_student_list",
            "item_count": len(rows),
            "total_count": total_count,
            "total_page": total_page,
            "url": str(response.url),
        }
    )
    students: list[AcademicExamStudent] = []
    for row in rows:
        student = _exam_student_from_row(row)
        if student:
            students.append(student)
    return students


def _load_offering_context(conn: sqlite3.Connection, teacher_id: int, class_offering_id: int) -> dict[str, Any] | None:
    row = conn.execute(
        """
        SELECT o.id AS class_offering_id,
               o.class_id,
               o.course_id,
               o.teacher_id,
               o.semester_id,
               o.semester,
               o.first_class_date,
               o.academic_teaching_class_name,
               c.name AS class_name,
               c.academic_class_code,
               c.academic_class_name,
               c.academic_college,
               c.academic_grade,
               c.academic_major,
               cr.name AS course_name,
               cr.academic_course_code,
               cr.credits,
               s.name AS semester_name,
               s.start_date AS semester_start_date,
               s.end_date AS semester_end_date,
               t.name AS teacher_name
        FROM class_offerings o
        JOIN classes c ON c.id = o.class_id
        JOIN courses cr ON cr.id = o.course_id
        JOIN teachers t ON t.id = o.teacher_id
        LEFT JOIN academic_semesters s ON s.id = o.semester_id
        WHERE o.id = ? AND o.teacher_id = ?
        LIMIT 1
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchone()
    return dict(row) if row else None


def _load_semester_for_offering(
    conn: sqlite3.Connection,
    teacher_id: int,
    offering_context: dict[str, Any],
) -> dict[str, Any] | None:
    semester_id = _optional_int(offering_context.get("semester_id"))
    if semester_id:
        row = conn.execute(
            "SELECT * FROM academic_semesters WHERE id = ? AND teacher_id = ? LIMIT 1",
            (semester_id, int(teacher_id)),
        ).fetchone()
        if row:
            return dict(row)
    name = offering_context.get("semester_name") or offering_context.get("semester")
    if name:
        return {
            "id": semester_id,
            "name": str(name),
            "start_date": offering_context.get("semester_start_date") or "",
            "end_date": offering_context.get("semester_end_date") or "",
        }
    row = conn.execute(
        """
        SELECT *
        FROM academic_semesters
        WHERE teacher_id = ?
          AND date(start_date) <= date(?)
          AND date(end_date) >= date(?)
        ORDER BY start_date DESC, id DESC
        LIMIT 1
        """,
        (int(teacher_id), china_now().date().isoformat(), china_now().date().isoformat()),
    ).fetchone()
    return dict(row) if row else None


def _score_exam_course(course: AcademicExamCourse, context: dict[str, Any]) -> int:
    score = 0
    local_course_code = _normalize_key(context.get("academic_course_code"))
    course_codes = {
        _normalize_key(course.course_code),
        _normalize_key(course.course_internal_id),
    }
    if local_course_code and local_course_code in course_codes:
        score += 60
    elif local_course_code and any(local_course_code in item or item in local_course_code for item in course_codes if item):
        score += 30

    local_course_name = _normalize_key(context.get("course_name"))
    remote_course_name = _normalize_key(course.course_name)
    if local_course_name and local_course_name == remote_course_name:
        score += 36
    elif local_course_name and remote_course_name and (
        local_course_name in remote_course_name or remote_course_name in local_course_name
    ):
        score += 20

    teaching_class_name = _normalize_key(context.get("academic_teaching_class_name"))
    remote_teaching_class = _normalize_key(course.teaching_class_name)
    if teaching_class_name and teaching_class_name == remote_teaching_class:
        score += 48
    elif teaching_class_name and remote_teaching_class and (
        teaching_class_name in remote_teaching_class or remote_teaching_class in teaching_class_name
    ):
        score += 20

    remote_class_composition = _normalize_key(course.class_composition)
    class_candidates = [
        context.get("academic_class_name"),
        context.get("academic_class_code"),
        context.get("class_name"),
    ]
    for class_candidate in class_candidates:
        normalized = _normalize_key(class_candidate)
        if normalized and remote_class_composition and (
            normalized in remote_class_composition or remote_class_composition in normalized
        ):
            score += 36
            break
    return score


def _serialize_exam_course_candidate(course: AcademicExamCourse, *, score: int = 0) -> dict[str, Any]:
    return {
        "exam_course_key": course.exam_course_key,
        "course_code": course.course_code,
        "course_name": course.course_name,
        "teaching_class_id": course.teaching_class_id,
        "teaching_class_name": course.teaching_class_name,
        "class_composition": course.class_composition,
        "credits": course.credits,
        "declared_student_count": course.declared_student_count,
        "schedule_text": course.schedule_text,
        "exam_method": course.exam_method,
        "score": score,
    }


def _select_exam_course(
    courses: list[AcademicExamCourse],
    context: dict[str, Any],
    *,
    requested_exam_course_key: str = "",
) -> tuple[AcademicExamCourse | None, list[dict[str, Any]], bool]:
    if requested_exam_course_key:
        for course in courses:
            if course.exam_course_key == requested_exam_course_key:
                return course, [_serialize_exam_course_candidate(course, score=999)], False
        candidates = [
            _serialize_exam_course_candidate(course, score=_score_exam_course(course, context))
            for course in courses
        ]
        candidates.sort(key=lambda item: int(item.get("score") or 0), reverse=True)
        return None, candidates[:12], True

    scored = [
        (course, _score_exam_course(course, context))
        for course in courses
    ]
    scored.sort(key=lambda item: item[1], reverse=True)
    candidates = [_serialize_exam_course_candidate(course, score=score) for course, score in scored[:12]]
    if not scored or scored[0][1] < 45:
        return None, candidates, True
    if len(scored) > 1 and scored[1][1] >= scored[0][1] - 8:
        return None, candidates, True
    return scored[0][0], candidates, False


def _local_roster_alignment(
    conn: sqlite3.Connection,
    *,
    class_id: int,
    students: list[AcademicExamStudent],
) -> dict[str, Any]:
    local_rows = conn.execute(
        """
        SELECT id, student_id_number, name, gender
        FROM students
        WHERE class_id = ?
          AND COALESCE(enrollment_status, 'active') = 'active'
        ORDER BY student_id_number ASC, id ASC
        """,
        (int(class_id),),
    ).fetchall()
    local_by_number = {
        _normalize_space(row["student_id_number"]): dict(row)
        for row in local_rows
        if _normalize_space(row["student_id_number"])
    }
    remote_by_number = {student.student_number: student for student in students if student.student_number}
    missing_local = [
        {
            "student_number": student.student_number,
            "student_name": student.student_name,
            "admin_class_name": student.admin_class_name,
        }
        for student in students
        if student.student_number and student.student_number not in local_by_number
    ]
    extra_local = [
        {
            "student_number": row["student_id_number"],
            "student_name": row["name"],
        }
        for number, row in local_by_number.items()
        if number not in remote_by_number
    ]
    return {
        "local_active_count": len(local_by_number),
        "exam_student_count": len(students),
        "matched_local_count": len(set(local_by_number).intersection(remote_by_number)),
        "missing_local_count": len(missing_local),
        "extra_local_count": len(extra_local),
        "missing_local_students": missing_local[:20],
        "extra_local_students": extra_local[:20],
    }


def _upsert_exam_roster_item(
    conn: sqlite3.Connection,
    *,
    teacher_id: int,
    semester: dict[str, Any],
    context: dict[str, Any],
    course: AcademicExamCourse,
    students: list[AcademicExamStudent],
    synced_at: str,
) -> int:
    semester_id = _optional_int(semester.get("id"))
    cursor = conn.execute(
        """
        INSERT INTO teacher_academic_exam_roster_items (
            teacher_id, semester_id, class_offering_id, course_id, class_id, school_code,
            academic_year, academic_year_name, academic_term, academic_term_name,
            exam_course_key, course_code, course_internal_id, course_name,
            teaching_class_id, teaching_class_name, class_composition, teacher_name,
            schedule_text, exam_method, grade_entry_status, credits,
            declared_student_count, roster_student_count, raw_json, source_url,
            sync_status, synced_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
        ON CONFLICT (teacher_id, school_code, academic_year, academic_term, exam_course_key)
        DO UPDATE SET
            semester_id = excluded.semester_id,
            class_offering_id = excluded.class_offering_id,
            course_id = excluded.course_id,
            class_id = excluded.class_id,
            academic_year_name = excluded.academic_year_name,
            academic_term_name = excluded.academic_term_name,
            course_code = excluded.course_code,
            course_internal_id = excluded.course_internal_id,
            course_name = excluded.course_name,
            teaching_class_id = excluded.teaching_class_id,
            teaching_class_name = excluded.teaching_class_name,
            class_composition = excluded.class_composition,
            teacher_name = excluded.teacher_name,
            schedule_text = excluded.schedule_text,
            exam_method = excluded.exam_method,
            grade_entry_status = excluded.grade_entry_status,
            credits = excluded.credits,
            declared_student_count = excluded.declared_student_count,
            roster_student_count = excluded.roster_student_count,
            raw_json = excluded.raw_json,
            source_url = excluded.source_url,
            sync_status = 'active',
            synced_at = excluded.synced_at,
            updated_at = excluded.updated_at
        """,
        (
            int(teacher_id),
            semester_id,
            int(context["class_offering_id"]),
            int(context["course_id"]),
            int(context["class_id"]),
            SCHOOL_CODE,
            course.academic_year,
            course.academic_year_name,
            course.academic_term,
            course.academic_term_name,
            course.exam_course_key,
            course.course_code,
            course.course_internal_id,
            course.course_name,
            course.teaching_class_id,
            course.teaching_class_name,
            course.class_composition,
            course.teacher_name,
            course.schedule_text,
            course.exam_method,
            course.grade_entry_status,
            course.credits,
            int(course.declared_student_count or len(students)),
            len(students),
            _json_dumps(course.raw_json),
            course.source_url,
            synced_at,
            synced_at,
        ),
    )
    if cursor.lastrowid:
        row = conn.execute(
            """
            SELECT id
            FROM teacher_academic_exam_roster_items
            WHERE teacher_id = ?
              AND school_code = ?
              AND academic_year = ?
              AND academic_term = ?
              AND exam_course_key = ?
            LIMIT 1
            """,
            (int(teacher_id), SCHOOL_CODE, course.academic_year, course.academic_term, course.exam_course_key),
        ).fetchone()
        if row:
            return int(row["id"])
    row = conn.execute(
        """
        SELECT id
        FROM teacher_academic_exam_roster_items
        WHERE teacher_id = ?
          AND school_code = ?
          AND academic_year = ?
          AND academic_term = ?
          AND exam_course_key = ?
        LIMIT 1
        """,
        (int(teacher_id), SCHOOL_CODE, course.academic_year, course.academic_term, course.exam_course_key),
    ).fetchone()
    return int(row["id"]) if row else 0


def _replace_exam_roster_students(
    conn: sqlite3.Connection,
    *,
    teacher_id: int,
    semester: dict[str, Any],
    context: dict[str, Any],
    exam_roster_item_id: int,
    course: AcademicExamCourse,
    students: list[AcademicExamStudent],
    synced_at: str,
) -> None:
    conn.execute(
        "DELETE FROM teacher_academic_exam_roster_students WHERE exam_roster_item_id = ?",
        (int(exam_roster_item_id),),
    )
    semester_id = _optional_int(semester.get("id"))
    for index, student in enumerate(students, start=1):
        local_student = conn.execute(
            """
            SELECT id
            FROM students
            WHERE class_id = ? AND student_id_number = ?
            LIMIT 1
            """,
            (int(context["class_id"]), student.student_number),
        ).fetchone()
        conn.execute(
            """
            INSERT INTO teacher_academic_exam_roster_students (
                teacher_id, semester_id, exam_roster_item_id, class_offering_id, class_id, student_id,
                school_code, academic_year, academic_term, exam_course_key,
                student_number, student_name, gender, admin_class_code, admin_class_name,
                college, grade, major, school_status, selection_type,
                seat_no, row_order, raw_json, source_url, synced_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(teacher_id),
                semester_id,
                int(exam_roster_item_id),
                int(context["class_offering_id"]),
                int(context["class_id"]),
                int(local_student["id"]) if local_student else None,
                SCHOOL_CODE,
                course.academic_year,
                course.academic_term,
                course.exam_course_key,
                student.student_number,
                student.student_name,
                student.gender,
                student.admin_class_code,
                student.admin_class_name,
                student.college,
                student.grade,
                student.major,
                student.school_status,
                student.selection_type,
                index,
                index,
                _json_dumps(student.raw_json),
                course.source_url,
                synced_at,
                synced_at,
            ),
        )


def _admin_class_name_from_students(students: list[dict[str, Any]], fallback: str = "") -> str:
    names = []
    for student in students:
        name = _normalize_space(student.get("admin_class_name"))
        if name and name not in names:
            names.append(name)
    if len(names) == 1:
        return names[0]
    if names:
        return "、".join(names[:4]) + ("等" if len(names) > 4 else "")
    return fallback


def _load_latest_exam_item(
    conn: sqlite3.Connection,
    *,
    teacher_id: int,
    class_offering_id: int,
) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT *
        FROM teacher_academic_exam_roster_items
        WHERE teacher_id = ?
          AND class_offering_id = ?
          AND sync_status = 'active'
        ORDER BY synced_at DESC, id DESC
        LIMIT 1
        """,
        (int(teacher_id), int(class_offering_id)),
    ).fetchone()


def _load_exam_students(conn: sqlite3.Connection, item_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT *
        FROM teacher_academic_exam_roster_students
        WHERE exam_roster_item_id = ?
        ORDER BY row_order ASC, id ASC
        """,
        (int(item_id),),
    ).fetchall()
    return [dict(row) for row in rows]


def _format_semester_exam_title(item: dict[str, Any], context: dict[str, Any]) -> str:
    academic_year_name = _normalize_space(item.get("academic_year_name"))
    if academic_year_name:
        year_part = academic_year_name if "学年" in academic_year_name else f"{academic_year_name}学年"
    else:
        year_start = _parse_int(item.get("academic_year"))
        year_part = f"{year_start}-{year_start + 1}学年" if year_start else "本学年"
    term = str(item.get("academic_term") or "")
    semester_name = str(context.get("semester_name") or context.get("semester") or "")
    if term in {"12", "2"} or re.search(r"(第\s*)?(二|2)\s*学期", semester_name):
        term_part = "第二学期"
    elif term in {"3", "1"} or re.search(r"(第\s*)?(一|1)\s*学期", semester_name):
        term_part = "第一学期"
    else:
        term_part = _normalize_space(item.get("academic_term_name")) or "本学期"
    return f"{year_part}{term_part}期末考试"


def _datetime_local(value: Any) -> str:
    text = _normalize_space(value)
    if not text:
        return ""
    text = text.replace(" ", "T")
    try:
        return datetime.fromisoformat(text[:16]).isoformat(timespec="minutes")
    except ValueError:
        return ""


def _last_session_default(conn: sqlite3.Connection, class_offering_id: int) -> tuple[str, str]:
    row = conn.execute(
        """
        SELECT session_date, academic_location
        FROM class_offering_sessions
        WHERE class_offering_id = ?
          AND COALESCE(schedule_status, 'scheduled') != 'cancelled'
        ORDER BY date(session_date) DESC, order_index DESC, id DESC
        LIMIT 1
        """,
        (int(class_offering_id),),
    ).fetchone()
    if not row:
        return "", ""
    session_date = parse_date_input(row["session_date"])
    if not session_date:
        return "", _normalize_space(row["academic_location"])
    return f"{session_date.isoformat()}T08:00", _normalize_space(row["academic_location"])


def _invigilation_default(
    conn: sqlite3.Connection,
    *,
    teacher_id: int,
    item: dict[str, Any],
) -> tuple[str, str]:
    semester_id = _optional_int(item.get("semester_id"))
    params: list[Any] = [int(teacher_id)]
    where = ["teacher_id = ?", "sync_status = 'active'"]
    if semester_id:
        where.append("semester_id = ?")
        params.append(semester_id)
    course_code = _normalize_space(item.get("course_code"))
    course_name = _normalize_space(item.get("course_name"))
    if course_code:
        where.append("course_code = ?")
        params.append(course_code)
    elif course_name:
        where.append("course_name = ?")
        params.append(course_name)
    else:
        return "", ""
    row = conn.execute(
        f"""
        SELECT starts_at, exam_date, location
        FROM teacher_academic_invigilation_items
        WHERE {' AND '.join(where)}
        ORDER BY COALESCE(starts_at, exam_date, synced_at) DESC, id DESC
        LIMIT 1
        """,
        tuple(params),
    ).fetchone()
    if not row:
        return "", ""
    starts_at = _datetime_local(row["starts_at"])
    if not starts_at and row["exam_date"]:
        starts_at = f"{row['exam_date']}T08:00"
    return starts_at, _normalize_space(row["location"])


def _default_export_fields(
    conn: sqlite3.Connection,
    *,
    teacher_id: int,
    context: dict[str, Any],
    item: dict[str, Any] | None,
    students: list[dict[str, Any]],
) -> dict[str, Any]:
    item_payload = item or {}
    invigilation_datetime, invigilation_location = _invigilation_default(
        conn,
        teacher_id=teacher_id,
        item=item_payload,
    ) if item_payload else ("", "")
    session_datetime, session_location = _last_session_default(conn, int(context["class_offering_id"]))
    admin_class_name = _admin_class_name_from_students(
        students,
        fallback=_normalize_space(item_payload.get("class_composition")) or _normalize_space(context.get("class_name")),
    )
    return {
        "exam_datetime_local": invigilation_datetime or session_datetime,
        "exam_location": invigilation_location or session_location,
        "chief_invigilator": _normalize_space(context.get("teacher_name")),
        "assistant_invigilator": "",
        "admin_class_name": admin_class_name,
    }


def _status_payload(
    conn: sqlite3.Connection,
    *,
    teacher_id: int,
    context: dict[str, Any],
    item: sqlite3.Row | None,
) -> dict[str, Any]:
    if item is None:
        return {
            "status": "empty",
            "message": "尚未同步教务系统考试名单。",
            "class_offering_id": int(context["class_offering_id"]),
            "default_export": _default_export_fields(
                conn,
                teacher_id=teacher_id,
                context=context,
                item=None,
                students=[],
            ),
        }
    item_dict = dict(item)
    students = _load_exam_students(conn, int(item["id"]))
    alignment_students = [
        AcademicExamStudent(
            student_number=str(student.get("student_number") or ""),
            student_name=str(student.get("student_name") or ""),
            admin_class_name=str(student.get("admin_class_name") or ""),
        )
        for student in students
    ]
    alignment = _local_roster_alignment(
        conn,
        class_id=int(context["class_id"]),
        students=alignment_students,
    )
    return {
        "status": "success",
        "class_offering_id": int(context["class_offering_id"]),
        "exam_roster_item_id": int(item["id"]),
        "course": {
            "exam_course_key": item["exam_course_key"],
            "course_code": item["course_code"],
            "course_name": item["course_name"],
            "teaching_class_name": item["teaching_class_name"],
            "class_composition": item["class_composition"],
            "credits": item["credits"],
            "exam_method": item["exam_method"],
        },
        "student_count": len(students),
        "students_preview": students[:8],
        "alignment": alignment,
        "synced_at": item["synced_at"],
        "default_export": _default_export_fields(
            conn,
            teacher_id=teacher_id,
            context=context,
            item=item_dict,
            students=students,
        ),
    }


def load_classroom_exam_roster_status(teacher_id: int, class_offering_id: int) -> dict[str, Any]:
    with get_db_connection() as conn:
        context = _load_offering_context(conn, int(teacher_id), int(class_offering_id))
        if not context:
            return {"status": "not_found", "message": "课堂不存在或无权访问。"}
        item = _load_latest_exam_item(conn, teacher_id=int(teacher_id), class_offering_id=int(class_offering_id))
        return _status_payload(conn, teacher_id=int(teacher_id), context=context, item=item)


async def sync_classroom_exam_roster_from_academic_system(
    teacher_id: int,
    class_offering_id: int,
    *,
    exam_course_key: str = "",
) -> dict[str, Any]:
    with get_db_connection() as conn:
        access_payload = load_teacher_academic_access_method(conn, int(teacher_id), school_code=SCHOOL_CODE)
        context = _load_offering_context(conn, int(teacher_id), int(class_offering_id))
        semester = _load_semester_for_offering(conn, int(teacher_id), context) if context else None

    if not context:
        return {"status": "not_found", "message": "课堂不存在或无权访问。"}
    if not access_payload:
        return {
            "status": "missing_credential",
            "message": "请先在系统设置中配置并验证教务系统账号，再同步考试名单。",
        }
    if not semester:
        semester_result = await prepare_current_semester_from_academic_system(int(teacher_id))
        if semester_result.get("status") == "success":
            with get_db_connection() as conn:
                context = _load_offering_context(conn, int(teacher_id), int(class_offering_id)) or context
                semester = _load_semester_for_offering(conn, int(teacher_id), context)
        if not semester:
            return {
                "status": "no_semester",
                "message": "当前课堂没有可对齐的学期，请先为课堂选择学期或同步本学期校历。",
            }

    try:
        async with open_authenticated_academic_client(access_payload) as (client, profile, login_result):
            courses, source_summary, term_params = await _fetch_exam_courses(client, semester)
            selected_course, candidates, needs_confirmation = _select_exam_course(
                courses,
                context,
                requested_exam_course_key=_normalize_space(exam_course_key),
            )
            if needs_confirmation or selected_course is None:
                return {
                    "status": "needs_confirmation",
                    "message": "系统未能唯一确认本课堂对应的教务系统考试课程，请教师选择后继续同步。",
                    "semester_id": _optional_int(semester.get("id")),
                    "semester_name": str(semester.get("name") or ""),
                    "candidates": candidates,
                    "source_summary": source_summary,
                }
            selected_course.academic_year = selected_course.academic_year or (term_params or {}).get("xnm", "")
            selected_course.academic_term = selected_course.academic_term or (term_params or {}).get("xqm", "")
            students = await _fetch_exam_students(client, selected_course, source_summary)
    except (ValueError, httpx.HTTPError, json.JSONDecodeError) as exc:
        return {
            "status": "academic_query_failed",
            "message": f"教务系统登录或考试名单读取失败：{str(exc)[:180]}",
        }

    synced_at = _now_iso()
    with get_db_connection() as conn:
        try:
            alignment = _local_roster_alignment(conn, class_id=int(context["class_id"]), students=students)
            item_id = _upsert_exam_roster_item(
                conn,
                teacher_id=int(teacher_id),
                semester=semester,
                context=context,
                course=selected_course,
                students=students,
                synced_at=synced_at,
            )
            _replace_exam_roster_students(
                conn,
                teacher_id=int(teacher_id),
                semester=semester,
                context=context,
                exam_roster_item_id=item_id,
                course=selected_course,
                students=students,
                synced_at=synced_at,
            )
            conn.commit()
            item = _load_latest_exam_item(conn, teacher_id=int(teacher_id), class_offering_id=int(class_offering_id))
            status_payload = _status_payload(conn, teacher_id=int(teacher_id), context=context, item=item)
        except sqlite3.Error:
            conn.rollback()
            raise

    return {
        **status_payload,
        "status": "success",
        "message": (
            f"已从教务系统同步《{selected_course.course_name}》考试名单，"
            f"共 {len(students)} 名考生。"
        ),
        "alignment": alignment,
        "source_summary": source_summary,
    }


def _parse_export_datetime(value: Any) -> datetime | None:
    text = _normalize_space(value).replace("T", " ")
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%Y-%m-%d":
                return parsed.replace(hour=8, minute=0)
            return parsed
        except ValueError:
            continue
    return None


def _format_exam_datetime(value: datetime) -> str:
    return value.strftime("%Y年%m月%d日 %H:%M")


def _safe_filename(value: Any) -> str:
    normalized = re.sub(r"[\\/:*?\"<>|\r\n]+", "_", str(value or "")).strip("._ ")
    return normalized[:80] or "考试名单"


def _apply_cell_style(
    cell,
    *,
    font: Font,
    alignment: Alignment,
    border: Border | None = None,
    fill: PatternFill | None = None,
) -> None:
    cell.font = font
    cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill


def build_exam_roster_signature_workbook(
    teacher_id: int,
    class_offering_id: int,
    *,
    export_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload = export_payload or {}
    selected_place: dict[str, Any] | None = None
    selected_place_key = _normalize_space(
        payload.get("exam_location_place_key") or payload.get("exam_place_key")
    )
    selected_place_id = _normalize_space(
        payload.get("exam_location_place_id") or payload.get("exam_place_id")
    )
    with get_db_connection() as conn:
        context = _load_offering_context(conn, int(teacher_id), int(class_offering_id))
        if not context:
            raise ValueError("课堂不存在或无权访问。")
        item = _load_latest_exam_item(conn, teacher_id=int(teacher_id), class_offering_id=int(class_offering_id))
        if item is None:
            raise ValueError("请先从教务系统同步并确认考试名单，再导出签名表。")
        item_dict = dict(item)
        students = _load_exam_students(conn, int(item["id"]))
        if not students:
            raise ValueError("当前考试名单没有考生，暂不能导出签名表。")
        if len(students) > MAX_EXAM_ROSTER_STUDENTS:
            raise ValueError(f"当前模板最多容纳 {MAX_EXAM_ROSTER_STUDENTS} 名考生，请拆分考场后再导出。")
        defaults = _default_export_fields(
            conn,
            teacher_id=int(teacher_id),
            context=context,
            item=item_dict,
            students=students,
        )
        if selected_place_key or selected_place_id:
            selected_place = load_teacher_teaching_place_by_key(
                conn,
                int(teacher_id),
                place_key=selected_place_key,
                place_id=selected_place_id,
            )
            if selected_place is None:
                raise ValueError("所选考试教室不在当前本地教学场地中，请重新同步场地或重新选择教室。")

    exam_datetime = _parse_export_datetime(payload.get("exam_datetime") or defaults.get("exam_datetime_local"))
    if exam_datetime is None:
        raise ValueError("请先确认考试时间。")
    exam_location = _normalize_space(
        (selected_place or {}).get("display_name")
        or payload.get("exam_location")
        or defaults.get("exam_location")
    )
    if not exam_location:
        raise ValueError("请填写考试地点。")
    chief_invigilator = _normalize_space(payload.get("chief_invigilator") or defaults.get("chief_invigilator"))
    assistant_invigilator = _normalize_space(payload.get("assistant_invigilator") or defaults.get("assistant_invigilator"))
    admin_class_name = _normalize_space(payload.get("admin_class_name") or defaults.get("admin_class_name"))

    course_code = _normalize_space(item_dict.get("course_code"))
    course_name = _normalize_space(item_dict.get("course_name") or context.get("course_name"))
    credits = _parse_float(item_dict.get("credits")) or _parse_float(context.get("credits")) or 0.0
    semester_title = _format_semester_exam_title(item_dict, context)

    wb = Workbook()
    ws = wb.active
    ws.title = "考试名单"
    ws.sheet_view.showGridLines = False
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    widths = [12.18, 9.18, 4.45, 5.82, 10.18, 5.18, 12.18, 9.18, 4.45, 5.82, 10.18, 5.18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    heights = {1: 24.3, 2: 28.5, 7: 3.8}
    for row_index in range(1, EXAM_ROSTER_TOTAL_ROWS + 1):
        ws.row_dimensions[row_index].height = heights.get(row_index, 20 if 3 <= row_index <= 6 else 16.5)

    ws.merge_cells("A1:L1")
    ws.merge_cells("A2:L2")
    ws.merge_cells("A3:J3")
    ws.merge_cells("K3:L3")
    ws.merge_cells("A4:F5")
    ws.merge_cells("G4:J5")
    ws.merge_cells("K4:L5")
    ws.merge_cells("A6:C6")
    ws.merge_cells("D6:F6")
    ws.merge_cells("G6:L6")

    font_title = Font(name="宋体", size=14, bold=True)
    font_header = Font(name="宋体", size=10, bold=True)
    font_meta = Font(name="宋体", size=10)
    font_student = Font(name="宋体", size=9)
    font_seat = Font(name="宋体", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    no_border = Border()
    header_fill = PatternFill(fill_type="solid", fgColor="EDEDED")

    ws["A1"] = "广西外国语学院考场学生名册"
    _apply_cell_style(ws["A1"], font=font_title, alignment=center)
    ws["A2"] = semester_title
    _apply_cell_style(ws["A2"], font=font_header, alignment=center)
    ws["A3"] = f"课程：{course_code} {course_name}".strip()
    _apply_cell_style(ws["A3"], font=font_meta, alignment=left)
    ws["K3"] = f"学分：{credits:.1f}"
    _apply_cell_style(ws["K3"], font=font_meta, alignment=right)
    ws["A4"] = f"考试时间：{_format_exam_datetime(exam_datetime)}"
    _apply_cell_style(ws["A4"], font=font_meta, alignment=left)
    ws["G4"] = f"考试地点：{exam_location}"
    _apply_cell_style(ws["G4"], font=font_meta, alignment=left)
    ws["K4"] = f"考试人数：{len(students)}"
    _apply_cell_style(ws["K4"], font=font_meta, alignment=right)
    ws["A6"] = f"主监考：{chief_invigilator}"
    _apply_cell_style(ws["A6"], font=font_meta, alignment=left)
    ws["D6"] = f"辅监考：{assistant_invigilator}"
    _apply_cell_style(ws["D6"], font=font_meta, alignment=left)
    ws["G6"] = f"行政班名称：{admin_class_name}"
    _apply_cell_style(ws["G6"], font=font_meta, alignment=left)

    headers = ["学号", "姓名", "性别", "座位号", "签名", "备注"] * 2
    for column_index, label in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=column_index, value=label)
        _apply_cell_style(cell, font=font_student, alignment=center, border=border, fill=header_fill)

    for row_index in range(1, 8):
        for column_index in range(1, 13):
            cell = ws.cell(row=row_index, column=column_index)
            cell.border = no_border
            if not cell.font or cell.font.name != "宋体":
                cell.font = font_meta
            if not cell.alignment:
                cell.alignment = center

    for row_index in range(EXAM_ROSTER_TABLE_START_ROW, EXAM_ROSTER_TOTAL_ROWS + 1):
        for column_index in range(1, 13):
            cell = ws.cell(row=row_index, column=column_index)
            cell_font = (
                font_seat
                if row_index >= EXAM_ROSTER_FIRST_STUDENT_ROW and column_index in {4, 10}
                else font_student
            )
            _apply_cell_style(cell, font=cell_font, alignment=center, border=border)

    for index, student in enumerate(students):
        if index < EXAM_ROSTER_STUDENT_ROWS_PER_SIDE:
            row_index = EXAM_ROSTER_FIRST_STUDENT_ROW + index
            base_col = 1
        else:
            row_index = EXAM_ROSTER_FIRST_STUDENT_ROW + (index - EXAM_ROSTER_STUDENT_ROWS_PER_SIDE)
            base_col = 7
        values = [
            student.get("student_number") or "",
            student.get("student_name") or "",
            student.get("gender") or "",
            index + 1,
            "",
            "",
        ]
        for offset, value in enumerate(values):
            ws.cell(row=row_index, column=base_col + offset, value=value)

    output_dir = Path(ROSTER_DIR) / "exam_rosters"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = china_now().strftime("%Y%m%d%H%M%S")
    filename_stem = (
        f"4. 签到表-"
        f"{_safe_filename(admin_class_name or context.get('class_name') or '班级')}-"
        f"{_safe_filename(course_code or '课程编号')}-"
        f"{_safe_filename(course_name or '课程名称')}"
    )
    filename = f"{filename_stem}.xlsx"
    output_path = output_dir / f"{filename_stem}-{timestamp}.xlsx"
    wb.save(output_path)
    return {
        "path": output_path,
        "filename": filename,
        "student_count": len(students),
        "exam_roster_item_id": int(item_dict["id"]),
    }
