from __future__ import annotations

import io
import json
import math
import re
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from fastapi import HTTPException

from .exam_json_service import normalize_exam_scoring_payload
from .excel_upload_service import open_upload_workbook_pair
from .material_identity_service import (
    build_final_material_export_filename,
    period_label,
)
from .semester_identity_service import parse_semester_identity


EXAM_GRADE_RECORD_TYPE = "exam_grade_record"
EXAM_GRADE_RECORD_LABEL = "机试（作品设计）考核登分表"
EXAM_GRADE_RECORD_SCHEMA_VERSION = "gxufl-exam-grade-record-v2"
EXAM_GRADE_RECORD_TABLE_MODE = "single_continuous_roster_table"

EXAM_GRADE_LAYOUT = {
    "page": "A4 portrait",
    "layout_source": "gxufl_exam_grade_record_xlsx",
    "margins_in": {
        "left": 0.5118110236,
        "right": 0.5118110236,
        "top": 0.5511811024,
        "bottom": 0.7480314961,
        "header": 0.3149606299,
        "footer": 0.3149606299,
    },
    "row_heights": {"title": 35, "metadata": 32, "header": 15, "student": 18},
    "base_column_widths": {"index": 7.0, "student_number": 21.7265625, "student_name": 19.08984375, "total": 11.0},
    "sample_section_width": 46.0,
    "table_mode": EXAM_GRADE_RECORD_TABLE_MODE,
}

_CHINESE_ORDINALS = [
    "一",
    "二",
    "三",
    "四",
    "五",
    "六",
    "七",
    "八",
    "九",
    "十",
    "十一",
    "十二",
    "十三",
    "十四",
    "十五",
    "十六",
    "十七",
    "十八",
    "十九",
    "二十",
]


@dataclass(frozen=True)
class ExamGradeRecordParseResult:
    metadata: dict[str, Any]
    content_markdown: str
    tables: list[dict[str, Any]]
    warnings: list[str]
    export_payload: dict[str, Any]
    formula_count: int


def build_exam_grade_record_export_filename(
    fields: dict[str, Any] | None,
    *,
    suffix: str = ".xlsx",
) -> str:
    """导出文件名：``2025-2026-2《课程》机试（作品设计）考核登分表-班级.xlsx``。

    与平时成绩记录表保持同一套公文命名，归档时一眼能看出学年学期/课程/班级。
    """
    return build_final_material_export_filename(
        document_type_label=EXAM_GRADE_RECORD_LABEL,
        fields=fields,
        suffix=suffix,
    )


def normalize_exam_grade_record_payload(
    *,
    metadata: dict[str, Any] | None,
    content_markdown: str = "",
    tables: list[dict[str, Any]] | None = None,
    export_payload: dict[str, Any] | None = None,
    classroom_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    base = dict(export_payload or {})
    fields = _compact_dict(_as_dict(base.get("fields")))
    fields.update({key: value for key, value in _as_dict(metadata).items() if not _is_blank(value)})
    fields.update(
        {
            key: value
            for key, value in _fields_from_classroom_context(classroom_context or {}).items()
            if _is_blank(fields.get(key))
        }
    )
    fields.setdefault("school", "广西外国语学院")
    fields.setdefault("assessment_method", "机试（作品设计）")
    fields.setdefault("title", f"{fields.get('school') or '广西外国语学院'}机试（作品设计）考核登分表")
    # 每次都重算：学年/班级/课程可能在属性面板被改过，文件名要跟着走。
    fields["export_filename"] = build_exam_grade_record_export_filename(fields)

    structured = _as_dict(base.get("structured"))
    sections = _normalize_sections(structured.get("sections") or _sections_from_tables(tables or []))
    if not sections:
        total = _coerce_float(fields.get("total_score")) or 100
        sections = [{"index": 1, "label": "一", "title": "一", "full_score": _score_to_int(total), "questions": []}]
    students = _normalize_student_records(
        structured.get("students") if isinstance(structured.get("students"), list) else _students_from_tables(tables or []),
        section_count=len(sections),
    )
    total_score = sum(_score_to_int(section.get("full_score")) for section in sections)
    fields["total_score"] = total_score
    if not fields.get("class_size") and students:
        fields["class_size"] = len(students)

    warnings = _merge_warnings(base.get("warnings"), structured.get("warnings"))
    for student in students:
        section_scores = _student_section_scores(student, section_count=len(sections))
        if section_scores:
            calculated_total = sum(_score_to_int(value) for value in section_scores)
            student["calculated_total_score"] = calculated_total
            total_value = student.get("total_score")
            if total_value not in (None, "") and _score_to_int(total_value) != calculated_total:
                warnings.append(
                    f"{student.get('student_name') or student.get('student_number') or '学生'} 的大题合计 "
                    f"{calculated_total} 与总分 {_score_to_int(total_value)} 不一致，导出时以大题合计公式核验。"
                )
    if not students:
        warnings.append("未识别到学生成绩行，请检查源 Excel 或考试成绩是否已经发布。")

    normalized_structured = {
        **structured,
        "template_schema_version": EXAM_GRADE_RECORD_SCHEMA_VERSION,
        "table_mode": EXAM_GRADE_RECORD_TABLE_MODE,
        "ordering_source": str(structured.get("ordering_source") or "source_list_order"),
        "sections": sections,
        "students": students,
        "source_exam": _as_dict(structured.get("source_exam")),
        "warnings": _dedupe(warnings),
        "score_adjustment_policy": {
            "integer_scores": True,
            "target_total_source": "submissions.score",
            "deduction_distribution": "even_by_big_question_with_remainder_on_higher_score",
        },
    }
    base.update(
        {
            "document_group": "final_material",
            "document_type": EXAM_GRADE_RECORD_TYPE,
            "document_type_label": EXAM_GRADE_RECORD_LABEL,
            "template_key": EXAM_GRADE_RECORD_TYPE,
            "fields": fields,
            # 考核登分表的用户契约始终只有一张全班连续总表。内部评分核验明细留在
            # structured.students 中供后台追溯，不再作为并列表或隐藏工作表交给用户。
            "tables": [_table_from_students("考核登分表", sections, students)],
            "layout_profile": dict(EXAM_GRADE_LAYOUT),
            "structured": normalized_structured,
            "queryable_fields": _exam_grade_queryable_fields(fields, normalized_structured),
            # 预览同样来自规范化后的全量名单，避免旧材料沿用历史的 120 人截断内容。
            "content_markdown": _build_content_markdown(
                fields,
                sections,
                students,
                normalized_structured.get("source_exam") or {},
            ),
            "compatibility": {
                **_as_dict(base.get("compatibility")),
                "source_format_preserved": True,
                "layout_source": "gxufl_exam_grade_record_xlsx",
                "requires_template_confirmation": False,
                "template_schema_version": EXAM_GRADE_RECORD_SCHEMA_VERSION,
            },
        }
    )
    return base


def list_exam_grade_record_candidates(conn, *, class_offering_id: int, teacher_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT a.id,
               a.title,
               a.status,
               a.created_at,
               a.due_at,
               a.exam_paper_id,
               ep.title AS exam_paper_title,
               ep.questions_json,
               (
                   SELECT COUNT(*)
                   FROM students roster
                   WHERE (roster.class_id = o.class_id OR EXISTS (SELECT 1 FROM class_offering_class_links cocl_m WHERE cocl_m.offering_id = o.id AND cocl_m.class_id = roster.class_id))
                     AND COALESCE(roster.enrollment_status, 'active') = 'active'
               ) AS roster_count,
               COUNT(DISTINCT CASE WHEN scored_student.id IS NOT NULL THEN s.student_pk_id ELSE NULL END) AS submission_count,
               COUNT(DISTINCT CASE WHEN scored_student.id IS NOT NULL AND s.score IS NOT NULL THEN s.student_pk_id ELSE NULL END) AS graded_count,
               AVG(CASE WHEN scored_student.id IS NOT NULL AND s.score IS NOT NULL THEN s.score ELSE NULL END) AS average_score
        FROM assignments a
        JOIN class_offerings o ON o.id = a.class_offering_id
        JOIN exam_papers ep ON ep.id = a.exam_paper_id
        LEFT JOIN submissions s ON s.assignment_id = a.id
        LEFT JOIN students scored_student
               ON scored_student.id = s.student_pk_id
              AND (scored_student.class_id = o.class_id OR EXISTS (SELECT 1 FROM class_offering_class_links cocl_m WHERE cocl_m.offering_id = o.id AND cocl_m.class_id = scored_student.class_id))
              AND COALESCE(scored_student.enrollment_status, 'active') = 'active'
        WHERE a.class_offering_id = ?
          AND o.teacher_id = ?
          AND COALESCE(a.exam_paper_id, '') != ''
        GROUP BY a.id, a.title, a.status, a.created_at, a.due_at, a.exam_paper_id, ep.title, ep.questions_json, o.class_id
        ORDER BY COALESCE(a.due_at, a.created_at, '') DESC, a.id DESC
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchall()
    items = []
    for row in rows:
        item = dict(row)
        sections = _sections_from_exam_data(item.get("questions_json"))
        total_score = sum(_score_to_int(section.get("full_score")) for section in sections)
        average_score = item.get("average_score")
        roster_count = _coerce_int(item.get("roster_count"))
        submission_count = _coerce_int(item.get("submission_count"))
        graded_count = _coerce_int(item.get("graded_count"))
        blocking_reason = ""
        if not sections or total_score <= 0:
            blocking_reason = "试卷未配置可识别的大题和分值"
        elif roster_count <= 0:
            blocking_reason = "课堂暂无在读学生"
        elif graded_count <= 0:
            blocking_reason = "考试尚无已评分成绩"
        eligible = not blocking_reason
        items.append(
            {
                "id": int(item["id"]),
                "title": item.get("title") or f"考试 {item['id']}",
                "status": item.get("status") or "",
                "created_at": item.get("created_at") or "",
                "due_at": item.get("due_at") or "",
                "kind": "exam",
                "exam_paper_id": item.get("exam_paper_id") or "",
                "exam_paper_title": item.get("exam_paper_title") or "",
                "section_count": len(sections),
                "total_score": total_score,
                "roster_count": roster_count,
                "submission_count": submission_count,
                "graded_count": graded_count,
                "missing_grade_count": max(0, roster_count - graded_count),
                "coverage_percent": round(graded_count * 100 / roster_count, 1) if roster_count else 0,
                "average_score": round(float(average_score), 2) if average_score is not None else None,
                "eligible": eligible,
                "blocking_reason": blocking_reason,
            }
        )
    return items


def build_exam_grade_record_payload(
    conn,
    *,
    class_offering_id: int,
    teacher_id: int,
    exam_assignment_id: int | str,
    classroom_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    assignment_id = _coerce_int(exam_assignment_id)
    if assignment_id <= 0:
        raise HTTPException(400, "请选择一个课堂考试。")
    context = _load_context(conn, class_offering_id=int(class_offering_id), teacher_id=int(teacher_id), classroom_context=classroom_context)
    assignment = _load_exam_assignment(
        conn,
        class_offering_id=int(class_offering_id),
        teacher_id=int(teacher_id),
        assignment_id=assignment_id,
    )
    sections = _sections_from_exam_data(assignment.get("questions_json"))
    if not sections:
        raise HTTPException(422, "所选考试没有可识别的大题分值，无法生成考核登分表。")
    total_score = sum(_score_to_int(section.get("full_score")) for section in sections)
    students = _load_roster(conn, class_offering_id=int(class_offering_id))
    if not students:
        raise HTTPException(422, "当前课堂没有在读学生，无法生成考核登分表。")
    submissions = _load_exam_submissions(conn, assignment_id=assignment_id)

    from .classroom_retake_service import get_confirmed_retake_students

    confirmed_retake_map = {
        item["student_id"]: item
        for item in get_confirmed_retake_students(conn, class_offering_id=int(class_offering_id))
    }

    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    missing_grade_students: list[str] = []
    graded_student_count = 0
    for index, student in enumerate(students, start=1):
        student_id = int(student["student_id"])
        submission = submissions.get(student_id)
        row: dict[str, Any] = {
            "index": index,
            "row_order": index,
            "student_id": student_id,
            "student_number": student.get("student_number") or "",
            "student_name": student.get("student_name") or "",
            "section_scores": ["" for _ in sections],
            "raw_section_scores": ["" for _ in sections],
            "total_score": "",
            "raw_total_score": "",
            "adjustment_points": 0,
            "score_adjustment_reason": "",
            "source_question_scores": [],
        }
        if not submission or submission.get("score") in (None, ""):
            roster_retake = confirmed_retake_map.get(student_id)
            if roster_retake is not None:
                # 已确认的重修/插班学生未参加考试：按教师确认的默认分入库，
                # 大题按满分比例拆分并校验总分一致。
                default_total = min(
                    _round_int_score(roster_retake["default_ordinary_score"]),
                    total_score,
                )
                max_scores = [_score_to_int(section.get("full_score")) for section in sections]
                raw_section_scores = _clamp_section_scores(
                    _distribute_total_by_weights(default_total, max_scores),
                    max_scores,
                )
                section_scores = _reconcile_section_scores(
                    raw_section_scores,
                    target_total=default_total,
                    max_scores=max_scores,
                )
                if sum(section_scores) != default_total:
                    raise HTTPException(
                        500,
                        f"{row['student_name'] or row['student_number']} 的重修默认分拆分校验失败，请稍后重试。",
                    )
                row.update(
                    {
                        "section_scores": section_scores,
                        "raw_section_scores": section_scores,
                        "total_score": default_total,
                        "raw_total_score": default_total,
                        "adjustment_points": 0,
                        "score_adjustment_reason": "重修/插班学生未参加，按教师确认默认分记录",
                        "is_retake": True,
                        "retake": {
                            "mode": "roster_default",
                            "default_score": float(roster_retake["default_ordinary_score"]),
                        },
                    }
                )
                warnings.append(
                    f"{row['student_name'] or row['student_number']} 为已确认的重修/插班学生，"
                    f"未提交本场考试，按默认分 {default_total} 分入库。"
                )
                graded_student_count += 1
                rows.append(row)
                continue
            missing_grade_students.append(str(row["student_name"] or row["student_number"] or f"第 {index} 名学生"))
            rows.append(row)
            continue

        target_total = _round_int_score(submission.get("score"))
        if target_total < 0:
            warnings.append(
                f"{row['student_name'] or row['student_number']} 的最终分 {target_total} 小于 0，已按 0 分生成。"
            )
            target_total = 0
        if target_total > total_score:
            warnings.append(
                f"{row['student_name'] or row['student_number']} 的最终分 {target_total} 超过试卷满分 {total_score}，已按满分生成。"
            )
            target_total = total_score
        raw_total = _resolve_raw_total(submission, target_total=target_total)
        parsed_scores, source_question_scores = _section_scores_from_feedback(
            submission.get("feedback_md"),
            sections=sections,
            raw_total=raw_total,
        )
        if parsed_scores:
            raw_section_scores = [_score_to_int(value) for value in parsed_scores]
        else:
            raw_section_scores = _distribute_total_by_weights(raw_total, [_score_to_int(section.get("full_score")) for section in sections])
            warnings.append(f"{row['student_name'] or row['student_number']} 未解析到逐题得分，已按大题满分比例拆分卷面分。")
        raw_section_scores = _clamp_section_scores(raw_section_scores, [_score_to_int(section.get("full_score")) for section in sections])
        section_scores = _reconcile_section_scores(
            raw_section_scores,
            target_total=target_total,
            max_scores=[_score_to_int(section.get("full_score")) for section in sections],
        )
        if sum(section_scores) != target_total:
            raise HTTPException(500, f"{row['student_name'] or row['student_number']} 的总分校验失败，请稍后重试。")
        adjustment = sum(raw_section_scores) - target_total
        row.update(
            {
                "section_scores": section_scores,
                "raw_section_scores": raw_section_scores,
                "total_score": target_total,
                "raw_total_score": sum(raw_section_scores),
                "adjustment_points": adjustment,
                "score_adjustment_reason": _score_adjustment_reason(submission, adjustment),
                "source_question_scores": source_question_scores,
            }
        )
        graded_student_count += 1
        rows.append(row)

    if graded_student_count <= 0:
        raise HTTPException(422, "所选考试尚无已评分成绩，请完成评分后再生成考核登分表。")
    if missing_grade_students:
        preview = "、".join(missing_grade_students[:8])
        suffix = f"等 {len(missing_grade_students)} 人" if len(missing_grade_students) > 8 else f"共 {len(missing_grade_students)} 人"
        warnings.append(f"{preview}{suffix}尚无考试总分，登分表保留学生信息且成绩留空。")

    source_exam = {
        "assignment_id": int(assignment["id"]),
        "assignment_title": assignment.get("title") or "",
        "exam_paper_id": assignment.get("exam_paper_id") or "",
        "exam_paper_title": assignment.get("exam_paper_title") or "",
        "section_count": len(sections),
        "total_score": total_score,
    }
    fields = {
        **context,
        "title": f"{context.get('school') or '广西外国语学院'}机试（作品设计）考核登分表",
        "assessment_method": "机试（作品设计）",
        "class_size": len(students),
        "source_exam_title": source_exam["assignment_title"],
        "total_score": total_score,
    }
    payload = normalize_exam_grade_record_payload(
        metadata=fields,
        content_markdown="",
        tables=[],
        export_payload={
            "fields": fields,
            "structured": {
                "sections": sections,
                "students": rows,
                "source_exam": source_exam,
                "ordering_source": "active_class_roster.student_number_then_id",
                "warnings": _dedupe(warnings),
            },
        },
    )
    payload["content_markdown"] = _build_content_markdown(fields, sections, rows, source_exam)
    return payload


def parse_exam_grade_record_file(file_path: Path, original_name: str) -> ExamGradeRecordParseResult:
    with open_upload_workbook_pair(
        file_path,
        original_name,
        material_label="考核登分表",
    ) as (wb_formula, wb_values):
        ws_formula, ws_values, header_row, matching_sheet_count = _locate_exam_grade_worksheet(
            wb_formula,
            wb_values,
        )
        if ws_formula is None or ws_values is None or not header_row:
            raise HTTPException(422, "未识别到“考核登分表”的序号、学号、姓名和总分表头。")
        columns = _header_columns(ws_formula, header_row)
        sections, section_columns, section_warnings = _parse_excel_sections(
            ws_formula,
            ws_values,
            header_row,
            columns,
        )
        if not sections:
            raise HTTPException(422, "未识别到带有效满分的大题得分列。")
        metadata = _parse_excel_metadata(ws_formula, header_row)
        metadata["source_sheet"] = ws_formula.title
        metadata["source_filename"] = str(original_name or file_path.name).strip()
        metadata["total_score"] = sum(_score_to_int(section.get("full_score")) for section in sections)
        students, formula_count, student_warnings = _parse_excel_students(
            ws_formula,
            ws_values,
            header_row,
            columns,
            sections=sections,
            section_columns=section_columns,
        )
    if not students:
        raise HTTPException(422, "未识别到任何学生成绩行，请确认文件不是空白模板。")
    metadata["class_size"] = len(students)
    warnings = [*section_warnings, *student_warnings]
    if matching_sheet_count > 1:
        warnings.append(
            f"文件中有 {matching_sheet_count} 张工作表符合考核登分表结构，"
            f"已按工作簿顺序解析《{metadata.get('source_sheet') or ''}》。"
        )
    if formula_count <= 0:
        warnings.append("源 Excel 未识别到总分公式，导出时将按大题合计公式重新生成。")
    tables = [_table_from_students("考核登分表", sections, students)]
    export_payload = normalize_exam_grade_record_payload(
        metadata=metadata,
        content_markdown="",
        tables=tables,
        export_payload={
            "fields": metadata,
            "structured": {
                "sections": sections,
                "students": students,
                "source_formula_count": formula_count,
                "warnings": warnings,
            },
        },
    )
    tables = list(export_payload.get("tables") or [])
    normalized_structured = _as_dict(export_payload.get("structured"))
    normalized_warnings = _merge_warnings(warnings, normalized_structured.get("warnings"))
    normalized_structured["warnings"] = normalized_warnings
    export_payload["structured"] = normalized_structured
    export_payload["warnings"] = normalized_warnings
    content_markdown = _build_content_markdown(
        _as_dict(export_payload.get("fields")),
        _normalize_sections(normalized_structured.get("sections")),
        _normalize_student_records(
            normalized_structured.get("students"),
            section_count=len(normalized_structured.get("sections") or []),
        ),
        _as_dict(normalized_structured.get("source_exam")),
    )
    export_payload["content_markdown"] = content_markdown
    return ExamGradeRecordParseResult(
        metadata=metadata,
        content_markdown=content_markdown,
        tables=tables,
        warnings=normalized_warnings,
        export_payload=export_payload,
        formula_count=formula_count,
    )


def build_exam_grade_record_xlsx(payload: dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError as exc:
        raise RuntimeError(f"缺少 XLSX 导出依赖 openpyxl: {exc}") from exc

    export_payload = normalize_exam_grade_record_payload(
        metadata={},
        content_markdown=str(payload.get("content_markdown") or ""),
        tables=payload.get("tables") if isinstance(payload.get("tables"), list) else [],
        export_payload=_as_dict(payload.get("export_payload")) or payload,
    )
    fields = _as_dict(export_payload.get("fields"))
    structured = _as_dict(export_payload.get("structured"))
    sections = _normalize_sections(structured.get("sections"))
    students = _normalize_student_records(structured.get("students"), section_count=len(sections))
    if not sections:
        sections = [{"index": 1, "label": "一", "title": "一", "full_score": _score_to_int(fields.get("total_score") or 100), "questions": []}]
    section_count = len(sections)
    last_col = 3 + section_count + 1
    total_col = last_col

    wb = Workbook()
    ws = wb.active
    ws.title = "考核登分表"
    for named_style in wb._named_styles:
        if getattr(named_style, "name", "") == "Normal":
            named_style.font = Font(name="宋体", size=11)
            break

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    margins = EXAM_GRADE_LAYOUT["margins_in"]
    ws.page_margins.left = margins["left"]
    ws.page_margins.right = margins["right"]
    ws.page_margins.top = margins["top"]
    ws.page_margins.bottom = margins["bottom"]
    ws.page_margins.header = margins["header"]
    ws.page_margins.footer = margins["footer"]
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "D5"
    ws.print_title_rows = "1:4"
    ws.print_options.horizontalCentered = True
    ws.oddFooter.center.text = "第 &P 页 / 共 &N 页"
    ws.oddFooter.center.size = 9
    ws.oddFooter.center.font = "宋体"

    widths = _exam_grade_column_widths(section_count)
    for col_index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    title_font = Font(name="宋体", size=18, bold=True, charset=134, family=3)
    meta_font = Font(name="黑体", size=12, charset=134, family=3)
    header_font = Font(name="宋体", size=11, bold=True, charset=134, family=3)
    section_font = Font(name="宋体", size=12, bold=True, color="FF0000", charset=134, family=3)
    body_font = Font(name="宋体", size=11, charset=134, family=3)
    green_fill = PatternFill(fill_type="solid", fgColor="FF92D050")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    for col in (1, 2, 3):
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    ws.cell(1, 1, str(fields.get("title") or "广西外国语学院机试（作品设计）考核登分表"))
    ws.cell(1, 1).font = title_font
    ws.cell(1, 1).alignment = center
    ws.row_dimensions[1].height = EXAM_GRADE_LAYOUT["row_heights"]["title"]

    ws.cell(2, 1, _exam_grade_metadata_line(fields))
    ws.cell(2, 1).font = meta_font
    ws.cell(2, 1).alignment = left
    ws.cell(2, 1).border = bottom_border
    ws.row_dimensions[2].height = EXAM_GRADE_LAYOUT["row_heights"]["metadata"]

    header_values = {1: "序号", 2: "学号", 3: "姓名"}
    for col, value in header_values.items():
        cell = ws.cell(3, col, value)
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for offset, section in enumerate(sections, start=0):
        col = 4 + offset
        header = ws.cell(3, col, str(section.get("label") or _ordinal(offset + 1)))
        header.font = section_font
        header.alignment = center
        header.border = border
        score_cell = ws.cell(4, col, _score_to_int(section.get("full_score")))
        score_cell.font = body_font
        score_cell.alignment = center
        score_cell.border = border
        score_cell.fill = green_fill
        score_cell.number_format = "0"
        validation = DataValidation(
            type="whole",
            operator="between",
            formula1="0",
            formula2=str(_score_to_int(section.get("full_score"))),
            allow_blank=True,
        )
        validation.error = f"请输入 0 到 {_score_to_int(section.get('full_score'))} 之间的整数分数。"
        validation.errorTitle = "大题分数超出范围"
        validation.prompt = f"本大题满分 {_score_to_int(section.get('full_score'))} 分。"
        validation.promptTitle = "录入大题得分"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        validation.add(f"{get_column_letter(col)}5:{get_column_letter(col)}{max(5, 4 + len(students))}")

    total_header = ws.cell(3, total_col, "总分")
    total_header.font = header_font
    total_header.alignment = center
    total_header.border = border
    total_score_cell = ws.cell(4, total_col, sum(_score_to_int(section.get("full_score")) for section in sections))
    total_score_cell.font = body_font
    total_score_cell.alignment = center
    total_score_cell.border = border
    total_score_cell.number_format = "0"
    for row in (3, 4):
        ws.row_dimensions[row].height = EXAM_GRADE_LAYOUT["row_heights"]["header"]
        for col in range(1, last_col + 1):
            ws.cell(row, col).border = border

    data_start = 5
    row_count = max(len(students), 1)
    for local_index in range(row_count):
        row_number = data_start + local_index
        student = students[local_index] if local_index < len(students) else {}
        section_scores = _student_section_scores(student, section_count=section_count)
        values = [
            local_index + 1 if student else "",
            student.get("student_number") or "",
            student.get("student_name") or "",
            *section_scores,
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row_number, col_index, value)
            cell.font = body_font
            cell.alignment = center
            cell.border = border
            if col_index == 2:
                cell.number_format = "@"
            elif col_index >= 4:
                cell.number_format = "0"
        total_cell = ws.cell(row_number, total_col)
        total_cell.font = body_font
        total_cell.alignment = center
        total_cell.border = border
        total_cell.number_format = "0"
        if student and any(value not in (None, "") for value in section_scores):
            first_score_col = get_column_letter(4)
            last_score_col = get_column_letter(total_col - 1)
            total_cell.value = f"={first_score_col}{row_number}" if section_count == 1 else f"=SUM({first_score_col}{row_number}:{last_score_col}{row_number})"
        ws.row_dimensions[row_number].height = EXAM_GRADE_LAYOUT["row_heights"]["student"]

    last_row = data_start + row_count - 1
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"

    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _load_context(conn, *, class_offering_id: int, teacher_id: int, classroom_context: dict[str, Any] | None = None) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT o.id,
               o.semester,
               o.teacher_id,
               o.class_id,
               c.name AS course_name,
               c.total_hours AS course_hours,
               c.credits,
               c.college AS course_college,
               c.department AS course_department,
               cl.name AS class_name,
               cl.college AS class_college,
               cl.department AS class_department,
               t.name AS teacher_name,
               t.college AS teacher_college,
               t.department AS teacher_department
        FROM class_offerings o
        JOIN courses c ON c.id = o.course_id
        JOIN classes cl ON cl.id = o.class_id
        JOIN teachers t ON t.id = o.teacher_id
        WHERE o.id = ?
          AND o.teacher_id = ?
        LIMIT 1
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchone()
    if not row:
        raise HTTPException(404, "课堂不存在或您无权生成考核登分表。")
    data = dict(row)
    context = {
        **_fields_from_classroom_context(classroom_context or {}),
        "class_offering_id": int(class_offering_id),
        "school": "广西外国语学院",
        "college": data.get("course_college") or data.get("class_college") or data.get("teacher_college") or "",
        "department": data.get("course_department") or data.get("class_department") or data.get("teacher_department") or "",
        "course_name": data.get("course_name") or "",
        "course_hours": data.get("course_hours") or "",
        "credits": data.get("credits") or "",
        "teacher_name": data.get("teacher_name") or "",
        "class_name": data.get("class_name") or "",
        **_period_fields(data.get("semester")),
    }
    return _compact_dict(context)


def _load_exam_assignment(conn, *, class_offering_id: int, teacher_id: int, assignment_id: int) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT a.id,
               a.title,
               a.status,
               a.exam_paper_id,
               a.created_at,
               a.due_at,
               ep.title AS exam_paper_title,
               ep.questions_json
        FROM assignments a
        JOIN class_offerings o ON o.id = a.class_offering_id
        JOIN exam_papers ep ON ep.id = a.exam_paper_id
        WHERE a.id = ?
          AND a.class_offering_id = ?
          AND o.teacher_id = ?
          AND COALESCE(a.exam_paper_id, '') != ''
        LIMIT 1
        """,
        (int(assignment_id), int(class_offering_id), int(teacher_id)),
    ).fetchone()
    if not row:
        raise HTTPException(404, "所选考试不存在、未绑定试卷或您无权使用。")
    return dict(row)


def _load_roster(conn, *, class_offering_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT s.id,
               s.student_id_number,
               s.name
        FROM students s
        JOIN class_offerings o ON (s.class_id = o.class_id OR EXISTS (SELECT 1 FROM class_offering_class_links cocl_m WHERE cocl_m.offering_id = o.id AND cocl_m.class_id = s.class_id))
        WHERE o.id = ?
          AND COALESCE(s.enrollment_status, 'active') = 'active'
        ORDER BY s.student_id_number, s.id
        """,
        (int(class_offering_id),),
    ).fetchall()
    return [
        {
            "student_id": int(row["id"]),
            "student_number": str(row["student_id_number"] or ""),
            "student_name": str(row["name"] or ""),
        }
        for row in rows
    ]


def _load_exam_submissions(conn, *, assignment_id: int) -> dict[int, dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT s.id,
               s.assignment_id,
               s.student_pk_id,
               s.student_name,
               s.status,
               s.score,
               s.feedback_md,
               s.score_before_late_penalty,
               s.late_penalty_points,
               s.is_late_submission,
               s.late_by_seconds,
               s.late_score_cap_applied,
               gr.work_score,
               gr.peer_avg,
               gr.final_score AS group_final_score,
               gr.peer_review_count
        FROM submissions s
        LEFT JOIN group_assignment_member_results gr
               ON gr.assignment_id = CAST(s.assignment_id AS TEXT)
              AND gr.student_pk_id = s.student_pk_id
        WHERE s.assignment_id = ?
        ORDER BY s.id ASC
        """,
        (int(assignment_id),),
    ).fetchall()
    result: dict[int, dict[str, Any]] = {}
    for row in rows:
        item = dict(row)
        student_id = _coerce_int(item.get("student_pk_id"))
        if student_id > 0:
            result[student_id] = item
    return result


def _sections_from_exam_data(raw_value: Any) -> list[dict[str, Any]]:
    exam_data = _load_json_object(raw_value)
    if not exam_data:
        return []
    if isinstance(exam_data.get("questions"), dict) and isinstance(exam_data["questions"].get("pages"), list):
        exam_data = exam_data["questions"]
    elif "pages" not in exam_data and isinstance(exam_data.get("questions"), list):
        exam_data = {"pages": [{"name": "试卷题目", "questions": exam_data.get("questions") or []}], "grading": exam_data.get("grading") or {}}
    try:
        normalized = normalize_exam_scoring_payload(exam_data, require_complete=False)
    except Exception:
        normalized = exam_data if isinstance(exam_data, dict) else {}
    pages = normalized.get("pages") if isinstance(normalized.get("pages"), list) else []
    root_total = _coerce_float(_as_dict(normalized.get("grading")).get("total_score"))
    sections: list[dict[str, Any]] = []
    global_index = 1
    for page_index, page in enumerate(pages, start=1):
        if not isinstance(page, dict):
            continue
        questions = []
        section_total = 0.0
        for question_index, raw_question in enumerate(page.get("questions") or [], start=1):
            if not isinstance(raw_question, dict):
                continue
            points = _question_points(raw_question)
            section_total += points
            questions.append(
                {
                    "id": str(raw_question.get("id") or f"p{page_index}_q{question_index}"),
                    "index": question_index,
                    "global_index": global_index,
                    "points": _score_to_int(points),
                    "title": _compact_text(raw_question.get("text") or raw_question.get("title") or "", limit=80),
                }
            )
            global_index += 1
        if questions:
            sections.append(
                {
                    "index": len(sections) + 1,
                    "label": _ordinal(len(sections) + 1),
                    "title": str(page.get("name") or page.get("title") or _ordinal(len(sections) + 1)).strip(),
                    "full_score": _score_to_int(section_total),
                    "questions": questions,
                }
            )
    if sections and sum(_score_to_int(section.get("full_score")) for section in sections) <= 0 and root_total:
        allocated = _distribute_total_by_weights(_score_to_int(root_total), [1 for _ in sections])
        for section, full_score in zip(sections, allocated):
            section["full_score"] = full_score
    if not sections and root_total:
        sections.append({"index": 1, "label": "一", "title": "一", "full_score": _score_to_int(root_total), "questions": []})
    return sections


def _section_scores_from_feedback(
    feedback_md: Any,
    *,
    sections: list[dict[str, Any]],
    raw_total: int,
) -> tuple[list[int], list[dict[str, Any]]]:
    feedback_scores = _feedback_scores_by_question(feedback_md)
    if not feedback_scores:
        return [], []
    section_scores = [0.0 for _ in sections]
    section_known = [False for _ in sections]
    question_items: list[dict[str, Any]] = []
    for section_index, section in enumerate(sections):
        for question in section.get("questions") or []:
            match = _lookup_question_score(question, feedback_scores)
            if not match:
                continue
            score = _coerce_float(match.get("score"))
            if score is None:
                continue
            section_scores[section_index] += score
            section_known[section_index] = True
            question_items.append(
                {
                    "section_label": section.get("label") or "",
                    "question_id": question.get("id") or "",
                    "question_index": question.get("global_index") or question.get("index") or "",
                    "score": _score_to_int(score),
                    "max_score": _score_to_int(match.get("max_score") or question.get("points") or 0),
                }
            )
    if not any(section_known):
        return [], []
    known_total = sum(section_scores)
    if raw_total and raw_total > _round_int_score(known_total) and not all(section_known):
        remaining = max(0, raw_total - _round_int_score(known_total))
        weights = [
            0 if section_known[index] else _score_to_int(section.get("full_score"))
            for index, section in enumerate(sections)
        ]
        allocated = _distribute_total_by_weights(remaining, weights)
        section_scores = [section_scores[index] + allocated[index] for index in range(len(sections))]
    return [_score_to_int(value) for value in section_scores], question_items


def _feedback_scores_by_question(feedback_md: Any) -> dict[str, dict[str, Any]]:
    raw = str(feedback_md or "").replace("\r\n", "\n").replace("\r", "\n")
    if not raw.strip():
        return {}
    result: dict[str, dict[str, Any]] = {}
    current_key: str | None = None
    current_lines: list[str] = []

    def flush() -> None:
        nonlocal current_key, current_lines
        if not current_key:
            current_lines = []
            return
        score, max_score = _extract_score_pair("\n".join(current_lines))
        if score is not None:
            _store_feedback_score(result, current_key, score, max_score)
        current_lines = []

    for line in raw.split("\n"):
        heading = _feedback_question_heading(line)
        if heading:
            flush()
            current_key = heading
            current_lines = [line]
            continue
        if current_key:
            current_lines.append(line)
    flush()
    if result:
        return result
    for question_no, score, max_score in re.findall(
        r"(?:第\s*)?([0-9A-Za-z_-]+)\s*(?:题|question|q)?[^\n]{0,80}?(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)",
        raw,
        flags=re.I,
    ):
        _store_feedback_score(result, question_no, float(score), float(max_score))
    return result


def _feedback_question_heading(line: Any) -> str | None:
    text = str(line or "").strip()
    match = re.match(r"^#{1,6}\s*(?:第\s*)?([0-9A-Za-z_\-]+)\s*(?:题|question|q)?(?:\s|$|[：:.\-、])", text, flags=re.I)
    if match:
        return match.group(1).strip()
    return None


def _extract_score_pair(text: Any) -> tuple[float | None, float | None]:
    raw = str(text or "")
    pair = re.search(r"(-?\d+(?:\.\d+)?)\s*/\s*(-?\d+(?:\.\d+)?)", raw)
    if pair:
        return float(pair.group(1)), float(pair.group(2))
    number = re.search(r"(?:本题得分|得分|score)\s*[：:]\s*(-?\d+(?:\.\d+)?)", raw, flags=re.I)
    if number:
        return float(number.group(1)), None
    return None, None


def _store_feedback_score(result: dict[str, dict[str, Any]], raw_key: Any, score: float, max_score: float | None) -> None:
    for alias in _question_aliases(raw_key):
        result[alias] = {"score": score, "max_score": max_score}


def _lookup_question_score(question: dict[str, Any], feedback_scores: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    aliases: set[str] = set()
    for value in (question.get("id"), question.get("index"), question.get("global_index")):
        aliases.update(_question_aliases(value))
    for alias in aliases:
        if alias in feedback_scores:
            return feedback_scores[alias]
    return None


def _question_aliases(value: Any) -> set[str]:
    raw = str(value or "").strip()
    if not raw:
        return set()
    lower = raw.lower()
    aliases = {lower, lower.replace(" ", ""), lower.replace("_", "-"), lower.replace("-", "_")}
    match = re.search(r"(\d+)$", lower)
    if match:
        number = str(int(match.group(1)))
        aliases.update({number, f"q{number}", f"question{number}", f"第{number}题"})
    if lower.isdigit():
        number = str(int(lower))
        aliases.update({number, f"q{number}", f"question{number}", f"第{number}题"})
    return aliases


def _resolve_raw_total(submission: dict[str, Any], *, target_total: int) -> int:
    for key in ("score_before_late_penalty", "work_score"):
        value = submission.get(key)
        if value not in (None, ""):
            return _round_int_score(value)
    return _round_int_score(target_total)


def _score_adjustment_reason(submission: dict[str, Any], adjustment: int) -> str:
    reasons: list[str] = []
    late_penalty = _round_int_score(submission.get("late_penalty_points") or 0)
    if late_penalty > 0 or _coerce_int(submission.get("is_late_submission")):
        reasons.append(f"迟交扣 {late_penalty} 分")
    work_score = _coerce_float(submission.get("work_score"))
    final_score = _coerce_float(submission.get("score"))
    peer_avg = _coerce_float(submission.get("peer_avg"))
    if work_score is not None and final_score is not None and abs(final_score - work_score) >= 0.5:
        delta = _round_int_score(work_score) - _round_int_score(final_score)
        if delta > 0:
            reasons.append(f"小组互评折算扣 {delta} 分")
        elif delta < 0:
            reasons.append(f"小组互评折算加 {abs(delta)} 分")
        if peer_avg is not None:
            reasons.append(f"互评均分 {peer_avg:g}")
    if not reasons and adjustment > 0:
        reasons.append(f"最终总分较卷面分扣 {adjustment} 分")
    elif not reasons and adjustment < 0:
        reasons.append(f"最终总分较卷面分加 {abs(adjustment)} 分")
    return "；".join(reasons)


def _reconcile_section_scores(raw_scores: list[int], *, target_total: int, max_scores: list[int]) -> list[int]:
    scores = _clamp_section_scores(raw_scores, max_scores)
    target = max(0, min(_coerce_int(target_total), sum(max_scores)))
    diff = sum(scores) - target
    while diff > 0:
        room = [index for index, value in enumerate(scores) if value > 0]
        if not room:
            break
        base = max(1, diff // len(room))
        changed = 0
        for index in room:
            amount = min(base, scores[index], diff - changed)
            if amount <= 0:
                continue
            scores[index] -= amount
            changed += amount
            if changed >= diff:
                break
        diff -= changed
        if diff <= 0:
            break
        for index in sorted(room, key=lambda item: (scores[item], max_scores[item], item), reverse=True):
            if diff <= 0:
                break
            if scores[index] <= 0:
                continue
            scores[index] -= 1
            diff -= 1
    diff = target - sum(scores)
    while diff > 0:
        room = [index for index, value in enumerate(scores) if value < max_scores[index]]
        if not room:
            break
        base = max(1, diff // len(room))
        changed = 0
        for index in room:
            capacity = max_scores[index] - scores[index]
            amount = min(base, capacity, diff - changed)
            if amount <= 0:
                continue
            scores[index] += amount
            changed += amount
            if changed >= diff:
                break
        diff -= changed
        if diff <= 0:
            break
        for index in sorted(room, key=lambda item: (max_scores[item] - scores[item], -scores[item], -item), reverse=True):
            if diff <= 0:
                break
            if scores[index] >= max_scores[index]:
                continue
            scores[index] += 1
            diff -= 1
    return scores


def _distribute_total_by_weights(total: int | float, weights: list[int]) -> list[int]:
    target = _round_int_score(total)
    if not weights:
        return []
    clean_weights = [max(0, _coerce_int(value)) for value in weights]
    if sum(clean_weights) <= 0:
        clean_weights = [1 for _ in weights]
    raw_values = [Decimal(target) * Decimal(weight) / Decimal(sum(clean_weights)) for weight in clean_weights]
    floors = [int(value.to_integral_value(rounding="ROUND_FLOOR")) for value in raw_values]
    remainder = target - sum(floors)
    order = sorted(range(len(raw_values)), key=lambda index: (raw_values[index] - floors[index], clean_weights[index], index), reverse=True)
    for index in order[:remainder]:
        floors[index] += 1
    return floors


def _clamp_section_scores(scores: list[int], max_scores: list[int]) -> list[int]:
    result = []
    for index, max_score in enumerate(max_scores):
        value = scores[index] if index < len(scores) else 0
        result.append(max(0, min(_coerce_int(value), max(0, _coerce_int(max_score)))))
    return result


def _locate_exam_grade_worksheet(
    wb_formula: Any,
    wb_values: Any,
) -> tuple[Any | None, Any | None, int | None, int]:
    value_sheets = {worksheet.title: worksheet for worksheet in wb_values.worksheets}
    candidates: list[tuple[Any, Any, int]] = []
    for index, ws_formula in enumerate(wb_formula.worksheets):
        header_row = _find_exam_grade_header_row(ws_formula)
        if not header_row:
            continue
        ws_values = value_sheets.get(ws_formula.title)
        if ws_values is None and index < len(wb_values.worksheets):
            ws_values = wb_values.worksheets[index]
        if ws_values is not None:
            candidates.append((ws_formula, ws_values, header_row))
    if not candidates:
        return None, None, None, 0
    ws_formula, ws_values, header_row = candidates[0]
    return ws_formula, ws_values, header_row, len(candidates)


def _find_exam_grade_header_row(ws: Any) -> int | None:
    max_row = min(int(ws.max_row or 0), 80)
    max_column = min(int(ws.max_column or 0), 64)
    for row in range(1, max_row + 1):
        columns = _header_columns(ws, row, strict=False, max_column=max_column)
        if set(columns) == {"index", "student_number", "student_name", "total"}:
            if columns["index"] < columns["student_number"] < columns["student_name"] < columns["total"]:
                return row
    return None


def _header_columns(
    ws: Any,
    header_row: int,
    *,
    strict: bool = True,
    max_column: int | None = None,
) -> dict[str, int]:
    result: dict[str, int] = {}
    aliases = {
        "index": {"序号", "编号"},
        "student_number": {"学号", "学生学号"},
        "student_name": {"姓名", "学生姓名"},
        "total": {"总分", "总成绩", "合计"},
    }
    scan_columns = max_column if max_column is not None else int(ws.max_column or 0)
    for col in range(1, scan_columns + 1):
        value = _identity_cell_text(ws.cell(header_row, col).value)
        if value in aliases["index"]:
            result["index"] = col
        elif value in aliases["student_number"]:
            result["student_number"] = col
        elif value in aliases["student_name"]:
            result["student_name"] = col
        elif value in aliases["total"]:
            result["total"] = col
    missing = [key for key in ("index", "student_number", "student_name", "total") if key not in result]
    if strict and missing:
        raise HTTPException(422, f"考核登分表缺少表头：{', '.join(missing)}")
    return result


def _parse_excel_sections(
    ws_formula: Any,
    ws_values: Any,
    header_row: int,
    columns: dict[str, int],
) -> tuple[list[dict[str, Any]], list[int], list[str]]:
    sections: list[dict[str, Any]] = []
    section_columns: list[int] = []
    warnings: list[str] = []
    for col in range(columns["student_name"] + 1, columns["total"]):
        label = str(ws_formula.cell(header_row, col).value or "").strip()
        if not label:
            continue
        raw_full_score = ws_values.cell(header_row + 1, col).value
        if raw_full_score in (None, ""):
            raw_full_score = ws_formula.cell(header_row + 1, col).value
        numeric_full_score = _coerce_float(raw_full_score)
        if numeric_full_score is None or numeric_full_score <= 0:
            raise HTTPException(
                422,
                f"大题“{label}”缺少有效满分，请检查第 {header_row + 1} 行第 {col} 列。",
            )
        full_score = _score_to_int(numeric_full_score)
        if abs(numeric_full_score - full_score) > 1e-9:
            warnings.append(f"大题“{label}”满分 {numeric_full_score:g} 不是整数，已按 {full_score} 分解析。")
        section_columns.append(col)
        sections.append(
            {
                "index": len(sections) + 1,
                "label": label,
                "title": label,
                "full_score": _score_to_int(full_score),
                "questions": [],
            }
        )
    if not sections:
        return [], [], warnings

    section_total = sum(_score_to_int(section.get("full_score")) for section in sections)
    declared_total_raw = ws_values.cell(header_row + 1, columns["total"]).value
    if declared_total_raw in (None, ""):
        declared_total_raw = ws_formula.cell(header_row + 1, columns["total"]).value
    declared_total = _coerce_float(declared_total_raw)
    if declared_total is not None and _score_to_int(declared_total) != section_total:
        warnings.append(
            f"表头声明总分 {_score_to_int(declared_total)} 与大题满分合计 {section_total} 不一致，"
            "已保留大题分值并标记复核。"
        )
    if section_total != 100:
        warnings.append(f"大题满分合计为 {section_total}，不是学校模板通常要求的 100 分，请复核源文件。")
    return sections, section_columns, warnings


def _parse_excel_metadata(ws: Any, header_row: int) -> dict[str, Any]:
    title = str(ws.cell(1, 1).value or "").strip() or "广西外国语学院机试（作品设计）考核登分表"
    meta_lines = []
    for row in range(2, max(2, header_row)):
        values = [str(ws.cell(row, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        line = " ".join(value for value in values if value)
        if line:
            meta_lines.append(line)
    combined = "\n".join(meta_lines)
    fields: dict[str, Any] = {"title": title, "school": "广西外国语学院", "assessment_method": "机试（作品设计）"}
    course = re.search(r"课程[:：]\s*(?:\[([^\]]+)\])?\s*(.+?)(?:\s+专业年级班级[:：]|\s+授课|$)", combined)
    if course:
        if course.group(1):
            fields["course_code"] = course.group(1).strip()
        fields["course_name"] = course.group(2).strip()
    class_match = re.search(r"专业年级班级[:：]\s*(.+?)(?:\s+授课|\n|$)", combined)
    if class_match:
        fields["class_name"] = class_match.group(1).strip()
    teacher = re.search(r"(?:授课老师|授课教师|任课教师)[:：]\s*(.+?)(?:\s+|$)", combined)
    if teacher:
        fields["teacher_name"] = teacher.group(1).strip()
    semester_identity = parse_semester_identity(combined)
    if semester_identity is not None:
        fields["academic_year"] = f"{semester_identity.start_year}-{semester_identity.end_year}"
        term_labels = {1: "第一学期", 2: "第二学期", 3: "第三学期"}
        fields["semester"] = term_labels[semester_identity.term]
    return _compact_dict(fields)


def _parse_excel_students(
    ws_formula: Any,
    ws_values: Any,
    header_row: int,
    columns: dict[str, int],
    *,
    sections: list[dict[str, Any]],
    section_columns: list[int],
) -> tuple[list[dict[str, Any]], int, list[str]]:
    students: list[dict[str, Any]] = []
    formula_count = 0
    warnings: list[str] = []
    row = header_row + 2
    blank_streak = 0
    seen_student_numbers: dict[str, int] = {}
    while row <= ws_formula.max_row:
        number = _student_number_text(ws_formula.cell(row, columns["student_number"]))
        name = str(ws_formula.cell(row, columns["student_name"]).value or "").strip()
        if not number and not name:
            blank_streak += 1
            if blank_streak >= 3:
                break
            row += 1
            continue
        blank_streak = 0
        if not number:
            warnings.append(f"第 {row} 行缺少学号，已保留姓名“{name}”并标记复核。")
        elif number in seen_student_numbers:
            warnings.append(f"学号 {number} 在第 {seen_student_numbers[number]}、{row} 行重复，请复核学生名单。")
        else:
            seen_student_numbers[number] = row
        if not name:
            warnings.append(f"第 {row} 行学号 {number} 缺少姓名，已保留该成绩行并标记复核。")

        section_scores: list[int | str] = []
        for section, section_column in zip(sections, section_columns):
            raw_score = ws_values.cell(row, section_column).value
            formula_score = ws_formula.cell(row, section_column).value
            if raw_score in (None, "") and not str(formula_score or "").startswith("="):
                raw_score = formula_score
            numeric_score = _coerce_float(raw_score)
            if raw_score not in (None, "") and numeric_score is None:
                warnings.append(
                    f"第 {row} 行大题“{section.get('label') or ''}”成绩“{raw_score}”不是有效数值，已留空。"
                )
                section_scores.append("")
                continue
            score = _score_or_blank(numeric_score)
            section_scores.append(score)
            if numeric_score is None:
                continue
            rounded_score = _score_to_int(numeric_score)
            full_score = _score_to_int(section.get("full_score"))
            if abs(numeric_score - rounded_score) > 1e-9:
                warnings.append(
                    f"第 {row} 行大题“{section.get('label') or ''}”成绩 {numeric_score:g} "
                    f"不是整数，已按 {rounded_score} 分解析。"
                )
            if rounded_score < 0 or rounded_score > full_score:
                warnings.append(
                    f"第 {row} 行大题“{section.get('label') or ''}”成绩 {rounded_score} "
                    f"超出 0 至 {full_score} 分，请复核源文件。"
                )
        total_formula = str(ws_formula.cell(row, columns["total"]).value or "")
        if total_formula.startswith("="):
            formula_count += 1
            if not _formula_sums_section_range(
                total_formula,
                row=row,
                first_column=section_columns[0],
                last_column=section_columns[-1],
            ):
                warnings.append(f"第 {row} 行总分公式未覆盖全部大题列，导出时将按标准求和公式生成。")
        total_value = _score_or_blank(ws_values.cell(row, columns["total"]).value)
        calculated_total = (
            sum(_score_to_int(value) for value in section_scores if value != "")
            if any(value != "" for value in section_scores)
            else ""
        )
        if total_value == "" and any(value != "" for value in section_scores):
            total_value = calculated_total
            if not total_formula.startswith("="):
                warnings.append(f"第 {row} 行总分为空，已按大题得分合计。")
        elif calculated_total != "" and total_value != "" and _score_to_int(total_value) != calculated_total:
            warnings.append(
                f"第 {row} 行总分 {_score_to_int(total_value)} 与大题合计 {calculated_total} 不一致，"
                "已保留源值并标记复核。"
            )
        students.append(
            {
                "index": _coerce_int(ws_formula.cell(row, columns["index"]).value, default=len(students) + 1),
                "source_row": row,
                "student_number": number,
                "student_name": name,
                "section_scores": section_scores,
                "raw_section_scores": section_scores,
                "total_score": total_value,
                "source_formula": total_formula,
            }
        )
        row += 1
    return students, formula_count, warnings


def _identity_cell_text(value: Any) -> str:
    return re.sub(r"[\s\u3000]+", "", str(value or "")).strip()


def _student_number_text(cell: Any) -> str:
    value = cell.value
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)) and float(value).is_integer():
        text = str(int(value))
        number_format = str(getattr(cell, "number_format", "") or "").split(";", 1)[0].strip()
        if re.fullmatch(r"0+", number_format):
            text = text.zfill(len(number_format))
        return text
    return str(value).strip()


def _formula_sums_section_range(
    formula: str,
    *,
    row: int,
    first_column: int,
    last_column: int,
) -> bool:
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        return True
    expected = (
        f"SUM({get_column_letter(first_column)}{row}:"
        f"{get_column_letter(last_column)}{row})"
    )
    normalized = re.sub(r"[\s$]+", "", str(formula or "")).upper()
    return expected.upper() in normalized


def _sections_from_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    for table in tables or []:
        rows = table.get("rows") if isinstance(table, dict) else None
        if not isinstance(rows, list) or len(rows) < 2:
            continue
        header = [str(value or "").strip() for value in rows[0]]
        if not {"序号", "学号", "姓名", "总分"}.issubset(set(header)):
            continue
        name_index = header.index("姓名")
        total_index = header.index("总分")
        full_scores = rows[1] if len(rows) > 1 and isinstance(rows[1], list) else []
        sections = []
        for col in range(name_index + 1, total_index):
            label = header[col] if col < len(header) else ""
            if label:
                sections.append(
                    {
                        "index": len(sections) + 1,
                        "label": label,
                        "title": label,
                        "full_score": _score_to_int(full_scores[col] if col < len(full_scores) else 0),
                        "questions": [],
                    }
                )
        if sections:
            return sections
    return []


def _students_from_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    students = []
    for table in tables or []:
        rows = table.get("rows") if isinstance(table, dict) else None
        if not isinstance(rows, list) or len(rows) < 3:
            continue
        header = [str(value or "").strip() for value in rows[0]]
        if not {"序号", "学号", "姓名", "总分"}.issubset(set(header)):
            continue
        name_index = header.index("姓名")
        total_index = header.index("总分")
        for row in rows[2:]:
            if not isinstance(row, list):
                continue
            number = str(row[1] if len(row) > 1 else "").strip()
            name = str(row[2] if len(row) > 2 else "").strip()
            if not number and not name:
                continue
            students.append(
                {
                    "index": _coerce_int(row[0] if row else None, default=len(students) + 1),
                    "student_number": number,
                    "student_name": name,
                    "section_scores": [_score_or_blank(row[col] if col < len(row) else "") for col in range(name_index + 1, total_index)],
                    "total_score": _score_or_blank(row[total_index] if total_index < len(row) else ""),
                }
            )
    return students


def _table_from_students(title: str, sections: list[dict[str, Any]], students: list[dict[str, Any]]) -> dict[str, Any]:
    header = ["序号", "学号", "姓名", *[str(section.get("label") or _ordinal(index + 1)) for index, section in enumerate(sections)], "总分"]
    full_scores = ["", "", "", *[_score_to_int(section.get("full_score")) for section in sections], sum(_score_to_int(section.get("full_score")) for section in sections)]
    rows = [header, full_scores]
    for row_order, student in enumerate(students, start=1):
        section_scores = _student_section_scores(student, section_count=len(sections))
        rows.append(
            [
                row_order,
                student.get("student_number") or "",
                student.get("student_name") or "",
                *section_scores,
                student.get("total_score") if student.get("total_score") not in (None, "") else sum(_score_to_int(value) for value in section_scores if value != ""),
            ]
        )
    return {"title": title, "rows": rows}


def _normalize_sections(values: Any) -> list[dict[str, Any]]:
    if not isinstance(values, list):
        return []
    sections = []
    for index, item in enumerate(values, start=1):
        if not isinstance(item, dict):
            continue
        full_score = _score_to_int(item.get("full_score") or item.get("score") or item.get("max_score"))
        questions = item.get("questions") if isinstance(item.get("questions"), list) else []
        sections.append(
            {
                **item,
                "index": _coerce_int(item.get("index"), default=index),
                "label": str(item.get("label") or _ordinal(index)).strip(),
                "title": str(item.get("title") or item.get("name") or item.get("label") or _ordinal(index)).strip(),
                "full_score": full_score,
                "questions": [question for question in questions if isinstance(question, dict)],
            }
        )
    return sections


def _normalize_student_records(values: Any, *, section_count: int) -> list[dict[str, Any]]:
    if not isinstance(values, list):
        return []
    result = []
    for index, item in enumerate(values, start=1):
        if not isinstance(item, dict):
            continue
        student_number = str(item.get("student_number") or "").strip()
        student_name = str(item.get("student_name") or "").strip()
        if not student_number and not student_name:
            continue
        section_scores = _normalize_score_list(item.get("section_scores"), section_count=section_count)
        raw_section_scores = _normalize_score_list(item.get("raw_section_scores"), section_count=section_count)
        total_score = _score_or_blank(item.get("total_score"))
        if total_score == "" and section_scores and any(value != "" for value in section_scores):
            total_score = sum(_score_to_int(value) for value in section_scores if value != "")
        source_index = _coerce_int(item.get("source_index") or item.get("index"), default=index)
        result.append(
            {
                **item,
                "source_index": source_index,
                "index": index,
                "row_order": index,
                "student_number": student_number,
                "student_name": student_name,
                "section_scores": section_scores,
                "raw_section_scores": raw_section_scores,
                "total_score": total_score,
            }
        )
    return result


def _normalize_score_list(values: Any, *, section_count: int) -> list[Any]:
    raw_values = list(values) if isinstance(values, list) else []
    result = [_score_or_blank(value) for value in raw_values[:section_count]]
    result += [""] * max(0, section_count - len(result))
    return result


def _student_section_scores(student: dict[str, Any], *, section_count: int) -> list[Any]:
    scores = list(student.get("section_scores") or [])[:section_count]
    scores += [""] * max(0, section_count - len(scores))
    return scores


def _exam_grade_metadata_line(fields: dict[str, Any]) -> str:
    """表头第二行。学年学期必须出现——否则归档后无法区分同课程的不同学期。"""
    course_code = str(fields.get("course_code") or "").strip()
    course_name = str(fields.get("course_name") or "").strip()
    course = f"[{course_code}]{course_name}" if course_code else course_name
    class_name = str(fields.get("class_name") or "").strip()
    teacher = str(fields.get("teacher_name") or "").strip()
    period = period_label(fields)
    second_line = f"授课老师：{teacher}"
    if period:
        second_line = f"{second_line}    学年学期：{period}"
    return f"课程：{course}    专业年级班级：{class_name}\n{second_line}"


def _exam_grade_column_widths(section_count: int) -> list[float]:
    base = EXAM_GRADE_LAYOUT["base_column_widths"]
    if section_count <= 1:
        section_width = EXAM_GRADE_LAYOUT["sample_section_width"]
    else:
        section_width = max(7.5, min(23.0, EXAM_GRADE_LAYOUT["sample_section_width"] / section_count))
    return [base["index"], base["student_number"], base["student_name"], *[section_width for _ in range(section_count)], base["total"]]


def _exam_grade_queryable_fields(fields: dict[str, Any], structured: dict[str, Any]) -> dict[str, Any]:
    sections = structured.get("sections") or []
    students = structured.get("students") or []
    return {
        "course_name": fields.get("course_name") or "",
        "class_name": fields.get("class_name") or "",
        "teacher_name": fields.get("teacher_name") or "",
        "academic_year": fields.get("academic_year") or "",
        "semester": fields.get("semester") or "",
        "source_exam": structured.get("source_exam") or {},
        "student_count": len(students),
        "table_mode": structured.get("table_mode") or EXAM_GRADE_RECORD_TABLE_MODE,
        "ordering_source": structured.get("ordering_source") or "source_list_order",
        "section_count": len(sections),
        "section_labels": [section.get("label") for section in sections if isinstance(section, dict)],
        "total_score": fields.get("total_score") or "",
    }


def _build_content_markdown(
    fields: dict[str, Any],
    sections: list[dict[str, Any]],
    students: list[dict[str, Any]],
    source_exam: dict[str, Any],
) -> str:
    lines = [
        "# 广西外国语学院机试（作品设计）考核登分表",
        "",
        f"- 课程：{fields.get('course_name') or ''}",
        f"- 专业年级班级：{fields.get('class_name') or ''}",
        f"- 授课老师：{fields.get('teacher_name') or ''}",
        f"- 来源考试：{source_exam.get('assignment_title') or fields.get('source_exam_title') or ''}",
        f"- 大题：{'、'.join(str(section.get('label') or '') for section in sections)}",
        f"- 学生人数：{len(students)}",
        "",
        "| 序号 | 学号 | 姓名 | "
        + " | ".join(str(section.get("label") or "") for section in sections)
        + " | 总分 |",
        "| --- | --- | --- | " + " | ".join("---" for _ in sections) + " | --- |",
    ]
    for student in students:
        section_scores = _student_section_scores(student, section_count=len(sections))
        lines.append(
            "| "
            + " | ".join(
                [
                    str(student.get("index") or ""),
                    str(student.get("student_number") or ""),
                    str(student.get("student_name") or ""),
                    *[_cell_text(value) for value in section_scores],
                    _cell_text(student.get("total_score")),
                ]
            )
            + " |"
        )
    return "\n".join(lines)


def _question_points(question: dict[str, Any]) -> float:
    grading = question.get("grading") if isinstance(question.get("grading"), dict) else {}
    for source in (grading, question):
        for key in ("points", "score", "max_score", "full_score", "分值", "满分"):
            value = _coerce_float(source.get(key))
            if value is not None:
                return value
    return 0.0


def _load_json_object(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not isinstance(value, str) or not value.strip():
        return {}
    try:
        payload = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def _fields_from_classroom_context(context: dict[str, Any]) -> dict[str, Any]:
    allowed = {
        "school",
        "college",
        "department",
        "course_name",
        "course_code",
        "class_name",
        "teacher_name",
        "academic_year",
        "semester",
        "course_hours",
        "credits",
        "class_size",
    }
    return _compact_dict({key: context.get(key) for key in allowed})


def _period_fields(value: Any) -> dict[str, str]:
    text = str(value or "")
    fields: dict[str, str] = {}
    match = re.search(r"(20\d{2}).*?(20\d{2})", text)
    if match:
        fields["academic_year"] = f"{match.group(1)}-{match.group(2)}"
    if re.search(r"(?:^|[-_])1(?:$|[-_])|第一|一", text):
        fields["semester"] = "第一学期"
    elif re.search(r"(?:^|[-_])2(?:$|[-_])|第二|二", text):
        fields["semester"] = "第二学期"
    return fields


def _score_or_blank(value: Any) -> int | str:
    if value in (None, ""):
        return ""
    number = _coerce_float(value)
    if number is None:
        return ""
    return _score_to_int(number)


def _score_to_int(value: Any) -> int:
    return _round_int_score(value)


def _round_int_score(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        number = Decimal(str(value))
    except Exception:
        return 0
    if not number.is_finite():
        return 0
    return int(number.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _coerce_int(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(number) or math.isinf(number):
        return default
    return int(number)


def _coerce_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def _compact_dict(value: dict[str, Any]) -> dict[str, Any]:
    return {key: item for key, item in value.items() if not _is_blank(item)}


def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _is_blank(value: Any) -> bool:
    return value in (None, "", [], {})


def _merge_warnings(*items: Any) -> list[str]:
    warnings: list[str] = []
    for item in items:
        if isinstance(item, list):
            warnings.extend(str(value) for value in item if str(value or "").strip())
        elif item:
            warnings.append(str(item))
    return _dedupe(warnings)


def _dedupe(values: list[str]) -> list[str]:
    result = []
    seen = set()
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _compact_text(value: Any, *, limit: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) > limit:
        return text[:limit].rstrip()
    return text


def _cell_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _ordinal(index: int) -> str:
    if 1 <= index <= len(_CHINESE_ORDINALS):
        return _CHINESE_ORDINALS[index - 1]
    return str(index)
