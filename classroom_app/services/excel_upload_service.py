"""把上传的 Excel 材料变成 openpyxl 可靠加载的工作簿。

上传文件按内容哈希落盘（`data/files/<ab>/<cd>/<hash>`，没有扩展名），而
openpyxl 对“路径入参”是按文件名后缀判格式的，直接传存储路径会报
“openpyxl does not support file format”。这里统一按文件魔数识别真实格式，
再以 BytesIO 交给 openpyxl（file-like 入参不做后缀校验），同时兜住
“把 .xls/CSV/HTML 改名成 .xlsx”这类伪装文件。
"""

from __future__ import annotations

from contextlib import contextmanager
import io
import re
import tempfile
import zipfile
from pathlib import Path
from pathlib import PurePosixPath
from typing import Any, Iterator

from fastapi import HTTPException

from .libreoffice_service import convert_office_file

XLSX_ZIP_MAGIC = b"PK\x03\x04"
XLS_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
MAX_UPLOAD_WORKBOOK_BYTES = 50 * 1024 * 1024
MAX_UPLOAD_WORKBOOK_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_UPLOAD_WORKBOOK_ARCHIVE_ENTRIES = 10_000
MAX_UPLOAD_WORKBOOK_SHEETS = 128
MAX_UPLOAD_WORKSHEET_ROWS = 20_000
MAX_UPLOAD_WORKSHEET_COLUMNS = 256
_REQUIRED_XLSX_PARTS = {
    "[Content_Types].xml",
    "_rels/.rels",
    "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels",
}


def sniff_excel_kind(head: bytes) -> str:
    """按魔数返回 'xlsx' / 'xls' / 'unknown'，不信任文件名后缀。"""
    if head.startswith(XLSX_ZIP_MAGIC):
        return "xlsx"
    if head.startswith(XLS_OLE_MAGIC):
        return "xls"
    return "unknown"


def load_upload_workbook_bytes(file_path: Path, original_name: str, *, material_label: str) -> bytes:
    """读取上传的 Excel 文件并返回 .xlsx 字节流，旧版 .xls 自动转换。"""
    source = Path(file_path)
    display_name = str(original_name or source.name).strip() or source.name
    if not source.is_file():
        raise HTTPException(410, f"{material_label}源文件缓存已不存在，请重新上传《{display_name}》。")
    size = source.stat().st_size
    if size <= 0:
        raise HTTPException(422, f"{material_label}文件《{display_name}》是空文件，请检查后重新上传。")
    if size > MAX_UPLOAD_WORKBOOK_BYTES:
        raise HTTPException(413, f"{material_label}文件《{display_name}》超过 50MB，无法解析，请精简后重新上传。")

    data = source.read_bytes()
    kind = sniff_excel_kind(data[:8])
    if kind == "xlsx":
        _validate_xlsx_archive(data, display_name=display_name, material_label=material_label)
        return data
    if kind == "unknown":
        raise HTTPException(
            415,
            f"{material_label}文件《{display_name}》不是有效的 Excel 工作簿"
            "（可能是 CSV/网页表格被改了扩展名）。请先用 Excel 打开并另存为 .xlsx 后重新上传。",
        )

    try:
        # 全局文件仓库按哈希保存且没有扩展名。LibreOffice 会按输入文件后缀选择
        # 过滤器，直接把哈希路径交给它会复制成 input.bin，旧版 Excel 因而仍会
        # 转换失败。先写入隔离的 .xls 临时文件，只保留本次转换所需的格式提示。
        with tempfile.TemporaryDirectory(prefix="lanshare-excel-upload-") as temp_root:
            conversion_source = Path(temp_root) / "input.xls"
            conversion_source.write_bytes(data)
            converted = convert_office_file(conversion_source, "xlsx", timeout=120)
        output = bytes(converted.output_bytes or b"")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            422,
            f"{material_label}文件《{display_name}》可能是旧版 .xls 或加密工作簿，自动转换失败："
            f"{_safe_exception_text(exc)}。"
            "请用 Excel 另存为 .xlsx 后重新上传。",
        ) from exc
    if not output.startswith(XLSX_ZIP_MAGIC):
        raise HTTPException(
            422,
            f"{material_label}文件《{display_name}》的旧版 .xls 转换结果异常。请用 Excel 另存为 .xlsx 后重新上传。",
        )
    if len(output) > MAX_UPLOAD_WORKBOOK_BYTES:
        raise HTTPException(413, f"{material_label}文件《{display_name}》转换后超过 50MB，无法安全解析。")
    _validate_xlsx_archive(output, display_name=display_name, material_label=material_label)
    return output


def load_upload_workbook(
    file_path: Path,
    original_name: str,
    *,
    material_label: str,
    data_only: bool = False,
) -> Any:
    """加载一个受限的 openpyxl 只读工作簿，不依赖存储路径扩展名。"""
    data = load_upload_workbook_bytes(file_path, original_name, material_label=material_label)
    return _load_openpyxl_workbook(
        data,
        original_name=original_name,
        file_path=file_path,
        material_label=material_label,
        data_only=data_only,
    )


def load_upload_workbook_pair(
    file_path: Path,
    original_name: str,
    *,
    material_label: str,
) -> tuple[Any, Any]:
    """返回 (公式视图, 取值视图) 两个只读工作簿，供成绩表公式解析器使用。"""
    data = load_upload_workbook_bytes(file_path, original_name, material_label=material_label)
    wb_formula = None
    wb_values = None
    try:
        wb_formula = _load_openpyxl_workbook(
            data,
            original_name=original_name,
            file_path=file_path,
            material_label=material_label,
            data_only=False,
        )
        wb_values = _load_openpyxl_workbook(
            data,
            original_name=original_name,
            file_path=file_path,
            material_label=material_label,
            data_only=True,
        )
        return wb_formula, wb_values
    except Exception:
        _close_workbooks(wb_formula, wb_values)
        raise


@contextmanager
def open_upload_workbook(
    file_path: Path,
    original_name: str,
    *,
    material_label: str,
    data_only: bool = False,
) -> Iterator[Any]:
    workbook = load_upload_workbook(
        file_path,
        original_name,
        material_label=material_label,
        data_only=data_only,
    )
    try:
        yield workbook
    finally:
        _close_workbooks(workbook)


@contextmanager
def open_upload_workbook_pair(
    file_path: Path,
    original_name: str,
    *,
    material_label: str,
) -> Iterator[tuple[Any, Any]]:
    workbooks = load_upload_workbook_pair(
        file_path,
        original_name,
        material_label=material_label,
    )
    try:
        yield workbooks
    finally:
        _close_workbooks(*workbooks)


def _load_openpyxl_workbook(
    data: bytes,
    *,
    original_name: str,
    file_path: Path,
    material_label: str,
    data_only: bool,
) -> Any:
    import openpyxl

    workbook = None
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(data),
            data_only=bool(data_only),
            read_only=True,
            keep_links=False,
        )
        _validate_workbook_shape(workbook, material_label=material_label)
    except Exception as exc:
        _close_workbooks(workbook)
        if isinstance(exc, HTTPException):
            raise
        raise HTTPException(
            422,
            f"{material_label}文件《{str(original_name or '').strip() or Path(file_path).name}》无法按 Excel 打开："
            f"{_safe_exception_text(exc)}。请确认文件可以在 Excel 中正常打开后重新上传。",
        ) from exc
    if not workbook.worksheets:
        raise HTTPException(422, f"{material_label}文件中没有可解析的工作表。")
    return workbook


def _validate_xlsx_archive(data: bytes, *, display_name: str, material_label: str) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_UPLOAD_WORKBOOK_ARCHIVE_ENTRIES:
                raise HTTPException(
                    413,
                    f"{material_label}文件《{display_name}》内部文件数量异常，已停止解析。",
                )
            names: set[str] = set()
            total_uncompressed = 0
            for info in infos:
                archive_path = PurePosixPath(info.filename.replace("\\", "/"))
                normalized_name = str(archive_path)
                if archive_path.is_absolute() or ".." in archive_path.parts:
                    raise HTTPException(422, f"{material_label}文件《{display_name}》包含不安全的内部路径。")
                if info.flag_bits & 0x1:
                    raise HTTPException(
                        422,
                        f"{material_label}文件《{display_name}》已加密，服务器无法读取。请解除密码后重新上传。",
                    )
                names.add(normalized_name)
                total_uncompressed += max(0, int(info.file_size or 0))
                if total_uncompressed > MAX_UPLOAD_WORKBOOK_UNCOMPRESSED_BYTES:
                    raise HTTPException(
                        413,
                        f"{material_label}文件《{display_name}》解压后超过 250MB，已停止解析。",
                    )
            missing = sorted(_REQUIRED_XLSX_PARTS - names)
            if missing:
                raise HTTPException(
                    422,
                    f"{material_label}文件《{display_name}》缺少 Excel 必要结构，可能已损坏或只是改了扩展名。",
                )
            corrupt_part = archive.testzip()
            if corrupt_part:
                raise HTTPException(
                    422,
                    f"{material_label}文件《{display_name}》内部数据校验失败，可能已损坏。",
                )
    except HTTPException:
        raise
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise HTTPException(
            422,
            f"{material_label}文件《{display_name}》不是完整的 Excel 工作簿：{_safe_exception_text(exc)}。",
        ) from exc


def _validate_workbook_shape(workbook: Any, *, material_label: str) -> None:
    worksheets = list(workbook.worksheets or [])
    if len(worksheets) > MAX_UPLOAD_WORKBOOK_SHEETS:
        raise HTTPException(413, f"{material_label}包含超过 {MAX_UPLOAD_WORKBOOK_SHEETS} 张工作表，无法安全解析。")
    for worksheet in worksheets:
        if worksheet.max_row is None or worksheet.max_column is None:
            # 一些由教务系统/第三方组件生成的合法 OOXML 没有写 worksheet
            # dimension。只读模式不会自行推断，需要显式扫描一次，否则业务解析器
            # 会把 max_row/max_column 当成空表。
            worksheet.calculate_dimension(force=True)
        max_row = int(worksheet.max_row or 0)
        max_column = int(worksheet.max_column or 0)
        if max_row > MAX_UPLOAD_WORKSHEET_ROWS or max_column > MAX_UPLOAD_WORKSHEET_COLUMNS:
            raise HTTPException(
                413,
                f"{material_label}工作表《{worksheet.title}》使用范围过大"
                f"（{max_row} 行 × {max_column} 列），无法安全解析。",
            )


def _close_workbooks(*workbooks: Any) -> None:
    for workbook in workbooks:
        if workbook is None:
            continue
        try:
            workbook.close()
        except Exception:
            pass


def _safe_exception_text(exc: Exception) -> str:
    return re.sub(r"[\r\n\t]+", " ", str(exc).strip())[:120] or exc.__class__.__name__
