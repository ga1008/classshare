from __future__ import annotations

import hashlib
import io
import math
import random
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException

from .excel_upload_service import open_upload_workbook_pair


ORDINARY_GRADE_RECORD_TYPE = "ordinary_grade_record"
ORDINARY_GRADE_RECORD_LABEL = "学生平时成绩记录表"
ORDINARY_GRADE_RECORD_SCHEMA_VERSION = "gxufl-ordinary-grade-record-v1"
ORDINARY_GRADE_RECORD_FILE_SEQUENCE = 7
ORDINARY_GRADE_PAGE_STUDENT_CAPACITY = 25
ORDINARY_GRADE_LAST_PAGE_MIN_BLANK_ROWS = 2
ORDINARY_GRADE_ATTENDANCE_ELIGIBILITY_PERCENT = 70.0
ORDINARY_GRADE_DEFAULT_MINIMUM_SCORE = 60.0
ORDINARY_GRADE_SCORE_FLOOR_ALGORITHM_VERSION = "balanced-deterministic-v1"
_ORDINARY_ASSESSMENT_TITLE_PATTERN = re.compile(r"期末|期中|测评|测试|考试|考核|测验|验收|试炼")
_ORDINARY_HOMEWORK_TITLE_PATTERN = re.compile(r"作业|练习|实战|实验|项目|实践|任务")
ORDINARY_GRADE_ASSIGNMENT_KINDS = {"assignment", "exam"}

ORDINARY_GRADE_NOTES = [
    "注：",
    "1.该门课程成绩设置比例为：平时成绩占40%，期末考试成绩占60% 。同一门课程的多个平行教学班，课程成绩设置比例须一致。  ",
    "2.该表作为学生是否有考试资格的依据，由任课教师按班级本着客观公正、实事求是的原则执行。",
    "3.该表可为电子表格。平时成绩具体项目由系（教研室）根据课程要求设定，可由出勤情况、平时作业、课堂表现、课堂测验、课程论文、期中考试等项组成（同一门课程多个平行教学班的项目一致），其比例由任课教师设定，并给出具体得分依据（计算办法）。“出勤成绩”须附上“翻转校园”导出的数据原始表。如“翻转校园”上有其他成绩，可自行替换该表格的其他内容，如“课堂测试”、“讨论”等，只需附上原始表即可。",
]

ORDINARY_GRADE_LAYOUT = {
    "page": "A4 portrait",
    "students_per_page": ORDINARY_GRADE_PAGE_STUDENT_CAPACITY,
    "last_page_min_blank_rows": ORDINARY_GRADE_LAST_PAGE_MIN_BLANK_ROWS,
    "columns": ["序号", "学号", "姓名", "“翻转校园”记录", "作业1", "作业2", "作业3", "测评1", "出勤成绩", "作业成绩", "测评成绩", "平时成绩"],
    "column_widths": [5.49, 14.07, 9.66, 8.49, 5.49, 5.49, 5.49, 5.49, 6.83, 5.41, 5.57, 10.74],
    "margins_in": {"left": 0.3541666667, "right": 0.1576388889, "top": 0.1965277778, "bottom": 0.0, "header": 0.5118110236, "footer": 0.1181102362},
    "scale_percent": 100,
    "page_bottom_spacer_row_height": 53.5,
    "page_top_spacer_row_height": 40.0,
}


@dataclass(frozen=True)
class OrdinaryGradeParseResult:
    metadata: dict[str, Any]
    content_markdown: str
    tables: list[dict[str, Any]]
    warnings: list[str]
    export_payload: dict[str, Any]
    formula_count: int


def build_ordinary_grade_record_export_filename(
    fields: dict[str, Any] | None,
    *,
    suffix: str = ".xlsx",
) -> str:
    values = _as_dict(fields)
    academic_year = str(values.get("academic_year") or "").strip() or "未设置学年"
    semester_no = _semester_number(str(values.get("semester") or ""))
    period = f"{academic_year}-{semester_no}" if semester_no != "__" else academic_year
    course_name = str(values.get("course_name") or "").strip() or "未命名课程"
    class_name = str(values.get("class_name") or "").strip() or "未命名班级"
    normalized_suffix = str(suffix or ".xlsx").strip().lower()
    if not normalized_suffix.startswith("."):
        normalized_suffix = f".{normalized_suffix}"
    if normalized_suffix != ".xlsx":
        normalized_suffix = ".xlsx"
    stem = (
        f"{ORDINARY_GRADE_RECORD_FILE_SEQUENCE}. {period}"
        f"《{course_name}》学生平时成绩记录表-{class_name}"
    )
    safe_stem = re.sub(r'[\\/:*?"<>|]+', "-", stem)
    safe_stem = re.sub(r"\s+", " ", safe_stem).strip(" .")
    return f"{safe_stem}{normalized_suffix}"


def normalize_ordinary_grade_record_payload(
    *,
    metadata: dict[str, Any] | None,
    content_markdown: str = "",
    tables: list[dict[str, Any]] | None = None,
    export_payload: dict[str, Any] | None = None,
    classroom_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    base = dict(export_payload or {})
    fields = _compact_dict(_as_dict(base.get("fields")))
    fields.update({key: value for key, value in _as_dict(metadata).items() if not _is_blank(value)})
    fields.update({key: value for key, value in _fields_from_classroom_context(classroom_context or {}).items() if _is_blank(fields.get(key))})
    fields.setdefault("school", "广西外国语学院")
    fields.setdefault("title", "广西外国语学院学生平时成绩记录表")
    fields.setdefault("ordinary_score_percent", 40)
    fields.setdefault("final_exam_percent", 60)
    fields.setdefault("attendance_weight", 0.4)
    fields.setdefault("homework_weight", 0.3)
    fields.setdefault("assessment_weight", 0.3)
    fields["export_filename"] = build_ordinary_grade_record_export_filename(fields)

    structured = _as_dict(base.get("structured"))
    parsed_students = _students_from_tables(tables or [])
    students = _normalize_student_records(
        structured.get("students") if isinstance(structured.get("students"), list) else parsed_students
    )
    if not fields.get("class_size") and students:
        fields["class_size"] = len([row for row in students if str(row.get("student_number") or "").strip()])

    source_assignments = _as_dict(structured.get("source_assignments"))
    formula_templates = _as_dict(structured.get("formula_templates")) or _formula_templates()
    warnings = _merge_warnings(base.get("warnings"), structured.get("warnings"))
    if not students:
        warnings.append("未识别到学生成绩行，请检查源 Excel 是否为官方平时成绩记录表。")

    normalized_structured = {
        **structured,
        "template_schema_version": ORDINARY_GRADE_RECORD_SCHEMA_VERSION,
        "students_per_page": ORDINARY_GRADE_PAGE_STUDENT_CAPACITY,
        "students": students,
        "source_assignments": source_assignments,
        "formula_templates": formula_templates,
        "notes": list(ORDINARY_GRADE_NOTES),
        "warnings": warnings,
        "score_weights": {
            "attendance": 0.4,
            "homework": 0.3,
            "assessment": 0.3,
        },
    }
    base.update(
        {
            "document_group": "final_material",
            "document_type": ORDINARY_GRADE_RECORD_TYPE,
            "document_type_label": ORDINARY_GRADE_RECORD_LABEL,
            "template_key": ORDINARY_GRADE_RECORD_TYPE,
            "fields": fields,
            "tables": tables or base.get("tables") or [],
            "layout_profile": dict(ORDINARY_GRADE_LAYOUT),
            "structured": normalized_structured,
            "queryable_fields": _ordinary_grade_queryable_fields(fields, normalized_structured),
            "content_markdown": content_markdown or base.get("content_markdown") or _build_content_markdown(fields, students, source_assignments),
            "compatibility": {
                **_as_dict(base.get("compatibility")),
                "source_format_preserved": True,
                "layout_source": "gxufl_ordinary_grade_record_xls",
                "requires_template_confirmation": False,
                "template_schema_version": ORDINARY_GRADE_RECORD_SCHEMA_VERSION,
            },
        }
    )
    return base


def infer_ordinary_grade_assignment_kind(row: dict[str, Any]) -> str:
    """Infer the ordinary-grade purpose without applying a teacher override."""
    title = str(row.get("title") or "").strip()
    if _ORDINARY_ASSESSMENT_TITLE_PATTERN.search(title):
        return "exam"
    if _ORDINARY_HOMEWORK_TITLE_PATTERN.search(title):
        return "assignment"
    return "exam" if row.get("exam_paper_id") else "assignment"


def normalize_ordinary_grade_kind_override(value: Any) -> str:
    """Normalize an API value; an empty value or ``auto`` clears the override."""
    normalized = str(value or "").strip().lower()
    if normalized in {"", "auto"}:
        return ""
    if normalized not in ORDINARY_GRADE_ASSIGNMENT_KINDS:
        raise ValueError("平时成绩用途只能设置为“自动识别”“平时作业”或“测验”。")
    return normalized


def ordinary_grade_assignment_kind_info(row: dict[str, Any]) -> dict[str, Any]:
    auto_kind = infer_ordinary_grade_assignment_kind(row)
    stored_override = str(row.get("ordinary_grade_kind_override") or "").strip().lower()
    override = stored_override if stored_override in ORDINARY_GRADE_ASSIGNMENT_KINDS else ""
    effective_kind = override or auto_kind
    return {
        "kind": effective_kind,
        "ordinary_grade_kind": effective_kind,
        "ordinary_grade_auto_kind": auto_kind,
        "ordinary_grade_kind_override": override,
        "ordinary_grade_kind_source": "manual" if override else "auto",
        "ordinary_grade_kind_updated_at": row.get("ordinary_grade_kind_updated_at") or "",
        "ordinary_grade_kind_updated_by_teacher_id": (
            _coerce_int(row.get("ordinary_grade_kind_updated_by_teacher_id")) or None
        ),
    }


def classify_ordinary_grade_assignment(row: dict[str, Any]) -> str:
    """Classify only the ordinary-grade purpose, never the student's task format."""
    return str(ordinary_grade_assignment_kind_info(row)["kind"])


def list_ordinary_grade_assignment_candidates(conn, *, class_offering_id: int, teacher_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT a.id,
               a.title,
               a.status,
               a.created_at,
               a.due_at,
               a.exam_paper_id,
               a.ordinary_grade_kind_override,
               a.ordinary_grade_kind_updated_at,
               a.ordinary_grade_kind_updated_by_teacher_id,
               a.grading_mode,
               COUNT(s.id) AS submission_count,
               SUM(CASE WHEN s.score IS NOT NULL THEN 1 ELSE 0 END) AS graded_count,
               AVG(CASE WHEN s.score IS NOT NULL THEN s.score ELSE NULL END) AS average_score
        FROM assignments a
        JOIN class_offerings o ON o.id = a.class_offering_id
        LEFT JOIN submissions s ON s.assignment_id = a.id
        WHERE a.class_offering_id = ?
          AND o.teacher_id = ?
        GROUP BY a.id, a.title, a.status, a.created_at, a.due_at, a.exam_paper_id,
                 a.ordinary_grade_kind_override, a.ordinary_grade_kind_updated_at,
                 a.ordinary_grade_kind_updated_by_teacher_id, a.grading_mode
        ORDER BY COALESCE(a.due_at, a.created_at, '') ASC, a.id ASC
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchall()
    candidates = []
    for row in rows:
        item = dict(row)
        average_score = item.get("average_score")
        kind_info = ordinary_grade_assignment_kind_info(item)
        candidates.append(
            {
                "id": int(item["id"]),
                "title": item.get("title") or f"作业 {item['id']}",
                "status": item.get("status") or "",
                "created_at": item.get("created_at") or "",
                "due_at": item.get("due_at") or "",
                **kind_info,
                "submission_count": _coerce_int(item.get("submission_count")),
                "graded_count": _coerce_int(item.get("graded_count")),
                "average_score": round(float(average_score), 2) if average_score is not None else None,
            }
        )
    return candidates


def build_ordinary_grade_record_payload(
    conn,
    *,
    class_offering_id: int,
    teacher_id: int,
    homework_assignment_ids: list[int | str],
    assessment_assignment_id: int | str,
    classroom_context: dict[str, Any] | None = None,
    attendance_sync: dict[str, Any] | None = None,
    generation_requirements: str = "",
    minimum_ordinary_score_enabled: bool = True,
    minimum_ordinary_score: float = ORDINARY_GRADE_DEFAULT_MINIMUM_SCORE,
    retake_students: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    homework_ids, assessment_id = validate_ordinary_grade_sources(
        homework_assignment_ids=homework_assignment_ids,
        assessment_assignment_id=assessment_assignment_id,
    )
    retake_targets = normalize_retake_students(retake_students)
    context = _load_context(conn, class_offering_id=int(class_offering_id), teacher_id=int(teacher_id), classroom_context=classroom_context)
    assignments = _load_source_assignments(
        conn,
        class_offering_id=int(class_offering_id),
        teacher_id=int(teacher_id),
        assignment_ids=[*homework_ids, assessment_id],
    )
    missing = [str(item) for item in [*homework_ids, assessment_id] if int(item) not in assignments]
    if missing:
        raise HTTPException(400, f"所选作业/测评不属于当前课堂或无权使用：{', '.join(missing)}")
    mismatched_homework = [
        assignments[int(assignment_id)].get("title") or str(assignment_id)
        for assignment_id in homework_ids
        if classify_ordinary_grade_assignment(assignments[int(assignment_id)]) != "assignment"
    ]
    if mismatched_homework:
        raise HTTPException(
            400,
            f"以下来源当前被归类为测验，不能放入平时作业：{'、'.join(mismatched_homework)}。请刷新来源或先修改平时成绩用途。",
        )
    assessment = assignments[int(assessment_id)]
    if classify_ordinary_grade_assignment(assessment) != "exam":
        raise HTTPException(
            400,
            f"“{assessment.get('title') or assessment_id}”当前被归类为平时作业，不能作为测验。请刷新来源或先修改平时成绩用途。",
        )

    students = _load_roster(conn, class_offering_id=int(class_offering_id), context=context)
    attendance_scores = _load_attendance_scores(conn, class_offering_id=int(class_offering_id), teacher_id=int(teacher_id))
    score_map = _load_assignment_scores(conn, assignment_ids=[*homework_ids, assessment_id])

    from .classroom_retake_service import get_confirmed_retake_students

    confirmed_retake_map = {
        item["student_id"]: item
        for item in get_confirmed_retake_students(conn, class_offering_id=int(class_offering_id))
    }

    roster_numbers = {str(student.get("student_number") or "").strip() for student in students}
    unknown_retake_numbers = [number for number in retake_targets if number not in roster_numbers]
    if unknown_retake_numbers:
        raise HTTPException(
            400,
            f"以下重修学生学号不在本课堂名单中：{'、'.join(unknown_retake_numbers[:8])}。"
            "请重新核对后再生成。",
        )

    score_floor = _normalize_score_floor_policy(
        enabled=minimum_ordinary_score_enabled,
        minimum_score=minimum_ordinary_score,
    )
    score_floor.update(
        {
            "algorithm_version": ORDINARY_GRADE_SCORE_FLOOR_ALGORITHM_VERSION,
            "deterministic": True,
            "eligible_count": 0,
            "adjusted_count": 0,
            "ineligible_count": 0,
            "already_satisfied_count": 0,
            "capped_count": 0,
        }
    )
    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    retake_rows: list[dict[str, Any]] = []
    for index, student in enumerate(students, start=1):
        student_id = int(student["student_id"])
        student_number = str(student.get("student_number") or "").strip()
        raw_homework_scores = [score_map.get((int(assignment_id), student_id)) for assignment_id in homework_ids]
        raw_assessment_score = score_map.get((int(assessment_id), student_id))
        source_homework_scores = [_score_or_zero(value) for value in raw_homework_scores]
        source_assessment_score = _score_or_zero(raw_assessment_score)
        if student_number in retake_targets:
            # 重修/免修学生：不参加平时任务，平时分由教师直接设定，
            # 按公式反向分配到出勤、三次作业和测评，最终计算分恰为设定值。
            distribution = distribute_retake_ordinary_score(
                retake_targets[student_number],
                seed_parts=(
                    int(class_offering_id),
                    student_id,
                    *homework_ids,
                    assessment_id,
                ),
            )
            retake_row = {
                "index": index,
                "student_id": student_id,
                "student_number": student_number,
                "student_name": student.get("student_name") or "",
                "attendance_raw_score": distribution["attendance_score"],
                "source_homework_scores": source_homework_scores,
                "source_assessment_score": source_assessment_score,
                "source_score_missing": {
                    "homework": [value is None for value in raw_homework_scores],
                    "assessment": raw_assessment_score is None,
                },
                "homework_scores": distribution["homework_scores"],
                "assessment_score": distribution["assessment_score"],
                "is_retake": True,
                "retake": distribution,
                "calculated_scores": {"ordinary_score": distribution["target_score"]},
                "score_floor_adjustment": {
                    "enabled": bool(score_floor["enabled"]),
                    "eligible": False,
                    "applied": False,
                    "capped": False,
                    "reason": "retake_override",
                    "requested_minimum_score": float(score_floor["minimum_score"]),
                    "original_score": round(
                        calculate_ordinary_grade_score(
                            _score_or_zero(attendance_scores.get(student_id)),
                            source_homework_scores,
                            source_assessment_score,
                        ),
                        4,
                    ),
                    "achieved_score": distribution["achieved_score"],
                    "homework_scores": distribution["homework_scores"],
                    "assessment_score": distribution["assessment_score"],
                    "algorithm_version": distribution["algorithm_version"],
                    "seed_fingerprint": distribution["seed_fingerprint"],
                },
            }
            rows.append(retake_row)
            retake_rows.append(retake_row)
            warnings.append(
                f"{student.get('student_name') or student_number}（{student_number}）为重修/免修学生，"
                f"平时成绩由教师设定为 {_format_score(distribution['target_score'])} 分，"
                "已按公式分配到出勤、作业与测评，不参与最低分配平。"
            )
            continue
        roster_retake = confirmed_retake_map.get(student_id)
        if roster_retake is not None:
            # 名单处已确认的重修/插班学生：已提交按实际批改分数，未提交按
            # 教师确认的默认分；出勤豁免、按默认分记录。
            default_score = float(roster_retake["default_ordinary_score"])
            homework_values = [
                _score_or_zero(value) if value is not None else default_score
                for value in raw_homework_scores
            ]
            assessment_value = (
                _score_or_zero(raw_assessment_score)
                if raw_assessment_score is not None
                else default_score
            )
            computed_score = round(
                calculate_ordinary_grade_score(default_score, homework_values, assessment_value),
                2,
            )
            retake_row = {
                "index": index,
                "student_id": student_id,
                "student_number": student_number,
                "student_name": student.get("student_name") or "",
                "attendance_raw_score": default_score,
                "source_homework_scores": [_score_or_zero(value) for value in raw_homework_scores],
                "source_assessment_score": _score_or_zero(raw_assessment_score),
                "source_score_missing": {
                    "homework": [value is None for value in raw_homework_scores],
                    "assessment": raw_assessment_score is None,
                },
                "homework_scores": homework_values,
                "assessment_score": assessment_value,
                "is_retake": True,
                "retake": {
                    "mode": "roster_default",
                    "default_score": default_score,
                    "computed_score": computed_score,
                },
                "calculated_scores": {"ordinary_score": computed_score},
                "score_floor_adjustment": {
                    "enabled": bool(score_floor["enabled"]),
                    "eligible": False,
                    "applied": False,
                    "capped": False,
                    "reason": "retake_override",
                    "requested_minimum_score": float(score_floor["minimum_score"]),
                    "original_score": computed_score,
                    "achieved_score": computed_score,
                    "homework_scores": homework_values,
                    "assessment_score": assessment_value,
                    "algorithm_version": "retake-roster-v1",
                    "seed_fingerprint": "",
                },
            }
            rows.append(retake_row)
            retake_rows.append(retake_row)
            warnings.append(
                f"{student.get('student_name') or student_number}（{student_number}）为已确认的重修/插班学生："
                f"未参加的作业/测评按默认 {_format_score(default_score)} 分计入，出勤按默认分记录，"
                f"已参加的任务保留真实成绩，本次平时成绩 {_format_score(computed_score)} 分。"
            )
            continue
        if any(value is None for value in raw_homework_scores) or raw_assessment_score is None:
            warnings.append(f"{student.get('student_name') or student.get('student_number')} 存在未批改或缺失的作业/测评成绩，原始成绩已按 0 分计入。")
        attendance_score = attendance_scores.get(student_id)
        if attendance_score is None:
            warnings.append(f"{student.get('student_name') or student.get('student_number')} 暂无智慧课堂签到记录，出勤成绩已按 0 分计入。")
        attendance_raw_score = _score_or_zero(attendance_score)
        adjustment = apply_ordinary_grade_score_floor(
            attendance_score=attendance_raw_score,
            homework_scores=source_homework_scores,
            assessment_score=source_assessment_score,
            enabled=bool(score_floor["enabled"]),
            minimum_score=float(score_floor["minimum_score"]),
            seed_parts=(
                int(class_offering_id),
                student_id,
                *homework_ids,
                assessment_id,
            ),
        )
        if adjustment["eligible"]:
            score_floor["eligible_count"] += 1
        else:
            score_floor["ineligible_count"] += 1
        if adjustment["applied"]:
            score_floor["adjusted_count"] += 1
        elif adjustment["reason"] == "already_at_or_above_floor":
            score_floor["already_satisfied_count"] += 1
        if adjustment["capped"]:
            score_floor["capped_count"] += 1
            warnings.append(
                f"{student.get('student_name') or student.get('student_number')} 的出勤成绩为 "
                f"{_format_score(attendance_raw_score)}，在不调整出勤且作业/测评最高 100 分的前提下，"
                f"平时成绩最高只能达到 {_format_score(adjustment['achieved_score'])} 分，"
                f"未达到教师设置的 {_format_score(score_floor['minimum_score'])} 分。"
            )
        rows.append(
            {
                "index": index,
                "student_id": student_id,
                "student_number": student.get("student_number") or "",
                "student_name": student.get("student_name") or "",
                "attendance_raw_score": attendance_raw_score,
                "source_homework_scores": source_homework_scores,
                "source_assessment_score": source_assessment_score,
                "source_score_missing": {
                    "homework": [value is None for value in raw_homework_scores],
                    "assessment": raw_assessment_score is None,
                },
                "homework_scores": adjustment["homework_scores"],
                "assessment_score": adjustment["assessment_score"],
                "score_floor_adjustment": adjustment,
            }
        )

    source_assignments = {
        "homework_assignment_ids": homework_ids,
        "homework_assignments": [_assignment_summary(assignments[int(item)]) for item in homework_ids],
        "assessment_assignment_id": assessment_id,
        "assessment_assignment": _assignment_summary(assignments[int(assessment_id)]),
    }
    fields = {
        **context,
        "class_size": len(students),
        "source_homework_titles": "；".join(item["title"] for item in source_assignments["homework_assignments"]),
        "source_assessment_title": source_assignments["assessment_assignment"]["title"],
        "minimum_ordinary_score_enabled": bool(score_floor["enabled"]),
        "minimum_ordinary_score": score_floor["minimum_score"],
        "attendance_eligibility_percent": score_floor["attendance_eligibility_percent"],
    }
    if str(generation_requirements or "").strip():
        fields["generation_requirements"] = str(generation_requirements).strip()
    retake_policy = {
        "count": len(retake_rows),
        "students": [
            {
                "student_number": row["student_number"],
                "student_name": row["student_name"],
                "mode": row["retake"].get("mode") or "manual",
                "target_score": row["retake"].get(
                    "target_score",
                    row["retake"].get("default_score"),
                ),
            }
            for row in retake_rows
        ],
    }
    if retake_rows:
        fields["retake_student_count"] = len(retake_rows)
    payload = normalize_ordinary_grade_record_payload(
        metadata=fields,
        content_markdown="",
        tables=[],
        export_payload={
            "fields": fields,
            "structured": {
                "students": rows,
                "source_assignments": source_assignments,
                "attendance_sync": dict(attendance_sync or {}),
                "generation_requirements": str(generation_requirements or "").strip(),
                "score_floor_policy": score_floor,
                "retake_policy": retake_policy,
                "formula_templates": _formula_templates(),
                "warnings": _dedupe(warnings),
            },
        },
    )
    payload["content_markdown"] = _build_content_markdown(fields, rows, source_assignments)
    return payload


def validate_ordinary_grade_sources(
    *,
    homework_assignment_ids: list[int | str],
    assessment_assignment_id: int | str,
) -> tuple[list[int], int]:
    homework_ids = []
    for value in homework_assignment_ids or []:
        item = _coerce_int(value)
        if item > 0 and item not in homework_ids:
            homework_ids.append(item)
    assessment_id = _coerce_int(assessment_assignment_id)
    if len(homework_ids) != 3:
        raise HTTPException(400, "平时作业必须选择 3 份。")
    if assessment_id <= 0:
        raise HTTPException(400, "测评必须选择 1 份。")
    if assessment_id in homework_ids:
        raise HTTPException(400, "三次作业和一次测评不能重合。")
    return homework_ids, assessment_id


def parse_ordinary_grade_record_file(file_path: Path, original_name: str) -> OrdinaryGradeParseResult:
    with open_upload_workbook_pair(
        file_path,
        original_name,
        material_label="平时成绩记录表",
    ) as (wb_formula, wb_values):
        ws_formula, ws_values, starts, matching_sheet_count = _locate_ordinary_grade_worksheet(
            wb_formula,
            wb_values,
        )
        if ws_formula is None or ws_values is None or not starts:
            raise HTTPException(422, "未识别到“广西外国语学院学生平时成绩记录表”标题。")

        metadata = _parse_block_metadata(ws_formula, starts[0])
        metadata["source_sheet"] = ws_formula.title
        metadata["source_filename"] = str(original_name or file_path.name).strip()
        students: list[dict[str, Any]] = []
        tables: list[dict[str, Any]] = []
        warnings: list[str] = []
        formula_count = 0
        for block_index, start in enumerate(starts, start=1):
            block_rows, block_formula_count, block_warnings = _parse_block_students(
                ws_formula,
                ws_values,
                start,
            )
            formula_count += block_formula_count
            warnings.extend(block_warnings)
            students.extend(block_rows)
            tables.append(_table_from_students(f"第 {block_index} 版", block_rows))

    if matching_sheet_count > 1:
        warnings.append(
            f"文件中有 {matching_sheet_count} 张工作表符合平时成绩记录表结构，"
            f"已按工作簿顺序解析《{metadata.get('source_sheet') or ''}》。"
        )
    if formula_count <= 0:
        warnings.append("源 Excel 未识别到公式列，导出时将按官方公式重新生成。")
    export_payload = normalize_ordinary_grade_record_payload(
        metadata=metadata,
        content_markdown="",
        tables=tables,
        export_payload={
            "fields": metadata,
            "structured": {
                "students": students,
                "formula_templates": _formula_templates(),
                "source_formula_count": formula_count,
                "warnings": warnings,
            },
        },
    )
    content_markdown = _build_content_markdown(metadata, students, _as_dict(export_payload.get("structured")).get("source_assignments") or {})
    export_payload["content_markdown"] = content_markdown
    return OrdinaryGradeParseResult(
        metadata=metadata,
        content_markdown=content_markdown,
        tables=tables,
        warnings=_dedupe(warnings),
        export_payload=export_payload,
        formula_count=formula_count,
    )


def build_ordinary_grade_record_xlsx(payload: dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.pagebreak import Break
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError as exc:
        raise RuntimeError(f"缺少 XLSX 导出依赖 openpyxl: {exc}") from exc

    export_payload = normalize_ordinary_grade_record_payload(
        metadata={},
        content_markdown=str(payload.get("content_markdown") or ""),
        tables=payload.get("tables") if isinstance(payload.get("tables"), list) else [],
        export_payload=_as_dict(payload.get("export_payload")) or payload,
    )
    fields = _as_dict(export_payload.get("fields"))
    structured = _as_dict(export_payload.get("structured"))
    students = _normalize_student_records(structured.get("students") if isinstance(structured.get("students"), list) else [])

    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "平时成绩"
    # Excel converts character-based column widths to physical print widths using
    # the workbook default font. The official .xls template uses 宋体 12 here,
    # while individual table cells override their own 10/12/18 pt typography.
    # Keeping the same default is therefore essential for print-width parity.
    wb._fonts[0] = Font(name="宋体", size=12, charset=134)

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False, autoPageBreaks=False)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = int(ORDINARY_GRADE_LAYOUT["scale_percent"])
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    margins = ORDINARY_GRADE_LAYOUT["margins_in"]
    ws.page_margins.left = margins["left"]
    ws.page_margins.right = margins["right"]
    ws.page_margins.top = margins["top"]
    ws.page_margins.bottom = margins["bottom"]
    ws.page_margins.header = margins["header"]
    ws.page_margins.footer = margins["footer"]
    ws.sheet_view.showGridLines = True

    for col_index, width in enumerate(ORDINARY_GRADE_LAYOUT["column_widths"], start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = float(width)

    pages = _chunk_students(students)
    if not pages:
        pages = [[]]
    for page_index, chunk in enumerate(pages):
        start_row = 1 + page_index * 37
        _write_page(ws, fields, chunk, start_row=start_row, page_index=page_index, total_students=len(students))
        if page_index < len(pages) - 1:
            ws.row_dimensions[start_row + 35].height = float(ORDINARY_GRADE_LAYOUT["page_bottom_spacer_row_height"])
            ws.row_dimensions[start_row + 36].height = float(ORDINARY_GRADE_LAYOUT["page_top_spacer_row_height"])
            ws.row_breaks.append(Break(id=start_row + 35))

    last_row = _last_used_row_for_pages(pages)
    ws.print_area = f"A1:L{last_row}"
    score_floor_policy = _as_dict(structured.get("score_floor_policy"))
    if score_floor_policy:
        _write_score_floor_audit_sheet(wb, students, score_floor_policy)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_score_floor_audit_sheet(
    workbook: Any,
    students: list[dict[str, Any]],
    policy: dict[str, Any],
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet("最低分配平审计")
    ws.sheet_state = "hidden"
    headers = [
        "序号",
        "学号",
        "姓名",
        "出勤率",
        "达到70%资格",
        "原始缺失项",
        "原作业1",
        "原作业2",
        "原作业3",
        "原测评",
        "原平时分",
        "调整后作业1",
        "调整后作业2",
        "调整后作业3",
        "调整后测评",
        "调整后平时分",
        "教师最低分",
        "处理结果",
        "算法版本",
        "随机种子指纹",
    ]
    ws.append(headers)
    reason_labels = {
        "disabled": "未启用最低分",
        "attendance_below_threshold": "出勤率不足70%，不配平",
        "already_at_or_above_floor": "原平时分已达标",
        "adjusted_to_floor": "已配平到最低分",
        "capped_by_attendance": "受真实出勤限制，已配平到可达到的最高分",
        "retake_override": "重修/免修学生，平时分由教师设定并按公式分配",
    }
    for student in students:
        adjustment = _as_dict(student.get("score_floor_adjustment"))
        source_homework = list(student.get("source_homework_scores") or student.get("homework_scores") or [])[:3]
        source_homework += [0.0] * (3 - len(source_homework))
        adjusted_homework = list(student.get("homework_scores") or [])[:3]
        adjusted_homework += [0.0] * (3 - len(adjusted_homework))
        source_assessment = student.get("source_assessment_score", student.get("assessment_score", 0.0))
        missing = _as_dict(student.get("source_score_missing"))
        missing_items = [
            f"作业{index + 1}"
            for index, is_missing in enumerate(list(missing.get("homework") or [])[:3])
            if is_missing
        ]
        if missing.get("assessment"):
            missing_items.append("测评")
        original_score = adjustment.get("original_score")
        if original_score is None:
            original_score = calculate_ordinary_grade_score(
                _score_or_zero(student.get("attendance_raw_score")),
                [_score_or_zero(value) for value in source_homework],
                _score_or_zero(source_assessment),
            )
        achieved_score = adjustment.get("achieved_score")
        if achieved_score is None:
            achieved_score = calculate_ordinary_grade_score(
                _score_or_zero(student.get("attendance_raw_score")),
                [_score_or_zero(value) for value in adjusted_homework],
                _score_or_zero(student.get("assessment_score")),
            )
        ws.append(
            [
                student.get("index") or "",
                student.get("student_number") or "",
                student.get("student_name") or "",
                student.get("attendance_raw_score", ""),
                "是" if adjustment.get("eligible") else "否",
                "、".join(missing_items) or "无",
                *source_homework,
                source_assessment,
                round(float(original_score), 4),
                *adjusted_homework,
                student.get("assessment_score", ""),
                round(float(achieved_score), 4),
                policy.get("minimum_score", ""),
                reason_labels.get(str(adjustment.get("reason") or ""), str(adjustment.get("reason") or "")),
                adjustment.get("algorithm_version") or policy.get("algorithm_version") or "",
                adjustment.get("seed_fingerprint") or "",
            ]
        )
    header_fill = PatternFill("solid", fgColor="E0E7FF")
    for cell in ws[1]:
        cell.font = Font(name="宋体", size=10, bold=True, color="312E81")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="宋体", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [7, 16, 12, 10, 14, 18, 10, 10, 10, 10, 11, 12, 12, 12, 12, 13, 12, 34, 28, 20]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{max(1, ws.max_row)}"


def _write_page(ws: Any, fields: dict[str, Any], students: list[dict[str, Any]], *, start_row: int, page_index: int, total_students: int) -> None:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Alignment, Border, Font, Side

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    title_font = Font(name="宋体", size=18, bold=True)
    period_font = Font(name="宋体", size=12, bold=True)
    meta_font = Font(name="宋体", size=12)
    header_font = Font(name="宋体", size=10, bold=True)
    body_font = Font(name="宋体", size=10)
    student_font = Font(name="宋体", size=12)

    merges = [
        (0, 1, 0, 12),
        (1, 1, 1, 12),
        (2, 1, 2, 12),
        (3, 1, 3, 12),
        (4, 1, 4, 3),
        (4, 5, 4, 7),
        (4, 9, 5, 9),
        (4, 10, 5, 10),
        (4, 11, 5, 11),
        (4, 12, 5, 12),
    ]
    for row1, col1, row2, col2 in merges:
        ws.merge_cells(start_row=start_row + row1, start_column=col1, end_row=start_row + row2, end_column=col2)

    ws.cell(start_row, 1, "广西外国语学院学生平时成绩记录表")
    ws.cell(start_row, 1).font = title_font
    ws.cell(start_row, 1).alignment = center
    ws.row_dimensions[start_row].height = 34.5 if page_index == 0 else 29.5

    period_value: Any = _period_line(fields)
    period_match = re.fullmatch(r"（20 (\d{2}) — 20 (\d{2}) 学年度第 ([^ ]+) 学期）", str(period_value))
    if period_match:
        normal_period_font = InlineFont(rFont="宋体", sz=12, b=True)
        underlined_period_font = InlineFont(rFont="宋体", sz=12, b=True, u="single")
        period_value = CellRichText(
            TextBlock(normal_period_font, "（20"),
            TextBlock(underlined_period_font, f" {period_match.group(1)} "),
            TextBlock(normal_period_font, "— 20"),
            TextBlock(underlined_period_font, f" {period_match.group(2)} "),
            TextBlock(normal_period_font, "学年度第"),
            TextBlock(underlined_period_font, f" {period_match.group(3)} "),
            TextBlock(normal_period_font, "学期）"),
        )
    ws.cell(start_row + 1, 1, period_value)
    ws.cell(start_row + 1, 1).font = period_font
    ws.cell(start_row + 1, 1).alignment = center
    ws.row_dimensions[start_row + 1].height = 29.0 if page_index == 0 else 23.0

    ws.cell(start_row + 2, 1, _meta_line_one(fields))
    ws.cell(start_row + 2, 1).font = meta_font
    ws.cell(start_row + 2, 1).alignment = left
    ws.row_dimensions[start_row + 2].height = 26.0 if page_index == 0 else 28.0

    ws.cell(start_row + 3, 1, _meta_line_two(fields, total_students=total_students))
    ws.cell(start_row + 3, 1).font = meta_font
    ws.cell(start_row + 3, 1).alignment = left
    ws.row_dimensions[start_row + 3].height = 35.0 if page_index == 0 else 25.5

    header_values = {
        (4, 1): "学生信息",
        (4, 4): "出勤成绩",
        (4, 5): "平时作业",
        (4, 8): "测评",
        (4, 9): "出勤成绩",
        (4, 10): "作业成绩",
        (4, 11): "测评成绩",
        (4, 12): "平时\n成绩=出勤成绩40%+作业成绩30%+测评成绩30%",
        (5, 1): "序号",
        (5, 2): "学号",
        (5, 3): "姓名",
        (5, 4): "“翻转校园”记录",
        (5, 5): 1,
        (5, 6): 2,
        (5, 7): 3,
        (5, 8): 1,
    }
    for (row_offset, col), value in header_values.items():
        cell = ws.cell(start_row + row_offset, col, value)
        cell.font = header_font
        cell.alignment = center
    for row_offset, height in ((4, 29.4 if page_index == 0 else 19.75), (5, 39.5 if page_index == 0 else 50.5)):
        ws.row_dimensions[start_row + row_offset].height = height

    table_row_count = _table_row_count_for_page(students, is_last=True)
    data_start = start_row + 6
    for local_index in range(table_row_count):
        row_number = data_start + local_index
        student = students[local_index] if local_index < len(students) else {}
        values = _student_values_for_export(student)
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row_number, col_index, value)
            cell.font = student_font if col_index in {2, 3} else body_font
            cell.alignment = center
            cell.border = border
            if col_index == 2:
                cell.number_format = "@"
            elif col_index in {5, 6, 7, 8, 9, 10, 11}:
                cell.number_format = '0_);[RED]\\(0\\)'
            elif col_index == 12:
                cell.number_format = '0.00_);[RED]\\(0.00\\)'
        if student:
            row_number_text = str(row_number)
            ws.cell(row_number, 9, f"=D{row_number_text}")
            ws.cell(row_number, 10, f"=AVERAGE(E{row_number_text}:G{row_number_text})")
            ws.cell(row_number, 11, f"=H{row_number_text}")
            ws.cell(row_number, 12, f"=I{row_number_text}*0.4+J{row_number_text}*0.3+K{row_number_text}*0.3")
        ws.row_dimensions[row_number].height = 18.0

    ws.cell(start_row + 3, 1).border = Border(bottom=thin)
    for row in range(start_row + 4, data_start + table_row_count):
        for col in range(1, 13):
            ws.cell(row, col).border = border

    note_start = data_start + table_row_count
    for offset, note in enumerate(ORDINARY_GRADE_NOTES):
        row_number = note_start + offset
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=12)
        cell = ws.cell(row_number, 1)
        if offset == 3:
            try:
                cell.value = CellRichText(
                    TextBlock(InlineFont(rFont="宋体", sz=10, b=True), "3.该表可为电子表格。"),
                    TextBlock(InlineFont(rFont="宋体", sz=10), note.removeprefix("3.该表可为电子表格。")),
                )
            except Exception:
                cell.value = note
        else:
            cell.value = note
        cell.font = body_font
        cell.alignment = top_left if offset else left
        ws.row_dimensions[row_number].height = [15.0, 14.25, 14.25, 67.5][offset] if page_index == 0 else [15.0, 15.0, 15.0, 68.0][offset]


def _locate_ordinary_grade_worksheet(
    wb_formula: Any,
    wb_values: Any,
) -> tuple[Any | None, Any | None, list[int], int]:
    value_sheets = {worksheet.title: worksheet for worksheet in wb_values.worksheets}
    candidates: list[tuple[Any, Any, list[int]]] = []
    for index, ws_formula in enumerate(wb_formula.worksheets):
        starts = _find_record_block_starts(ws_formula)
        if not starts:
            continue
        ws_values = value_sheets.get(ws_formula.title)
        if ws_values is None and index < len(wb_values.worksheets):
            ws_values = wb_values.worksheets[index]
        if ws_values is not None:
            candidates.append((ws_formula, ws_values, starts))
    if not candidates:
        return None, None, [], 0
    ws_formula, ws_values, starts = candidates[0]
    return ws_formula, ws_values, starts, len(candidates)


def _find_record_block_starts(ws: Any) -> list[int]:
    starts = []
    for row in range(1, ws.max_row + 1):
        value = str(ws.cell(row, 1).value or "").strip()
        if "广西外国语学院学生平时成绩记录表" in value:
            starts.append(row)
    return starts


def _parse_block_metadata(ws: Any, start: int) -> dict[str, Any]:
    fields: dict[str, Any] = {"title": "广西外国语学院学生平时成绩记录表", "school": "广西外国语学院"}
    period = str(ws.cell(start + 1, 1).value or "")
    year_match = re.search(r"(20)\s*(\d{2}).*?(20)\s*(\d{2})", period)
    if year_match:
        fields["academic_year"] = f"{year_match.group(1)}{year_match.group(2)}-{year_match.group(3)}{year_match.group(4)}"
    semester_match = re.search(r"第\s*([一二三四五六七八九十\d]+)\s*学期", period)
    if semester_match:
        fields["semester"] = _semester_label(semester_match.group(1))
    line_one = str(ws.cell(start + 2, 1).value or "")
    line_two = str(ws.cell(start + 3, 1).value or "")
    patterns = {
        "college": r"二级学院（部）：\s*(.+?)\s+课程名称",
        "course_name": r"课程名称：\s*(.+?)\s+学时",
        "course_hours": r"学时\s*([0-9.]+)",
        "credits": r"学分：\s*([0-9.]+)",
        "teacher_name": r"任课教师：\s*(.+?)\s+专业年级班级",
        "class_name": r"专业年级班级：\s*(.+?)\s+班级人数",
        "class_size": r"班级人数：\s*([0-9]+)",
    }
    combined = f"{line_one} {line_two}"
    for key, pattern in patterns.items():
        match = re.search(pattern, combined)
        if match:
            fields[key] = match.group(1).strip()
    return fields


def _parse_block_students(ws_formula: Any, ws_values: Any, start: int) -> tuple[list[dict[str, Any]], int, list[str]]:
    students: list[dict[str, Any]] = []
    warnings: list[str] = []
    formula_count = 0
    row = start + 6
    while row <= ws_formula.max_row:
        first = ws_formula.cell(row, 1).value
        if str(first or "").strip().startswith("注"):
            break
        if _find_record_block_starts(ws_formula) and row != start and "广西外国语学院学生平时成绩记录表" in str(first or ""):
            break
        number = str(ws_formula.cell(row, 2).value or "").strip()
        name = str(ws_formula.cell(row, 3).value or "").strip()
        if number and name:
            formulas = {
                "attendance_score": str(ws_formula.cell(row, 9).value or ""),
                "homework_score": str(ws_formula.cell(row, 10).value or ""),
                "assessment_score": str(ws_formula.cell(row, 11).value or ""),
                "ordinary_score": str(ws_formula.cell(row, 12).value or ""),
            }
            formula_count += sum(1 for value in formulas.values() if value.startswith("="))
            if not _formulas_match_expected(row, formulas):
                warnings.append(f"第 {row} 行公式与官方模板不完全一致，导出时将按标准公式生成。")
            students.append(
                {
                    "index": _coerce_int(first, default=len(students) + 1),
                    "student_number": number,
                    "student_name": name,
                    "attendance_raw_score": _score_or_blank(ws_values.cell(row, 4).value),
                    "homework_scores": [
                        _score_or_blank(ws_values.cell(row, 5).value),
                        _score_or_blank(ws_values.cell(row, 6).value),
                        _score_or_blank(ws_values.cell(row, 7).value),
                    ],
                    "assessment_score": _score_or_blank(ws_values.cell(row, 8).value),
                    "calculated_scores": {
                        "attendance_score": _score_or_blank(ws_values.cell(row, 9).value),
                        "homework_score": _score_or_blank(ws_values.cell(row, 10).value),
                        "assessment_score": _score_or_blank(ws_values.cell(row, 11).value),
                        "ordinary_score": _score_or_blank(ws_values.cell(row, 12).value),
                    },
                    "source_formulas": formulas,
                }
            )
        row += 1
    return students, formula_count, warnings


def _load_context(conn, *, class_offering_id: int, teacher_id: int, classroom_context: dict[str, Any] | None = None) -> dict[str, Any]:
    row = conn.execute(
        """
        SELECT o.id,
               o.semester,
               o.teacher_id,
               o.class_id,
               c.name AS course_name,
               c.total_hours AS course_hours,
               c.credits,
               c.college AS course_college,
               c.department AS course_department,
               cl.name AS class_name,
               cl.college AS class_college,
               cl.department AS class_department,
               t.name AS teacher_name,
               t.college AS teacher_college,
               t.department AS teacher_department
        FROM class_offerings o
        JOIN courses c ON c.id = o.course_id
        JOIN classes cl ON cl.id = o.class_id
        JOIN teachers t ON t.id = o.teacher_id
        WHERE o.id = ?
          AND o.teacher_id = ?
        LIMIT 1
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchone()
    if not row:
        raise HTTPException(404, "课堂不存在或您无权生成平时成绩记录表。")
    data = dict(row)
    context = {
        **_fields_from_classroom_context(classroom_context or {}),
        "class_offering_id": int(class_offering_id),
        "school": "广西外国语学院",
        "college": data.get("course_college") or data.get("class_college") or data.get("teacher_college") or "",
        "department": data.get("course_department") or data.get("class_department") or data.get("teacher_department") or "",
        "course_name": data.get("course_name") or "",
        "course_hours": data.get("course_hours") or "",
        "credits": data.get("credits") or "",
        "teacher_name": data.get("teacher_name") or "",
        "class_name": data.get("class_name") or "",
        **_period_fields(data.get("semester")),
    }
    return _compact_dict(context)


def _load_source_assignments(conn, *, class_offering_id: int, teacher_id: int, assignment_ids: list[int]) -> dict[int, dict[str, Any]]:
    placeholders = ",".join("?" for _ in assignment_ids)
    rows = conn.execute(
        f"""
        SELECT a.id,
               a.title,
               a.status,
               a.exam_paper_id,
               a.ordinary_grade_kind_override,
               a.ordinary_grade_kind_updated_at,
               a.ordinary_grade_kind_updated_by_teacher_id,
               a.created_at,
               a.due_at
        FROM assignments a
        JOIN class_offerings o ON o.id = a.class_offering_id
        WHERE a.id IN ({placeholders})
          AND a.class_offering_id = ?
          AND o.teacher_id = ?
        """,
        (*[int(item) for item in assignment_ids], int(class_offering_id), int(teacher_id)),
    ).fetchall()
    return {int(row["id"]): dict(row) for row in rows}


def _load_roster(conn, *, class_offering_id: int, context: dict[str, Any]) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT s.id,
               s.student_id_number,
               s.name
        FROM students s
        JOIN class_offerings o ON o.class_id = s.class_id
        WHERE o.id = ?
          AND COALESCE(s.enrollment_status, 'active') = 'active'
        ORDER BY s.student_id_number, s.id
        """,
        (int(class_offering_id),),
    ).fetchall()
    return [
        {
            "student_id": int(row["id"]),
            "student_number": str(row["student_id_number"] or ""),
            "student_name": str(row["name"] or ""),
        }
        for row in rows
    ]


def _load_attendance_scores(conn, *, class_offering_id: int, teacher_id: int) -> dict[int, float]:
    session_rows = conn.execute(
        """
        SELECT *
        FROM smart_classroom_checkin_sessions
        WHERE class_offering_id = ?
          AND teacher_id = ?
          AND session_id IS NOT NULL
        ORDER BY session_id, COALESCE(checkin_time, '') DESC, COALESCE(synced_at, '') DESC, id DESC
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchall()
    latest_by_session: dict[int, dict[str, Any]] = {}
    for row in session_rows:
        item = dict(row)
        session_id = _coerce_int(item.get("session_id"))
        if session_id > 0 and session_id not in latest_by_session:
            latest_by_session[session_id] = item
    checkin_ids = [int(item["id"]) for item in latest_by_session.values()]
    if not checkin_ids:
        return {}
    placeholders = ",".join("?" for _ in checkin_ids)
    rows = conn.execute(
        f"""
        SELECT checkin_session_id, student_id, status
        FROM smart_classroom_checkin_students
        WHERE checkin_session_id IN ({placeholders})
          AND student_id IS NOT NULL
        """,
        checkin_ids,
    ).fetchall()
    statuses_by_student: dict[int, dict[int, str]] = {}
    for row in rows:
        student_id = _coerce_int(row["student_id"])
        checkin_session_id = _coerce_int(row["checkin_session_id"])
        if student_id <= 0 or checkin_session_id <= 0:
            continue
        statuses_by_student.setdefault(student_id, {})[checkin_session_id] = str(row["status"] or "").strip().upper()
    scores = {}
    total_sessions = len(checkin_ids)
    for student_id, statuses in statuses_by_student.items():
        checked = sum(1 for status in statuses.values() if status == "CHECKED")
        scores[student_id] = round(checked * 100.0 / total_sessions, 2)
    return scores


def _load_assignment_scores(conn, *, assignment_ids: list[int]) -> dict[tuple[int, int], float]:
    placeholders = ",".join("?" for _ in assignment_ids)
    rows = conn.execute(
        f"""
        SELECT assignment_id, student_pk_id, score
        FROM submissions
        WHERE assignment_id IN ({placeholders})
          AND score IS NOT NULL
        """,
        [int(item) for item in assignment_ids],
    ).fetchall()
    result = {}
    for row in rows:
        assignment_id = _coerce_int(row["assignment_id"])
        student_id = _coerce_int(row["student_pk_id"])
        if assignment_id > 0 and student_id > 0:
            try:
                result[(assignment_id, student_id)] = float(row["score"])
            except (TypeError, ValueError):
                pass
    return result


def _student_values_for_export(student: dict[str, Any]) -> list[Any]:
    if not student:
        return [""] * 12
    homework = list(student.get("homework_scores") or [])[:3]
    homework += [""] * (3 - len(homework))
    return [
        student.get("index") or "",
        student.get("student_number") or "",
        student.get("student_name") or "",
        student.get("attendance_raw_score", ""),
        *homework,
        student.get("assessment_score", ""),
        "",
        "",
        "",
        "",
    ]


def _table_row_count_for_page(students: list[dict[str, Any]], *, is_last: bool) -> int:
    if len(students) >= ORDINARY_GRADE_PAGE_STUDENT_CAPACITY:
        return ORDINARY_GRADE_PAGE_STUDENT_CAPACITY
    return max(len(students) + ORDINARY_GRADE_LAST_PAGE_MIN_BLANK_ROWS, ORDINARY_GRADE_LAST_PAGE_MIN_BLANK_ROWS)


def _chunk_students(students: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    return [
        students[index : index + ORDINARY_GRADE_PAGE_STUDENT_CAPACITY]
        for index in range(0, len(students), ORDINARY_GRADE_PAGE_STUDENT_CAPACITY)
    ]


def _last_used_row_for_pages(pages: list[list[dict[str, Any]]]) -> int:
    if not pages:
        return 35
    last_start = 1 + (len(pages) - 1) * 37
    return last_start + 6 + _table_row_count_for_page(pages[-1], is_last=True) + len(ORDINARY_GRADE_NOTES) - 1


def _period_line(fields: dict[str, Any]) -> str:
    academic_year = str(fields.get("academic_year") or "").strip()
    match = re.search(r"(20)(\d{2}).*?(20)(\d{2})", academic_year)
    if match:
        start = f"{match.group(1)} {match.group(2)}"
        end = f"{match.group(3)} {match.group(4)}"
    else:
        start = "20 __"
        end = "20 __"
    semester = str(fields.get("semester") or "").strip()
    semester_no = _semester_number(semester)
    return f"（{start} — {end} 学年度第 {semester_no} 学期）"


def _meta_line_one(fields: dict[str, Any]) -> str:
    college = str(fields.get("college") or "").strip()
    course_name = str(fields.get("course_name") or "").strip()
    hours = _number_text(fields.get("course_hours"))
    credits = _credit_text(fields.get("credits"))
    return f"二级学院（部）：{college}   课程名称：{course_name}   学时 {hours} 学分：{credits}"


def _meta_line_two(fields: dict[str, Any], *, total_students: int) -> str:
    teacher = str(fields.get("teacher_name") or "").strip()
    class_name = str(fields.get("class_name") or "").strip()
    class_size = _coerce_int(fields.get("class_size"), default=total_students) or total_students
    return f"任课教师：{teacher}      专业年级班级：{class_name}    班级人数：{class_size}"


def _period_fields(value: Any) -> dict[str, str]:
    text = str(value or "")
    fields: dict[str, str] = {}
    match = re.search(r"(20\d{2}).*?(20\d{2})", text)
    if match:
        fields["academic_year"] = f"{match.group(1)}-{match.group(2)}"
    if re.search(r"(?:^|[-_])1(?:$|[-_])|第一|一", text):
        fields["semester"] = "第一学期"
    elif re.search(r"(?:^|[-_])2(?:$|[-_])|第二|二", text):
        fields["semester"] = "第二学期"
    return fields


def _semester_number(value: str) -> str:
    raw = str(value or "").strip()
    if re.search(r"(?:第\s*)?(?:一|1)\s*学期", raw) or raw == "1":
        return "1"
    if re.search(r"(?:第\s*)?(?:二|2)\s*学期", raw) or raw == "2":
        return "2"
    match = re.search(r"(?:^|[-_])(1|2)(?:$|[-_])", raw)
    return match.group(1) if match else "__"


def _semester_label(value: str) -> str:
    number = _semester_number(value)
    if number == "1":
        return "第一学期"
    if number == "2":
        return "第二学期"
    return f"第{value}学期"


def _students_from_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    students = []
    for table in tables or []:
        for row in table.get("rows") or []:
            if not isinstance(row, list) or len(row) < 8:
                continue
            if str(row[0]).strip() in {"序号", ""}:
                continue
            if not str(row[1] or "").strip() or not str(row[2] or "").strip():
                continue
            students.append(
                {
                    "index": _coerce_int(row[0], default=len(students) + 1),
                    "student_number": str(row[1] or "").strip(),
                    "student_name": str(row[2] or "").strip(),
                    "attendance_raw_score": _score_or_blank(row[3]),
                    "homework_scores": [_score_or_blank(row[4]), _score_or_blank(row[5]), _score_or_blank(row[6])],
                    "assessment_score": _score_or_blank(row[7]),
                }
            )
    return students


def _normalize_student_records(values: Any) -> list[dict[str, Any]]:
    result = []
    if not isinstance(values, list):
        return result
    for index, item in enumerate(values, start=1):
        if not isinstance(item, dict):
            continue
        homework = item.get("homework_scores") if isinstance(item.get("homework_scores"), list) else []
        homework = [_score_or_blank(value) for value in homework[:3]]
        homework += [""] * (3 - len(homework))
        student_number = str(item.get("student_number") or "").strip()
        student_name = str(item.get("student_name") or "").strip()
        if not student_number and not student_name:
            continue
        result.append(
            {
                **item,
                "index": _coerce_int(item.get("index"), default=index),
                "student_number": student_number,
                "student_name": student_name,
                "attendance_raw_score": _score_or_blank(item.get("attendance_raw_score")),
                "homework_scores": homework,
                "assessment_score": _score_or_blank(item.get("assessment_score")),
            }
        )
    return result


def _table_from_students(title: str, students: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [["序号", "学号", "姓名", "出勤原始成绩", "作业1", "作业2", "作业3", "测评1", "出勤成绩公式", "作业成绩公式", "测评成绩公式", "平时成绩公式"]]
    for student in students:
        formulas = _as_dict(student.get("source_formulas"))
        rows.append(
            [
                student.get("index") or "",
                student.get("student_number") or "",
                student.get("student_name") or "",
                student.get("attendance_raw_score", ""),
                *(student.get("homework_scores") or ["", "", ""])[:3],
                student.get("assessment_score", ""),
                formulas.get("attendance_score") or "=D行",
                formulas.get("homework_score") or "=AVERAGE(E行:G行)",
                formulas.get("assessment_score") or "=H行",
                formulas.get("ordinary_score") or "=I行*0.4+J行*0.3+K行*0.3",
            ]
        )
    return {"title": title, "rows": rows}


def _formula_templates() -> dict[str, str]:
    return {
        "attendance_score": "=D{row}",
        "homework_score": "=AVERAGE(E{row}:G{row})",
        "assessment_score": "=H{row}",
        "ordinary_score": "=I{row}*0.4+J{row}*0.3+K{row}*0.3",
    }


def calculate_ordinary_grade_score(
    attendance_score: float,
    homework_scores: list[float],
    assessment_score: float,
) -> float:
    homework = [float(value) for value in list(homework_scores or [])[:3]]
    homework += [0.0] * (3 - len(homework))
    return (
        float(attendance_score) * 0.4
        + (sum(homework) / 3.0) * 0.3
        + float(assessment_score) * 0.3
    )


def apply_ordinary_grade_score_floor(
    *,
    attendance_score: float,
    homework_scores: list[float],
    assessment_score: float,
    enabled: bool,
    minimum_score: float,
    seed_parts: tuple[Any, ...] = (),
) -> dict[str, Any]:
    """Adjust task scores only, using a reproducible balanced random distribution."""

    attendance = float(attendance_score)
    source_homework = [float(value) for value in list(homework_scores or [])[:3]]
    source_homework += [0.0] * (3 - len(source_homework))
    source_assessment = float(assessment_score)
    source_values = [*source_homework, source_assessment]
    adjusted = list(source_values)
    requested_floor = float(minimum_score)
    eligible = attendance + 1e-9 >= ORDINARY_GRADE_ATTENDANCE_ELIGIBILITY_PERCENT
    original_score = calculate_ordinary_grade_score(attendance, source_homework, source_assessment)
    max_values = [max(100.0, value) for value in adjusted]
    max_achievable = calculate_ordinary_grade_score(attendance, max_values[:3], max_values[3])
    effective_target = min(requested_floor, max_achievable)
    capped = bool(enabled and eligible and requested_floor > max_achievable + 1e-9)
    reason = "disabled"
    applied = False

    seed_text = "|".join(
        [
            ORDINARY_GRADE_SCORE_FLOOR_ALGORITHM_VERSION,
            *(str(part) for part in seed_parts),
            f"{requested_floor:.4f}",
        ]
    )
    seed_digest = hashlib.sha256(seed_text.encode("utf-8")).digest()
    seed_fingerprint = seed_digest.hex()[:16]

    if not enabled:
        reason = "disabled"
    elif not eligible:
        reason = "attendance_below_threshold"
    elif original_score + 1e-9 >= effective_target:
        reason = "already_at_or_above_floor"
    else:
        rng = random.Random(int.from_bytes(seed_digest[:8], "big"))
        weights = [0.1, 0.1, 0.1, 0.3]
        for _ in range(800):
            current_score = calculate_ordinary_grade_score(attendance, adjusted[:3], adjusted[3])
            if current_score + 1e-9 >= effective_target:
                break
            available = [index for index, value in enumerate(adjusted) if value < 100.0 - 1e-9]
            if not available:
                break
            lowest = min(adjusted[index] for index in available)
            balanced = [index for index in available if adjusted[index] <= lowest + 5.0]
            chosen = balanced[rng.randrange(len(balanced))]
            remaining = effective_target - current_score
            needed_points = max(1, int(math.ceil((remaining - 1e-9) / weights[chosen])))
            capacity = max(1, int(math.ceil(100.0 - adjusted[chosen])))
            random_step = rng.randint(1, min(4, capacity))
            step = min(100.0 - adjusted[chosen], needed_points, random_step)
            adjusted[chosen] = round(adjusted[chosen] + step, 4)
        achieved = calculate_ordinary_grade_score(attendance, adjusted[:3], adjusted[3])
        applied = any(adjusted[index] > source_values[index] + 1e-9 for index in range(4))
        reason = "capped_by_attendance" if capped and achieved + 1e-9 < requested_floor else "adjusted_to_floor"

    achieved_score = calculate_ordinary_grade_score(attendance, adjusted[:3], adjusted[3])
    labels = ["homework_1", "homework_2", "homework_3", "assessment"]
    changes = [
        {
            "item": labels[index],
            "before": round(source_values[index], 4),
            "after": round(adjusted[index], 4),
            "delta": round(adjusted[index] - source_values[index], 4),
        }
        for index in range(4)
        if adjusted[index] > source_values[index] + 1e-9
    ]
    return {
        "enabled": bool(enabled),
        "attendance_threshold": ORDINARY_GRADE_ATTENDANCE_ELIGIBILITY_PERCENT,
        "eligible": eligible,
        "applied": applied,
        "capped": capped,
        "reason": reason,
        "requested_minimum_score": round(requested_floor, 4),
        "effective_target_score": round(effective_target, 4),
        "original_score": round(original_score, 4),
        "achieved_score": round(achieved_score, 4),
        "maximum_achievable_score": round(max_achievable, 4),
        "homework_scores": [round(value, 4) for value in adjusted[:3]],
        "assessment_score": round(adjusted[3], 4),
        "changes": changes,
        "algorithm_version": ORDINARY_GRADE_SCORE_FLOOR_ALGORITHM_VERSION,
        "seed_fingerprint": seed_fingerprint,
    }


def normalize_retake_students(retake_students: Any) -> dict[str, float]:
    """Validate teacher-provided retake overrides into {student_number: target_score}."""
    if not retake_students:
        return {}
    if not isinstance(retake_students, list):
        raise HTTPException(400, "重修学生设置格式不正确。")
    result: dict[str, float] = {}
    for item in retake_students:
        if not isinstance(item, dict):
            raise HTTPException(400, "重修学生设置格式不正确。")
        number = str(item.get("student_number") or "").strip()
        if not number:
            raise HTTPException(400, "重修学生设置缺少学号。")
        try:
            score = float(item.get("ordinary_score"))
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, f"重修学生 {number} 的平时分必须是 0 到 100 之间的数字。") from exc
        if not math.isfinite(score) or score < 0 or score > 100:
            raise HTTPException(400, f"重修学生 {number} 的平时分必须在 0 到 100 之间。")
        if number in result:
            raise HTTPException(400, f"重修学生 {number} 被重复设置，请检查后重试。")
        result[number] = round(score, 2)
    return result


def distribute_retake_ordinary_score(
    target_score: float,
    seed_parts: tuple[Any, ...] = (),
) -> dict[str, Any]:
    """Deterministically split a teacher-set ordinary score across attendance,
    three homework slots and the assessment so the weighted formula
    (attendance×0.4 + homework avg×0.3 + assessment×0.3) lands exactly on target.

    Homework deltas sum to zero (average stays on target) and the assessment
    delta is compensated through attendance, so the identity holds precisely."""
    target = round(float(target_score), 2)
    seed_text = "|".join(
        [
            "retake-v1",
            *(str(part) for part in seed_parts),
            f"{target:.4f}",
        ]
    )
    seed_digest = hashlib.sha256(seed_text.encode("utf-8")).digest()
    rng = random.Random(int.from_bytes(seed_digest[:8], "big"))

    margin = int(min(4.0, target, 100.0 - target))
    d1 = rng.randint(-margin, margin) if margin > 0 else 0
    d2 = rng.randint(max(-margin, -margin - d1), min(margin, margin - d1)) if margin > 0 else 0
    d3 = -(d1 + d2)
    homework = [round(target + delta, 2) for delta in (d1, d2, d3)]

    assessment_candidates = [
        delta
        for delta in (-4, -2, 0, 2, 4)
        if abs(delta) <= margin and 0.0 <= target - 0.75 * delta <= 100.0 and 0.0 <= target + delta <= 100.0
    ] or [0]
    assessment_delta = assessment_candidates[rng.randrange(len(assessment_candidates))]
    assessment = round(target + assessment_delta, 2)
    attendance = round(target - 0.75 * assessment_delta, 2)

    achieved = calculate_ordinary_grade_score(attendance, homework, assessment)
    if abs(achieved - target) > 1e-6 or not all(0.0 <= value <= 100.0 for value in (*homework, assessment, attendance)):
        homework = [target, target, target]
        assessment = target
        attendance = target
        achieved = calculate_ordinary_grade_score(attendance, homework, assessment)
    return {
        "target_score": target,
        "attendance_score": attendance,
        "homework_scores": homework,
        "assessment_score": assessment,
        "achieved_score": round(achieved, 4),
        "seed_fingerprint": seed_digest.hex()[:16],
        "algorithm_version": "retake-v1",
    }


def list_classroom_roster_students(conn, *, class_offering_id: int) -> list[dict[str, Any]]:
    """Public roster listing for wizard UIs (e.g. retake student picker)."""
    return _load_roster(conn, class_offering_id=int(class_offering_id), context={})


ORDINARY_GRADE_MANUAL_EDIT_ALGORITHM_VERSION = "manual-edit-v1"
ORDINARY_GRADE_MANUAL_EDIT_LOG_LIMIT = 100


def solve_ordinary_grade_components(
    *,
    target_score: float,
    attendance_score: float,
    homework_scores: list[float],
    assessment_score: float,
) -> dict[str, Any]:
    """Deterministically back-solve raw components so the official formula
    (attendance×0.4 + homework avg×0.3 + assessment×0.3) rounds to the target.

    Attendance is real check-in data, so it stays untouched whenever the target
    is reachable by moving only homework/assessment; otherwise the whole row is
    set uniformly to the target (the only exact fallback), flagged for audit."""
    target = round(float(target_score), 2)
    attendance = round(_score_or_zero(attendance_score), 2)
    homework = [round(_score_or_zero(value), 2) for value in list(homework_scores or [])[:3]]
    homework += [0.0] * (3 - len(homework))
    assessment = round(_score_or_zero(assessment_score), 2)

    attendance_adjusted = False
    lowest = calculate_ordinary_grade_score(attendance, [0.0, 0.0, 0.0], 0.0)
    highest = calculate_ordinary_grade_score(attendance, [100.0, 100.0, 100.0], 100.0)
    if target < lowest - 1e-9 or target > highest + 1e-9:
        attendance_adjusted = True
        attendance = target
        homework = [target, target, target]
        assessment = target
    else:
        base_homework = list(homework)
        base_assessment = assessment

        def _shifted(shift: float) -> tuple[list[float], float]:
            shifted_homework = [min(100.0, max(0.0, value + shift)) for value in base_homework]
            shifted_assessment = min(100.0, max(0.0, base_assessment + shift))
            return shifted_homework, shifted_assessment

        low, high = -100.0, 100.0
        for _ in range(60):
            middle = (low + high) / 2.0
            candidate_homework, candidate_assessment = _shifted(middle)
            if calculate_ordinary_grade_score(attendance, candidate_homework, candidate_assessment) < target:
                low = middle
            else:
                high = middle
        homework, assessment = _shifted(high)
        homework = [round(value, 2) for value in homework]
        assessment = round(assessment, 2)
        # Rounding to 2dp can leave a tiny residual; a 0.01 nudge on one
        # homework slot moves the formula by 0.001, so this always converges
        # to within half of the displayed 0.01 precision.
        for _ in range(400):
            residual = target - calculate_ordinary_grade_score(attendance, homework, assessment)
            if abs(residual) <= 0.0049:
                break
            step = 0.01 if residual > 0 else -0.01
            order = [3, 0, 1, 2] if abs(residual) > 0.02 else [0, 1, 2, 3]
            moved = False
            for slot in order:
                if slot == 3:
                    candidate = round(assessment + step, 2)
                    if 0.0 <= candidate <= 100.0:
                        assessment = candidate
                        moved = True
                        break
                else:
                    candidate = round(homework[slot] + step, 2)
                    if 0.0 <= candidate <= 100.0:
                        homework[slot] = candidate
                        moved = True
                        break
            if not moved:
                break

    achieved = calculate_ordinary_grade_score(attendance, homework, assessment)
    return {
        "target_score": target,
        "attendance_score": attendance,
        "homework_scores": homework,
        "assessment_score": assessment,
        "achieved_score": round(achieved, 4),
        "attendance_adjusted": attendance_adjusted,
        "exact": abs(round(achieved, 2) - target) < 0.005,
        "algorithm_version": ORDINARY_GRADE_MANUAL_EDIT_ALGORITHM_VERSION,
    }


def _manual_score_value(value: Any, *, label: str, student_number: str) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(400, f"学生 {student_number} 的{label}必须是 0 到 100 之间的数字。") from exc
    if not math.isfinite(number) or number < 0 or number > 100:
        raise HTTPException(400, f"学生 {student_number} 的{label}必须在 0 到 100 之间。")
    return round(number, 2)


def _current_ordinary_score(student: dict[str, Any]) -> float:
    calculated = _as_dict(student.get("calculated_scores"))
    stored = calculated.get("ordinary_score")
    if stored not in (None, ""):
        return round(_score_or_zero(stored), 2)
    return round(
        calculate_ordinary_grade_score(
            _score_or_zero(student.get("attendance_raw_score")),
            [_score_or_zero(value) for value in (student.get("homework_scores") or [])],
            _score_or_zero(student.get("assessment_score")),
        ),
        2,
    )


def apply_ordinary_grade_manual_edits(
    export_payload: dict[str, Any],
    edits: list[dict[str, Any]],
    *,
    editor_name: str = "",
    edited_at: str = "",
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Apply teacher edits to the stored ordinary grade record payload.

    Two modes per student row:
    - components: teacher edited attendance/homework/assessment raw scores; the
      ordinary score is recomputed with the official formula.
    - ordinary: teacher edited the final ordinary score; raw components are
      back-solved deterministically so the formula lands on the target.

    The final transcript reads ``calculated_scores.ordinary_score`` first, so
    both modes persist it — downstream materials always see the edited data.
    One-click refresh rebuilds the payload from live classroom scores, which
    intentionally discards these manual edits."""
    if not isinstance(edits, list) or not edits:
        raise HTTPException(400, "没有需要保存的成绩修改。")
    base = _as_dict(export_payload)
    if not base:
        raise HTTPException(409, "该材料没有可编辑的平时成绩数据。")
    structured = dict(_as_dict(base.get("structured")))
    students = [dict(item) for item in structured.get("students") or [] if isinstance(item, dict)]
    if not students:
        raise HTTPException(409, "该平时成绩记录表没有学生成绩行，无法编辑。")
    by_number: dict[str, int] = {}
    for position, student in enumerate(students):
        number = str(student.get("student_number") or "").strip()
        if number and number not in by_number:
            by_number[number] = position

    timestamp = str(edited_at or "").strip() or datetime.now().isoformat(timespec="seconds")
    retake_policy = dict(_as_dict(structured.get("retake_policy")))
    retake_policy_students = [
        dict(item) for item in retake_policy.get("students") or [] if isinstance(item, dict)
    ]
    edited_rows: list[dict[str, Any]] = []
    attendance_adjusted_students: list[str] = []
    log_entries: list[dict[str, Any]] = []

    for edit in edits:
        if not isinstance(edit, dict):
            raise HTTPException(400, "成绩修改格式不正确。")
        number = str(edit.get("student_number") or "").strip()
        if not number:
            raise HTTPException(400, "成绩修改缺少学号。")
        if number not in by_number:
            raise HTTPException(400, f"学号 {number} 不在本平时成绩记录表中。")
        student = students[by_number[number]]
        before = {
            "attendance_raw_score": student.get("attendance_raw_score", ""),
            "homework_scores": list(student.get("homework_scores") or []),
            "assessment_score": student.get("assessment_score", ""),
            "ordinary_score": _current_ordinary_score(student),
        }

        # Apply any provided raw components first, so an ordinary-score edit
        # back-solves from the freshest component values in the same request.
        if edit.get("attendance_raw_score") is not None:
            student["attendance_raw_score"] = _manual_score_value(
                edit.get("attendance_raw_score"), label="出勤成绩", student_number=number
            )
        homework = [_score_or_blank(value) for value in list(student.get("homework_scores") or [])[:3]]
        homework += [""] * (3 - len(homework))
        raw_homework_edit = edit.get("homework_scores")
        if raw_homework_edit is not None:
            if not isinstance(raw_homework_edit, list) or len(raw_homework_edit) > 3:
                raise HTTPException(400, f"学生 {number} 的作业成绩必须是不超过 3 项的列表。")
            for slot, value in enumerate(raw_homework_edit):
                if value is None:
                    continue
                homework[slot] = _manual_score_value(
                    value, label=f"作业{slot + 1}成绩", student_number=number
                )
        student["homework_scores"] = homework
        if edit.get("assessment_score") is not None:
            student["assessment_score"] = _manual_score_value(
                edit.get("assessment_score"), label="测评成绩", student_number=number
            )

        target = edit.get("ordinary_score")
        mode = "ordinary" if target is not None else "components"
        attendance_adjusted = False
        if mode == "ordinary":
            target_value = _manual_score_value(target, label="平时成绩", student_number=number)
            solution = solve_ordinary_grade_components(
                target_score=target_value,
                attendance_score=_score_or_zero(student.get("attendance_raw_score")),
                homework_scores=[_score_or_zero(value) for value in student["homework_scores"]],
                assessment_score=_score_or_zero(student.get("assessment_score")),
            )
            attendance_adjusted = bool(solution["attendance_adjusted"])
            if attendance_adjusted:
                student["attendance_raw_score"] = solution["attendance_score"]
                attendance_adjusted_students.append(number)
            student["homework_scores"] = list(solution["homework_scores"])
            student["assessment_score"] = solution["assessment_score"]
            ordinary_score = target_value
            retake = _as_dict(student.get("retake"))
            if student.get("is_retake") and retake:
                student["retake"] = {**retake, "target_score": target_value}
            for entry in retake_policy_students:
                if (
                    str(entry.get("student_number") or "").strip() == number
                    and str(entry.get("mode") or "manual") != "roster_default"
                ):
                    entry["target_score"] = target_value
        else:
            ordinary_score = round(
                calculate_ordinary_grade_score(
                    _score_or_zero(student.get("attendance_raw_score")),
                    [_score_or_zero(value) for value in student["homework_scores"]],
                    _score_or_zero(student.get("assessment_score")),
                ),
                2,
            )
            retake = _as_dict(student.get("retake"))
            if student.get("is_retake") and retake and retake.get("target_score") is not None:
                student["retake"] = {**retake, "target_score": ordinary_score}
            for entry in retake_policy_students:
                if (
                    str(entry.get("student_number") or "").strip() == number
                    and str(entry.get("mode") or "manual") != "roster_default"
                ):
                    entry["target_score"] = ordinary_score

        attendance_value = _score_or_zero(student.get("attendance_raw_score"))
        homework_values = [_score_or_zero(value) for value in student["homework_scores"]]
        student["calculated_scores"] = {
            **_as_dict(student.get("calculated_scores")),
            "attendance_score": round(attendance_value, 2),
            "homework_score": round(sum(homework_values) / 3.0, 2),
            "assessment_score": round(_score_or_zero(student.get("assessment_score")), 2),
            "ordinary_score": ordinary_score,
        }
        after = {
            "attendance_raw_score": student.get("attendance_raw_score", ""),
            "homework_scores": list(student.get("homework_scores") or []),
            "assessment_score": student.get("assessment_score", ""),
            "ordinary_score": ordinary_score,
        }
        student["manual_edit"] = {
            "mode": mode,
            "edited_at": timestamp,
            "editor_name": str(editor_name or "").strip(),
            "attendance_adjusted": attendance_adjusted,
            "algorithm_version": ORDINARY_GRADE_MANUAL_EDIT_ALGORITHM_VERSION,
            "before": before,
            "after": after,
        }
        edited_rows.append(
            {
                "student_number": number,
                "student_name": str(student.get("student_name") or ""),
                "mode": mode,
                "ordinary_score": ordinary_score,
                "attendance_adjusted": attendance_adjusted,
            }
        )
        log_entries.append(
            {
                "student_number": number,
                "student_name": str(student.get("student_name") or ""),
                "mode": mode,
                "edited_at": timestamp,
                "editor_name": str(editor_name or "").strip(),
                "before": before,
                "after": after,
            }
        )

    if not edited_rows:
        raise HTTPException(400, "没有需要保存的成绩修改。")

    manual_edit_log = [
        dict(item) for item in structured.get("manual_edit_log") or [] if isinstance(item, dict)
    ]
    manual_edit_log.extend(log_entries)
    manual_edit_log = manual_edit_log[-ORDINARY_GRADE_MANUAL_EDIT_LOG_LIMIT:]

    edited_names = "、".join(
        f"{row['student_name'] or row['student_number']}（{row['student_number']}）"
        for row in edited_rows[:5]
    )
    if len(edited_rows) > 5:
        edited_names += f" 等 {len(edited_rows)} 人"
    warning_text = (
        f"教师于 {timestamp[:16].replace('T', ' ')} 手动调整了 {len(edited_rows)} 名学生的平时成绩"
        f"（{edited_names}）。一键更新时将重新以课堂实际作业、测验成绩覆盖手动修改。"
    )
    warnings = _merge_warnings(structured.get("warnings"), [warning_text])

    if retake_policy_students or retake_policy:
        retake_policy["students"] = retake_policy_students
    new_structured = {
        **structured,
        "students": students,
        "retake_policy": retake_policy,
        "manual_edit_log": manual_edit_log,
        "warnings": warnings,
    }
    tables: list[dict[str, Any]] = []
    if base.get("tables"):
        tables = [
            _table_from_students(f"第 {page_index + 1} 版", chunk)
            for page_index, chunk in enumerate(_chunk_students(students))
        ]
    normalized = normalize_ordinary_grade_record_payload(
        metadata={},
        content_markdown="",
        tables=tables,
        export_payload={**base, "structured": new_structured},
    )
    normalized_structured = _as_dict(normalized.get("structured"))
    normalized["content_markdown"] = _build_content_markdown(
        _as_dict(normalized.get("fields")),
        normalized_structured.get("students") or [],
        _as_dict(normalized_structured.get("source_assignments")),
    )
    summary = {
        "edited_count": len(edited_rows),
        "edited_students": edited_rows,
        "attendance_adjusted_students": attendance_adjusted_students,
        "message": (
            f"已保存 {len(edited_rows)} 名学生的成绩修改，平时成绩已按公式自动均衡。"
            + ("其中部分学生因目标分超出作业/测评可调范围，出勤分也被同步调整。" if attendance_adjusted_students else "")
        ),
    }
    return normalized, summary


def serialize_ordinary_grade_students_for_editor(export_payload: dict[str, Any]) -> dict[str, Any]:
    """Flatten the stored payload into editor-friendly rows for the raw-score
    editing dialog (secondary preview modal)."""
    base = _as_dict(export_payload)
    fields = _as_dict(base.get("fields"))
    structured = _as_dict(base.get("structured"))
    rows: list[dict[str, Any]] = []
    for student in structured.get("students") or []:
        if not isinstance(student, dict):
            continue
        homework = [_score_or_blank(value) for value in list(student.get("homework_scores") or [])[:3]]
        homework += [""] * (3 - len(homework))
        attendance = _score_or_blank(student.get("attendance_raw_score"))
        assessment = _score_or_blank(student.get("assessment_score"))
        homework_values = [_score_or_zero(value) for value in homework]
        manual_edit = _as_dict(student.get("manual_edit"))
        rows.append(
            {
                "index": _coerce_int(student.get("index")),
                "student_number": str(student.get("student_number") or ""),
                "student_name": str(student.get("student_name") or ""),
                "attendance_raw_score": attendance,
                "homework_scores": homework,
                "assessment_score": assessment,
                "computed": {
                    "attendance": round(_score_or_zero(attendance), 2),
                    "homework": round(sum(homework_values) / 3.0, 2),
                    "assessment": round(_score_or_zero(assessment), 2),
                    "ordinary": _current_ordinary_score(student),
                },
                "is_retake": bool(student.get("is_retake")),
                "source_score_missing": _as_dict(student.get("source_score_missing")),
                "manual_edit": (
                    {
                        "mode": str(manual_edit.get("mode") or ""),
                        "edited_at": str(manual_edit.get("edited_at") or ""),
                        "attendance_adjusted": bool(manual_edit.get("attendance_adjusted")),
                    }
                    if manual_edit
                    else None
                ),
            }
        )
    return {
        "document_type": ORDINARY_GRADE_RECORD_TYPE,
        "document_type_label": ORDINARY_GRADE_RECORD_LABEL,
        "fields": {
            "course_name": fields.get("course_name") or "",
            "class_name": fields.get("class_name") or "",
            "teacher_name": fields.get("teacher_name") or "",
            "academic_year": fields.get("academic_year") or "",
            "semester": fields.get("semester") or "",
            "class_size": fields.get("class_size") or len(rows),
        },
        "score_weights": {"attendance": 0.4, "homework": 0.3, "assessment": 0.3},
        "students": rows,
        "manual_edit_count": sum(1 for row in rows if row["manual_edit"]),
    }


def _normalize_score_floor_policy(*, enabled: bool, minimum_score: float) -> dict[str, Any]:
    try:
        value = float(minimum_score)
    except (TypeError, ValueError) as exc:
        raise HTTPException(400, "最低平时分必须是 0 到 100 之间的数字。") from exc
    if not math.isfinite(value) or value < 0 or value > 100:
        raise HTTPException(400, "最低平时分必须在 0 到 100 之间。")
    return {
        "enabled": bool(enabled),
        "minimum_score": round(value, 2),
        "attendance_eligibility_percent": ORDINARY_GRADE_ATTENDANCE_ELIGIBILITY_PERCENT,
        "attendance_is_adjusted": False,
        "missing_source_scores_default_to_zero": True,
    }


def _formulas_match_expected(row: int, formulas: dict[str, str]) -> bool:
    expected = {key: value.format(row=row) for key, value in _formula_templates().items()}
    return all(str(formulas.get(key) or "").replace("$", "").upper() == expected[key].upper() for key in expected)


def _ordinary_grade_queryable_fields(fields: dict[str, Any], structured: dict[str, Any]) -> dict[str, Any]:
    return {
        "course_name": fields.get("course_name") or "",
        "class_name": fields.get("class_name") or "",
        "teacher_name": fields.get("teacher_name") or "",
        "academic_year": fields.get("academic_year") or "",
        "semester": fields.get("semester") or "",
        "class_size": fields.get("class_size") or "",
        "export_filename": fields.get("export_filename") or "",
        "source_assignments": structured.get("source_assignments") or {},
        "student_count": len(structured.get("students") or []),
        "formula_templates": structured.get("formula_templates") or {},
        "attendance_sync": structured.get("attendance_sync") or {},
        "generation_requirements": structured.get("generation_requirements") or "",
        "score_floor_policy": structured.get("score_floor_policy") or {},
    }


def _build_content_markdown(fields: dict[str, Any], students: list[dict[str, Any]], source_assignments: dict[str, Any]) -> str:
    lines = [
        "# 广西外国语学院学生平时成绩记录表",
        "",
        f"- 课程名称：{fields.get('course_name') or ''}",
        f"- 专业年级班级：{fields.get('class_name') or ''}",
        f"- 任课教师：{fields.get('teacher_name') or ''}",
        f"- 学年学期：{fields.get('academic_year') or ''} {fields.get('semester') or ''}",
        f"- 学生人数：{len(students)}",
    ]
    homework = source_assignments.get("homework_assignments") if isinstance(source_assignments, dict) else None
    assessment = source_assignments.get("assessment_assignment") if isinstance(source_assignments, dict) else None
    if homework:
        lines.append(f"- 平时作业来源：{'；'.join(str(item.get('title') or '') for item in homework if isinstance(item, dict))}")
    if isinstance(assessment, dict) and assessment:
        lines.append(f"- 测评来源：{assessment.get('title') or ''}")
    if str(fields.get("generation_requirements") or "").strip():
        lines.append(f"- 生成要求：{str(fields.get('generation_requirements')).strip()}")
    score_floor_policy = _as_dict(fields.get("score_floor_policy"))
    if fields.get("minimum_ordinary_score_enabled") or score_floor_policy.get("enabled"):
        minimum_score = fields.get("minimum_ordinary_score", score_floor_policy.get("minimum_score", 60))
        threshold = fields.get(
            "attendance_eligibility_percent",
            score_floor_policy.get("attendance_eligibility_percent", ORDINARY_GRADE_ATTENDANCE_ELIGIBILITY_PERCENT),
        )
        lines.append(f"- 最低分保护：出勤率达到 {_format_score(threshold)}% 后，最低平时分 {_format_score(minimum_score)} 分；出勤不调整。")
    retake_students_note = [
        f"{student.get('student_name') or student.get('student_number')}（{student.get('student_number')}，{_format_score(_as_dict(student.get('retake')).get('target_score'))} 分）"
        for student in students
        if isinstance(student, dict) and student.get("is_retake")
    ]
    if retake_students_note:
        lines.append(f"- 重修/免修学生（平时分由教师设定）：{'；'.join(retake_students_note)}")
    lines.extend(["", "| 序号 | 学号 | 姓名 | 出勤 | 作业1 | 作业2 | 作业3 | 测评 |", "| --- | --- | --- | --- | --- | --- | --- | --- |"])
    for student in students[:80]:
        homework_scores = list(student.get("homework_scores") or [])[:3]
        homework_scores += [""] * (3 - len(homework_scores))
        lines.append(
            "| "
            + " | ".join(
                [
                    str(student.get("index") or ""),
                    str(student.get("student_number") or ""),
                    str(student.get("student_name") or ""),
                    _cell_text(student.get("attendance_raw_score")),
                    *[_cell_text(value) for value in homework_scores],
                    _cell_text(student.get("assessment_score")),
                ]
            )
            + " |"
        )
    return "\n".join(lines)


def _assignment_summary(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": int(row["id"]),
        "title": row.get("title") or f"作业 {row['id']}",
        **ordinary_grade_assignment_kind_info(row),
        "status": row.get("status") or "",
        "due_at": row.get("due_at") or "",
    }


def _score_or_blank(value: Any) -> float | str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if math.isnan(number) or math.isinf(number):
        return ""
    return round(number, 2)


def _score_or_zero(value: Any) -> float:
    score = _score_or_blank(value)
    return float(score) if score != "" else 0.0


def _format_score(value: Any) -> str:
    number = _score_or_zero(value)
    return str(int(number)) if number.is_integer() else f"{number:.2f}".rstrip("0").rstrip(".")


def _number_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    return str(int(number)) if number.is_integer() else str(number)


def _credit_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):.1f}"
    except (TypeError, ValueError):
        return str(value)


def _cell_text(value: Any) -> str:
    if value == "":
        return ""
    return _number_text(value)


def _fields_from_classroom_context(context: dict[str, Any]) -> dict[str, Any]:
    raw = _as_dict(context)
    return _compact_dict(
        {
            "school": raw.get("school_name") or raw.get("school") or "广西外国语学院",
            "college": raw.get("college") or "",
            "department": raw.get("department") or "",
            "course_name": raw.get("course_name") or "",
            "class_name": raw.get("class_name") or "",
            "teacher_name": raw.get("teacher_name") or "",
            "academic_year": raw.get("academic_year") or "",
            "semester": raw.get("semester") or "",
            "course_hours": raw.get("course_hours") or raw.get("total_hours") or "",
            "credits": raw.get("credits") or "",
        }
    )


def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _compact_dict(value: dict[str, Any]) -> dict[str, Any]:
    return {key: val for key, val in value.items() if not _is_blank(val)}


def _is_blank(value: Any) -> bool:
    return value in (None, "", [], {})


def _coerce_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def _merge_warnings(*values: Any) -> list[str]:
    merged: list[str] = []
    for value in values:
        if not value:
            continue
        items = value if isinstance(value, list) else [value]
        for item in items:
            text = str(item or "").strip()
            if text and text not in merged:
                merged.append(text)
    return merged


def _dedupe(values: list[str]) -> list[str]:
    result = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in result:
            result.append(text)
    return result


__all__ = [name for name in globals() if not name.startswith("__")]
