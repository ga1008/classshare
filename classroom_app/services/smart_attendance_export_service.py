from __future__ import annotations

import io
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MEDIA_TYPE = "application/pdf"

_CHECKED_STATUS = "CHECKED"
_ABSENT_STATUS = "UNCHECKED"
_LATE_OR_EARLY_STATUS = "LATE_OR_EARLY"
_LEAVE_STATUSES = {"SICK_LEAVE", "PERSONAL_LEAVE"}
_STATUS_SYMBOLS = {
    _CHECKED_STATUS: "√",
    _ABSENT_STATUS: "×",
    _LATE_OR_EARLY_STATUS: "⊕",
    "SICK_LEAVE": "○",
    "PERSONAL_LEAVE": "○",
}
_SPREADSHEET_FONT_NAME = "宋体"
_PDF_FONT_NAME = "lanshare-song"
_PDF_FALLBACK_FONT = "china-s"


def _resolve_pdf_font_file() -> str | None:
    env_font = os.getenv("LANSHARE_PDF_FONT_FILE", "").strip()
    candidates = [
        env_font,
        r"C:\Windows\Fonts\simsun.ttc",
        r"C:\Windows\Fonts\simsunb.ttf",
        r"C:\Windows\Fonts\simhei.ttf",
        "/usr/share/fonts/truetype/arphic/uming.ttc",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return candidate
    return None


def _ensure_pdf_font(page: fitz.Page) -> str:
    font_file = _resolve_pdf_font_file()
    if not font_file:
        return _PDF_FALLBACK_FONT
    try:
        page.insert_font(fontname=_PDF_FONT_NAME, fontfile=font_file)
        return _PDF_FONT_NAME
    except Exception:
        return _PDF_FALLBACK_FONT


@dataclass(frozen=True)
class SmartAttendanceExport:
    filename: str
    media_type: str
    content: bytes


def _coerce_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def _normalize_status(value: Any) -> str:
    return str(value or "").strip().upper()


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value or "")
    cleaned = re.sub(r"_+", "_", cleaned).strip("._ ")
    return cleaned or "智慧课堂出勤成绩"


def _parse_datetime(value: Any) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    normalized = raw.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _format_checkin_label(value: Any) -> str:
    dt = _parse_datetime(value)
    if dt:
        return f"{dt.month}/{dt.day}\n{dt.hour:02d}:{dt.minute:02d}"
    raw = str(value or "").strip()
    if not raw:
        return "-"
    date_match = re.search(r"(\d{1,2})[-/](\d{1,2})", raw)
    time_match = re.search(r"(\d{1,2}):(\d{2})", raw)
    if date_match and time_match:
        return f"{int(date_match.group(1))}/{int(date_match.group(2))}\n{int(time_match.group(1)):02d}:{time_match.group(2)}"
    return raw[:12]


def _session_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    dt = _parse_datetime(row.get("checkin_time"))
    return (
        dt or datetime.max,
        _coerce_int(row.get("week_index"), 999),
        _coerce_int(row.get("weekday"), 999),
        _coerce_int(row.get("section_index"), 999),
        _coerce_int(row.get("session_id"), 999999),
        _coerce_int(row.get("id"), 999999),
    )


def _build_title(semester_name: str) -> str:
    text = _clean_text(semester_name)
    year_match = re.search(r"(\d{4})\D+(\d{4})", text)
    term_match = re.search(r"(?:第)?\s*([一二12])\s*学期", text)
    if not term_match and year_match:
        term_match = re.search(r"(?:^|[-_/\s])([12])(?:$|\D)", text[year_match.end() :])
    if year_match and term_match:
        term_map = {"一": "1", "二": "2", "1": "1", "2": "2"}
        term = term_map.get(term_match.group(1), term_match.group(1))
        return f"{year_match.group(1)}-{year_match.group(2)}学年第{term}学期平时成绩记录表"
    if text:
        return f"{text}平时成绩记录表"
    return "平时成绩记录表"


def _score_for_statuses(statuses: Iterable[str]) -> float | None:
    normalized = [_normalize_status(status) for status in statuses if _normalize_status(status)]
    if not normalized:
        return None
    checked = sum(1 for status in normalized if status == _CHECKED_STATUS)
    return round(checked * 100.0 / len(normalized), 2)


def _score_display(score: float | None) -> str:
    return "" if score is None else f"{score:.2f}"


def _symbol_for_status(status: Any) -> str:
    return _STATUS_SYMBOLS.get(_normalize_status(status), "")


def _student_number_display(value: Any) -> str:
    text = _clean_text(value)
    if len(text) > 10 and re.fullmatch(r"[0-9A-Za-z]+", text):
        return "\n".join(text[index : index + 5] for index in range(0, len(text), 5))
    return text


def _load_latest_sessions(conn, *, class_offering_id: int, teacher_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT *
        FROM smart_classroom_checkin_sessions
        WHERE class_offering_id = ?
          AND teacher_id = ?
          AND session_id IS NOT NULL
        ORDER BY session_id,
                 COALESCE(checkin_time, '') DESC,
                 COALESCE(synced_at, '') DESC,
                 id DESC
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchall()
    latest_by_session: dict[int, dict[str, Any]] = {}
    for row in rows:
        row_dict = dict(row)
        session_id = _coerce_int(row_dict.get("session_id"))
        if session_id <= 0 or session_id in latest_by_session:
            continue
        latest_by_session[session_id] = row_dict
    return sorted(latest_by_session.values(), key=_session_sort_key)


def _load_export_context(conn, *, class_offering_id: int, teacher_id: int) -> dict[str, Any]:
    offering = conn.execute(
        """
        SELECT o.id,
               o.teacher_id,
               o.class_id,
               COALESCE(s.name, o.semester, '') AS semester_name,
               c.name AS course_name,
               COALESCE(c.academic_course_code, '') AS course_code,
               cl.name AS class_name,
               COALESCE(cl.academic_college, '') AS class_college,
               t.name AS teacher_name,
               t.email AS teacher_email
        FROM class_offerings o
        JOIN courses c ON c.id = o.course_id
        JOIN classes cl ON cl.id = o.class_id
        JOIN teachers t ON t.id = o.teacher_id
        LEFT JOIN academic_semesters s ON s.id = o.semester_id
        WHERE o.id = ?
          AND o.teacher_id = ?
        LIMIT 1
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchone()
    if offering is None:
        raise ValueError("未找到当前教师可导出的课堂。")

    offering_dict = dict(offering)
    credential = conn.execute(
        """
        SELECT username
        FROM teacher_smart_classroom_credentials
        WHERE teacher_id = ?
          AND enabled = 1
        ORDER BY COALESCE(last_verified_at, updated_at, created_at, '') DESC, id DESC
        LIMIT 1
        """,
        (int(teacher_id),),
    ).fetchone()
    watermark_account = _clean_text(credential["username"]) if credential else ""
    if not watermark_account:
        watermark_account = _clean_text(offering_dict.get("teacher_email")) or str(teacher_id)
    offering_dict["watermark_text"] = f"{_clean_text(offering_dict.get('teacher_name')) or '教师'}：{watermark_account}"
    offering_dict["title"] = _build_title(str(offering_dict.get("semester_name") or ""))
    return offering_dict


def _build_student_rows(
    conn,
    *,
    class_offering_id: int,
    sessions: list[dict[str, Any]],
    context: dict[str, Any],
) -> list[dict[str, Any]]:
    roster_rows = conn.execute(
        """
        SELECT s.id,
               s.student_id_number,
               s.name,
               COALESCE(NULLIF(s.academic_college, ''), NULLIF(cl.academic_college, ''), '') AS college,
               COALESCE(NULLIF(s.academic_class_name, ''), NULLIF(cl.academic_class_name, ''), cl.name, '') AS class_name
        FROM students s
        JOIN class_offerings o ON o.class_id = s.class_id
        JOIN classes cl ON cl.id = s.class_id
        WHERE o.id = ?
          AND COALESCE(s.enrollment_status, 'active') = 'active'
        ORDER BY s.student_id_number, s.id
        """,
        (int(class_offering_id),),
    ).fetchall()

    students: list[dict[str, Any]] = []
    by_id: dict[int, dict[str, Any]] = {}
    by_number: dict[str, dict[str, Any]] = {}

    for roster in roster_rows:
        number = _clean_text(roster["student_id_number"])
        item = {
            "student_id": int(roster["id"]),
            "student_number": number,
            "student_name": _clean_text(roster["name"]),
            "college": _clean_text(roster["college"]) or _clean_text(context.get("class_college")),
            "class_name": _clean_text(roster["class_name"]) or _clean_text(context.get("class_name")),
            "statuses": {},
            "remote_only": False,
        }
        students.append(item)
        by_id[int(roster["id"])] = item
        if number:
            by_number[number] = item

    if sessions:
        session_ids = [int(row["id"]) for row in sessions]
        placeholders = ",".join("?" for _ in session_ids)
        status_rows = conn.execute(
            f"""
            SELECT checkin_session_id,
                   student_id,
                   student_number,
                   student_name,
                   status
            FROM smart_classroom_checkin_students
            WHERE checkin_session_id IN ({placeholders})
            ORDER BY student_number, id
            """,
            session_ids,
        ).fetchall()
    else:
        status_rows = []

    for status_row in status_rows:
        row_dict = dict(status_row)
        student_id = _coerce_int(row_dict.get("student_id"))
        number = _clean_text(row_dict.get("student_number"))
        item = by_id.get(student_id) if student_id else None
        if item is None and number:
            item = by_number.get(number)
        if item is None:
            item = {
                "student_id": None,
                "student_number": number,
                "student_name": _clean_text(row_dict.get("student_name")),
                "college": _clean_text(context.get("class_college")),
                "class_name": _clean_text(context.get("class_name")),
                "statuses": {},
                "remote_only": True,
            }
            students.append(item)
            if number:
                by_number[number] = item
        item["statuses"][int(row_dict["checkin_session_id"])] = _normalize_status(row_dict.get("status"))

    for item in students:
        ordered_statuses = [item["statuses"].get(int(session["id"]), "") for session in sessions]
        score = _score_for_statuses(ordered_statuses)
        item["score"] = score
        item["ordinary_score"] = score
        item["ordered_statuses"] = ordered_statuses

    return students


def _build_dataset(conn, *, class_offering_id: int, teacher_id: int) -> dict[str, Any]:
    context = _load_export_context(conn, class_offering_id=class_offering_id, teacher_id=teacher_id)
    sessions = _load_latest_sessions(conn, class_offering_id=class_offering_id, teacher_id=teacher_id)
    if not sessions:
        raise ValueError("当前课堂还没有可导出的智慧课堂点名记录，请先同步点名。")
    students = _build_student_rows(
        conn,
        class_offering_id=class_offering_id,
        sessions=sessions,
        context=context,
    )
    if not students:
        raise ValueError("当前课堂没有可导出的学生名单。")
    return {
        "context": context,
        "sessions": sessions,
        "students": students,
    }


def _column_headers(dataset: dict[str, Any]) -> list[str]:
    return [
        "序号",
        "学院",
        "班级",
        "学号",
        "姓名",
        *[_format_checkin_label(session.get("checkin_time")) for session in dataset["sessions"]],
        "成绩",
        "平时\n成绩",
    ]


def _notes(dataset: dict[str, Any]) -> list[str]:
    return [
        "注： 1. 此表记录平时成绩，满分为100分，授课教师要详实记录，期末作为学生平时成绩判定的依据，并装订入试卷册；",
        "    2. 平时成绩占__%，期末成绩占__%；平时成绩中出勤占100.0%，课堂测验占0.0%，作业占0.0%，课堂表现占0.0%其他占__%（如有，请列出具体名称）；",
        "    3. 同一门课程多个平行教学班的平时成绩项目须一致，评分标准由教师制定，教师各负责任教，授课教师在开学第一周向学生公布平时成绩记录与评判标准；",
        "    4. 出勤√、缺课×、迟到与早退⊕、请假○，作业、实验、测验及课堂表现等用A、B、C、D、E五级给与评价或给出相应的分数，实验学时超过总学时的1/3者，实验成绩单独计算。",
    ]


def _build_xlsx(dataset: dict[str, Any]) -> bytes:
    context = dataset["context"]
    sessions = dataset["sessions"]
    students = dataset["students"]
    headers = _column_headers(dataset)
    total_columns = len(headers)
    score_start_col = 6 + len(sessions)

    wb = Workbook()
    for named_style in wb._named_styles:
        if getattr(named_style, "name", "") == "Normal":
            named_style.font = Font(name=_SPREADSHEET_FONT_NAME, size=10)
            break
    ws = wb.active
    ws.title = "平时成绩"

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:4"
    ws.freeze_panes = "F5"
    ws.oddHeader.center.text = f'&"{_SPREADSHEET_FONT_NAME},Regular"&KCCCCCC&18{context["watermark_text"]}'
    ws.sheet_view.showGridLines = False

    title_font = Font(name=_SPREADSHEET_FONT_NAME, size=20, bold=True)
    meta_font = Font(name=_SPREADSHEET_FONT_NAME, size=11)
    body_font = Font(name=_SPREADSHEET_FONT_NAME, size=10)
    symbol_font = Font(name=_SPREADSHEET_FONT_NAME, size=13)
    header_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_side = Side(style="thin", color="222222")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws.cell(1, 1, context["title"])
    ws.cell(1, 1).font = title_font
    ws.cell(1, 1).alignment = center
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(5, total_columns))
    ws.cell(2, 1, f"任课教师：{context['teacher_name']}        课程名称：{context['course_name']}")
    ws.cell(2, 1).font = meta_font
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24

    for col in range(1, 6):
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        cell = ws.cell(3, col, headers[col - 1])
        cell.font = body_font
        cell.alignment = center
        cell.fill = header_fill

    if sessions:
        ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=5 + len(sessions))
        ws.cell(3, 6, "出勤情况（日期）")
        ws.cell(3, 6).font = body_font
        ws.cell(3, 6).alignment = center
        ws.cell(3, 6).fill = header_fill
        for index, session in enumerate(sessions, start=6):
            cell = ws.cell(4, index, _format_checkin_label(session.get("checkin_time")))
            cell.font = Font(name=_SPREADSHEET_FONT_NAME, size=9)
            cell.alignment = center
            cell.fill = header_fill

    for col, text in ((score_start_col, "成绩"), (score_start_col + 1, "平时\n成绩")):
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        cell = ws.cell(3, col, text)
        cell.font = body_font
        cell.alignment = center
        cell.fill = header_fill

    for row_index, student in enumerate(students, start=5):
        values = [
            row_index - 4,
            student["college"],
            student["class_name"],
            _student_number_display(student["student_number"]),
            student["student_name"],
            *[_symbol_for_status(status) for status in student["ordered_statuses"]],
            student["score"],
            student["ordinary_score"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col_index, value)
            cell.font = symbol_font if 6 <= col_index < score_start_col else body_font
            cell.alignment = center
            if col_index == 4:
                cell.number_format = "@"
            if col_index >= score_start_col and value is not None:
                cell.number_format = "0.00"
        ws.row_dimensions[row_index].height = 44

    note_start_row = 5 + len(students) + 1
    for offset, note in enumerate(_notes(dataset)):
        row = note_start_row + offset
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
        cell = ws.cell(row, 1, note)
        cell.font = Font(name=_SPREADSHEET_FONT_NAME, size=8.5)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22

    for row in ws.iter_rows(min_row=3, max_row=4 + len(students), min_col=1, max_col=total_columns):
        for cell in row:
            cell.border = border
            if cell.row <= 4:
                cell.fill = header_fill

    fixed_widths = {
        1: 5.5,
        2: 8.5,
        3: 9.0,
        4: 11.0,
        5: 8.5,
        score_start_col: 7.5,
        score_start_col + 1: 7.5,
    }
    for col, width in fixed_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    available_attendance_width = 92.0
    attendance_width = available_attendance_width / max(len(sessions), 1)
    for col in range(6, score_start_col):
        ws.column_dimensions[get_column_letter(col)].width = attendance_width

    ws.print_area = f"A1:{get_column_letter(total_columns)}{note_start_row + len(_notes(dataset)) - 1}"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _text_units(text: str) -> float:
    total = 0.0
    for char in str(text or ""):
        total += 1.0 if ord(char) < 128 else 1.9
    return total


def _wrap_text(text: Any, max_units: float) -> list[str]:
    normalized = str(text or "").strip()
    if not normalized:
        return [""]
    lines: list[str] = []
    current = ""
    current_units = 0.0
    for char in normalized:
        if char == "\n":
            lines.append(current)
            current = ""
            current_units = 0.0
            continue
        unit = 1.0 if ord(char) < 128 else 1.9
        if current and current_units + unit > max_units:
            lines.append(current)
            current = char
            current_units = unit
        else:
            current += char
            current_units += unit
    if current or not lines:
        lines.append(current)
    return lines[:6]


def _fit_font_size(text: Any, width: float, base_size: float, minimum: float = 5.2) -> float:
    units = max(_text_units(str(text or "").replace("\n", "")), 1.0)
    estimated = (width - 4) / units * 1.68
    return max(min(base_size, estimated), minimum)


def _draw_textbox(
    page: fitz.Page,
    rect: fitz.Rect,
    text: Any,
    *,
    font_size: float = 8.0,
    fontname: str | None = None,
    align: int = fitz.TEXT_ALIGN_CENTER,
    bold: bool = False,
    color: tuple[float, float, float] = (0, 0, 0),
    max_lines: int | None = None,
) -> None:
    max_units = max(1.0, rect.width / max(font_size * 0.58, 1))
    lines = str(text or "").split("\n") if "\n" in str(text or "") else _wrap_text(text, max_units)
    if max_lines:
        lines = lines[:max_lines]
    line_height = font_size * 1.18
    total_height = line_height * max(len(lines), 1)
    top = rect.y0 + max(1.0, (rect.height - total_height) / 2)
    target = fitz.Rect(rect.x0 + 1.5, top, rect.x1 - 1.5, rect.y1 - 1)
    resolved_font = fontname or _PDF_FALLBACK_FONT
    text_value = "\n".join(lines)
    size = font_size
    while size >= 4.8:
        remaining = page.insert_textbox(
            target,
            text_value,
            fontsize=size,
            fontname=resolved_font,
            color=color,
            align=align,
        )
        if remaining >= 0:
            return
        size -= 0.45


def _draw_watermark(page: fitz.Page, text: str, *, fontname: str) -> None:
    if not text:
        return
    matrix = fitz.Matrix(1, 1).prerotate(-18)
    for y in range(100, int(page.rect.height) - 80, 170):
        for x in range(90, int(page.rect.width) - 250, 360):
            point = fitz.Point(float(x), float(y))
            page.insert_text(
                point,
                text,
                fontsize=22,
                fontname=fontname,
                fill=(0.62, 0.62, 0.62),
                fill_opacity=0.22,
                morph=(point, matrix),
            )


def _pdf_column_widths(page_width: float, margin_x: float, session_count: int) -> list[float]:
    content_width = page_width - margin_x * 2
    fixed = [27.0, 36.0, 38.0, 42.0, 39.0]
    score = [28.0, 28.0]
    attendance_width = (content_width - sum(fixed) - sum(score)) / max(session_count, 1)
    return [*fixed, *([attendance_width] * session_count), *score]


def _draw_pdf_header(
    page: fitz.Page,
    dataset: dict[str, Any],
    *,
    margin_x: float,
    title_top: float,
    fontname: str,
) -> None:
    context = dataset["context"]
    title_rect = fitz.Rect(margin_x, title_top, page.rect.width - margin_x, title_top + 30)
    page.insert_textbox(
        title_rect,
        context["title"],
        fontsize=20,
        fontname=fontname,
        color=(0, 0, 0),
        align=fitz.TEXT_ALIGN_CENTER,
    )
    meta = f"任课教师： {context['teacher_name']}        课程名称： {context['course_name']}"
    page.insert_textbox(
        fitz.Rect(margin_x, title_top + 38, page.rect.width - margin_x, title_top + 60),
        meta,
        fontsize=10.5,
        fontname=fontname,
        color=(0, 0, 0),
        align=fitz.TEXT_ALIGN_LEFT,
    )


def _draw_pdf_table(
    page: fitz.Page,
    dataset: dict[str, Any],
    rows: list[dict[str, Any]],
    *,
    table_top: float,
    margin_x: float,
    row_height: float,
    fontname: str,
) -> float:
    sessions = dataset["sessions"]
    widths = _pdf_column_widths(page.rect.width, margin_x, len(sessions))
    x_positions = [margin_x]
    for width in widths[:-1]:
        x_positions.append(x_positions[-1] + width)
    header_one = 15.0
    header_two = 24.0
    header_height = header_one + header_two
    border_color = (0.05, 0.05, 0.05)
    header_labels = ["序号", "学院", "班级", "学号", "姓名"]

    for col, label in enumerate(header_labels):
        rect = fitz.Rect(x_positions[col], table_top, x_positions[col] + widths[col], table_top + header_height)
        page.draw_rect(rect, color=border_color, width=0.55)
        _draw_textbox(page, rect, label, font_size=8.4, fontname=fontname)

    attendance_start = x_positions[5]
    attendance_width = sum(widths[5 : 5 + len(sessions)])
    attendance_rect = fitz.Rect(attendance_start, table_top, attendance_start + attendance_width, table_top + header_one)
    page.draw_rect(attendance_rect, color=border_color, width=0.55)
    _draw_textbox(page, attendance_rect, "出勤情况（日期）", font_size=8.2, fontname=fontname)

    date_font = max(5.3, min(8.0, widths[5] * 0.22)) if sessions else 7.0
    for idx, session in enumerate(sessions):
        col = 5 + idx
        rect = fitz.Rect(
            x_positions[col],
            table_top + header_one,
            x_positions[col] + widths[col],
            table_top + header_height,
        )
        page.draw_rect(rect, color=border_color, width=0.55)
        _draw_textbox(
            page,
            rect,
            _format_checkin_label(session.get("checkin_time")),
            font_size=date_font,
            fontname=fontname,
            max_lines=2,
        )

    for col, label in ((5 + len(sessions), "成绩"), (6 + len(sessions), "平时\n成绩")):
        rect = fitz.Rect(x_positions[col], table_top, x_positions[col] + widths[col], table_top + header_height)
        page.draw_rect(rect, color=border_color, width=0.55)
        _draw_textbox(page, rect, label, font_size=8.0, fontname=fontname)

    y = table_top + header_height
    for row in rows:
        values = [
            row["index"],
            row["college"],
            row["class_name"],
            _student_number_display(row["student_number"]),
            row["student_name"],
            *[_symbol_for_status(status) for status in row["ordered_statuses"]],
            _score_display(row["score"]),
            _score_display(row["ordinary_score"]),
        ]
        for col, value in enumerate(values):
            rect = fitz.Rect(x_positions[col], y, x_positions[col] + widths[col], y + row_height)
            page.draw_rect(rect, color=border_color, width=0.48)
            if 5 <= col < 5 + len(sessions):
                _draw_textbox(page, rect, value, font_size=9.4, fontname=fontname)
            elif col >= 5 + len(sessions):
                _draw_textbox(page, rect, value, font_size=_fit_font_size(value, widths[col], 8.0, 5.6), fontname=fontname)
            else:
                _draw_textbox(page, rect, value, font_size=_fit_font_size(value, widths[col], 8.0, 5.6), fontname=fontname)
        y += row_height
    return y


def _draw_pdf_notes(page: fitz.Page, dataset: dict[str, Any], *, y: float, margin_x: float, fontname: str) -> None:
    note_rect = fitz.Rect(margin_x, y + 8, page.rect.width - margin_x, page.rect.height - 18)
    page.insert_textbox(
        note_rect,
        "\n".join(_notes(dataset)),
        fontsize=8.0,
        fontname=fontname,
        color=(0.05, 0.05, 0.05),
        align=fitz.TEXT_ALIGN_LEFT,
    )


def _pdf_capacity(page_height: float, table_top: float, *, row_height: float, notes: bool) -> int:
    reserved_notes = 76.0 if notes else 0.0
    available = page_height - 18.0 - reserved_notes - table_top - 39.0
    return max(1, int(math.floor(available / row_height)))


def _build_pdf(dataset: dict[str, Any]) -> bytes:
    students = [{**item, "index": index} for index, item in enumerate(dataset["students"], start=1)]
    doc = fitz.open()
    page_width = 841.89
    page_height = 595.28
    margin_x = 24.0
    row_height = 32.0
    table_top = 90.0
    remaining = students[:]
    page_index = 0
    while remaining:
        page = doc.new_page(width=page_width, height=page_height)
        fontname = _ensure_pdf_font(page)
        _draw_watermark(page, dataset["context"]["watermark_text"], fontname=fontname)
        _draw_pdf_header(page, dataset, margin_x=margin_x, title_top=18.0, fontname=fontname)

        cap_with_notes = _pdf_capacity(page_height, table_top, row_height=row_height, notes=True)
        cap_full = _pdf_capacity(page_height, table_top, row_height=row_height, notes=False)
        if len(remaining) <= cap_with_notes:
            chunk = remaining
            remaining = []
            is_last = True
        else:
            take_count = min(cap_full, len(remaining))
            if len(remaining) <= cap_full:
                take_count = max(1, len(remaining) - cap_with_notes)
            chunk = remaining[:take_count]
            remaining = remaining[take_count:]
            is_last = False

        table_bottom = _draw_pdf_table(
            page,
            dataset,
            chunk,
            table_top=table_top,
            margin_x=margin_x,
            row_height=row_height,
            fontname=fontname,
        )
        if is_last:
            _draw_pdf_notes(page, dataset, y=table_bottom, margin_x=margin_x, fontname=fontname)
        page_index += 1

    if not doc.page_count:
        page = doc.new_page(width=page_width, height=page_height)
        fontname = _ensure_pdf_font(page)
        _draw_watermark(page, dataset["context"]["watermark_text"], fontname=fontname)
    return doc.tobytes(garbage=4, deflate=True)


def build_smart_attendance_export(
    conn,
    *,
    class_offering_id: int,
    teacher_id: int,
    file_format: str,
) -> SmartAttendanceExport:
    normalized_format = str(file_format or "xlsx").strip().lower()
    if normalized_format not in {"xlsx", "excel", "pdf"}:
        raise ValueError("导出格式仅支持 Excel 或 PDF。")
    dataset = _build_dataset(conn, class_offering_id=int(class_offering_id), teacher_id=int(teacher_id))
    context = dataset["context"]
    base_filename = _safe_filename(f"{context['semester_name']}_{context['course_name']}_{context['class_name']}_平时成绩记录表")
    if normalized_format == "pdf":
        return SmartAttendanceExport(
            filename=f"{base_filename}.pdf",
            media_type=PDF_MEDIA_TYPE,
            content=_build_pdf(dataset),
        )
    return SmartAttendanceExport(
        filename=f"{base_filename}.xlsx",
        media_type=XLSX_MEDIA_TYPE,
        content=_build_xlsx(dataset),
    )
