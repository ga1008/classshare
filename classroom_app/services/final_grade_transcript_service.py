from __future__ import annotations

import hashlib
import io
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import HTTPException

from ..db.connection import get_configured_db_engine
from .excel_upload_service import open_upload_workbook
from .exam_grade_record_service import EXAM_GRADE_RECORD_TYPE
from .ordinary_grade_record_service import (
    ORDINARY_GRADE_RECORD_TYPE,
    calculate_ordinary_grade_score,
)


FINAL_GRADE_TRANSCRIPT_TYPE = "final_grade_transcript"
FINAL_GRADE_TRANSCRIPT_LABEL = "期末成绩单"
FINAL_GRADE_TRANSCRIPT_SCHEMA_VERSION = "gxufl-final-grade-transcript-v1"
FINAL_GRADE_TRANSCRIPT_REMARKS = ("违纪", "免训", "免训*", "取消考试资格", "作弊", "缺考")
FINAL_GRADE_TRANSCRIPT_HEADERS = (
    "序号",
    "班级",
    "学号",
    "姓名",
    "平时(必填)",
    "期末(必填)",
    "备注",
)
FINAL_GRADE_TRANSCRIPT_LAYOUT = {
    "layout_source": "gxufl_student_grade_entry_template_xlsx",
    "sheet_name": "学生成绩录入模板",
    "column_widths": [20.0] * 7,
    "row_heights": {"header": 25.0, "student": 20.0},
    "margins_in": {
        "left": 0.7,
        "right": 0.7,
        "top": 0.75,
        "bottom": 0.75,
        "header": 0.3,
        "footer": 0.3,
    },
}


@dataclass(frozen=True)
class FinalGradeTranscriptParseResult:
    metadata: dict[str, Any]
    content_markdown: str
    tables: list[dict[str, Any]]
    warnings: list[str]
    export_payload: dict[str, Any]
    formula_count: int = 0


def normalize_final_grade_transcript_payload(
    *,
    metadata: dict[str, Any] | None,
    content_markdown: str = "",
    tables: list[dict[str, Any]] | None = None,
    export_payload: dict[str, Any] | None = None,
    classroom_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    base = dict(export_payload or {})
    fields = _compact_dict(_as_dict(base.get("fields")))
    fields.update({key: value for key, value in _as_dict(metadata).items() if not _is_blank(value)})
    fields.update(
        {
            key: value
            for key, value in _as_dict(classroom_context).items()
            if _is_blank(fields.get(key)) and not _is_blank(value)
        }
    )
    fields.setdefault("school", "广西外国语学院")
    fields.setdefault("title", FINAL_GRADE_TRANSCRIPT_LABEL)

    structured = _as_dict(base.get("structured"))
    students = _normalize_students(
        structured.get("students")
        if isinstance(structured.get("students"), list)
        else _students_from_tables(tables or [])
    )
    if students and not fields.get("class_size"):
        fields["class_size"] = len(students)
    warnings = _dedupe(
        [
            *_string_list(base.get("warnings")),
            *_string_list(structured.get("warnings")),
        ]
    )
    if not students:
        warnings.append("未识别到学生成绩行，请确认文件使用教务系统学生成绩录入模板。")

    normalized_structured = {
        **structured,
        "template_schema_version": FINAL_GRADE_TRANSCRIPT_SCHEMA_VERSION,
        "headers": list(FINAL_GRADE_TRANSCRIPT_HEADERS),
        "students": students,
        "warnings": _dedupe(warnings),
        "source_lineage": _as_dict(structured.get("source_lineage")),
        "identity_policy": {
            "primary_key": "student_number",
            "secondary_check": "student_name",
            "ordering_source": "academic_exam_roster.row_order",
        },
    }
    base.update(
        {
            "document_group": "final_material",
            "document_type": FINAL_GRADE_TRANSCRIPT_TYPE,
            "document_type_label": FINAL_GRADE_TRANSCRIPT_LABEL,
            "template_key": FINAL_GRADE_TRANSCRIPT_TYPE,
            "fields": fields,
            "tables": tables or base.get("tables") or [],
            "layout_profile": dict(FINAL_GRADE_TRANSCRIPT_LAYOUT),
            "structured": normalized_structured,
            "queryable_fields": {
                "school": fields.get("school") or "",
                "college": fields.get("college") or "",
                "department": fields.get("department") or "",
                "course_name": fields.get("course_name") or "",
                "course_nature": fields.get("course_nature") or "",
                "class_name": fields.get("class_name") or "",
                "academic_year": fields.get("academic_year") or "",
                "semester": fields.get("semester") or "",
                "class_offering_id": fields.get("class_offering_id") or "",
                "student_count": len(students),
                "ordinary_grade_record_id": _as_dict(normalized_structured.get("source_lineage"))
                .get("ordinary_grade_record", {})
                .get("record_id", ""),
                "exam_grade_record_id": _as_dict(normalized_structured.get("source_lineage"))
                .get("exam_grade_record", {})
                .get("record_id", ""),
                "exam_roster_item_id": _as_dict(normalized_structured.get("source_lineage"))
                .get("exam_roster", {})
                .get("item_id", ""),
            },
            "content_markdown": content_markdown
            or base.get("content_markdown")
            or _build_content_markdown(fields, students),
            "compatibility": {
                **_as_dict(base.get("compatibility")),
                "source_format_preserved": True,
                "layout_source": FINAL_GRADE_TRANSCRIPT_LAYOUT["layout_source"],
                "requires_template_confirmation": False,
                "template_schema_version": FINAL_GRADE_TRANSCRIPT_SCHEMA_VERSION,
            },
        }
    )
    return base


def parse_final_grade_transcript_file(
    file_path: Path,
    original_name: str,
    import_metadata: dict[str, Any] | None = None,
) -> FinalGradeTranscriptParseResult:
    with open_upload_workbook(
        file_path,
        original_name,
        material_label="期末成绩单",
        data_only=False,
    ) as workbook:
        worksheet, header_row, columns, matching_sheet_count = _locate_transcript_worksheet(workbook)
        if worksheet is None or not header_row:
            raise HTTPException(422, "未识别到“序号、班级、学号、姓名、平时(必填)、期末(必填)、备注”表头。")

        warnings: list[str] = []
        if matching_sheet_count > 1:
            warnings.append(
                f"文件中有 {matching_sheet_count} 张工作表符合期末成绩单结构，"
                f"已按工作簿顺序解析《{worksheet.title}》。"
            )
        students: list[dict[str, Any]] = []
        seen_numbers: dict[str, int] = {}
        class_names: list[str] = []
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            values = {
                key: worksheet.cell(row_number, column).value
                for key, column in columns.items()
            }
            if all(_is_blank(values.get(key)) for key in ("student_number", "student_name", "ordinary_score", "final_score", "remark")):
                continue
            student_number = _text(values.get("student_number"))
            student_name = _text(values.get("student_name"))
            class_name = _text(values.get("class_name"))
            if not student_number or not student_name:
                raise HTTPException(422, f"第 {row_number} 行缺少学号或姓名，无法建立可靠学生关联。")
            if student_number in seen_numbers:
                raise HTTPException(
                    422,
                    f"学号 {student_number} 在第 {seen_numbers[student_number]}、{row_number} 行重复，无法安全导入。",
                )
            seen_numbers[student_number] = row_number
            if class_name and class_name not in class_names:
                class_names.append(class_name)
            ordinary_score = _score_or_blank(values.get("ordinary_score"), row_number=row_number, label="平时")
            final_score = _score_or_blank(values.get("final_score"), row_number=row_number, label="期末")
            remark = _text(values.get("remark"))
            if remark and remark not in FINAL_GRADE_TRANSCRIPT_REMARKS:
                warnings.append(f"第 {row_number} 行备注“{remark}”不在模板下拉选项中，已原样保留。")
            students.append(
                {
                    "index": _positive_int(values.get("index"), len(students) + 1),
                    "source_row": row_number,
                    "class_name": class_name,
                    "student_number": student_number,
                    "student_name": student_name,
                    "ordinary_score": ordinary_score,
                    "final_score": final_score,
                    "remark": remark,
                }
            )
        if not students:
            raise HTTPException(422, "没有识别到任何学生成绩行。")

        metadata = _compact_dict(dict(import_metadata or {}))
        if len(class_names) == 1:
            metadata.setdefault("class_name", class_names[0])
        elif len(class_names) > 1:
            metadata.setdefault("class_name", "、".join(class_names))
            warnings.append(f"文件包含 {len(class_names)} 个班级，已逐行保留班级字段。")
        metadata.setdefault("source_filename", original_name)
        metadata.setdefault("teacher_name", _teacher_from_filename(original_name))
        metadata.setdefault("course_name", _course_from_filename(original_name))
        metadata.setdefault("school", "广西外国语学院")
        metadata["class_size"] = len(students)
        missing_fields = [
            label
            for key, label in (
                ("academic_year", "学年"),
                ("semester", "学期"),
                ("school", "学校"),
                ("college", "学院"),
                ("department", "系部"),
                ("class_name", "班级"),
                ("course_name", "课程"),
                ("course_nature", "课程属性"),
            )
            if _is_blank(metadata.get(key))
        ]
        if missing_fields:
            raise HTTPException(422, f"导入信息不完整，请补充：{'、'.join(missing_fields)}。")

        table = _table_from_students(students)
        payload = normalize_final_grade_transcript_payload(
            metadata=metadata,
            tables=[table],
            export_payload={
                "fields": metadata,
                "structured": {
                    "students": students,
                    "warnings": warnings,
                    "import_contract": {
                        "header_row": header_row,
                        "source_sheet": worksheet.title,
                        "source_row_count": len(students),
                        "all_columns_preserved": True,
                    },
                },
            },
        )
        content_markdown = _build_content_markdown(metadata, students)
        payload["content_markdown"] = content_markdown
        return FinalGradeTranscriptParseResult(
            metadata=metadata,
            content_markdown=content_markdown,
            tables=[table],
            warnings=_dedupe(warnings),
            export_payload=payload,
        )


def build_final_grade_transcript_xlsx(payload: dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.writer.theme import theme_xml
    except ImportError as exc:
        raise RuntimeError(f"缺少 XLSX 导出依赖 openpyxl: {exc}") from exc

    export_payload = normalize_final_grade_transcript_payload(
        metadata={},
        content_markdown=str(payload.get("content_markdown") or ""),
        tables=payload.get("tables") if isinstance(payload.get("tables"), list) else [],
        export_payload=_as_dict(payload.get("export_payload")) or payload,
    )
    fields = _as_dict(export_payload.get("fields"))
    structured = _as_dict(export_payload.get("structured"))
    students = _normalize_students(structured.get("students"))

    workbook = Workbook()
    official_theme = (
        theme_xml.replace('typeface="Cambria"', 'typeface="Aptos Display"')
        .replace('typeface="Calibri"', 'typeface="Aptos Narrow"')
        .replace(
            'script="Hans" typeface="&#x5B8B;&#x4F53;"',
            'script="Hans" typeface="等线 Light"',
            1,
        )
        .replace(
            'script="Hans" typeface="&#x5B8B;&#x4F53;"',
            'script="Hans" typeface="等线"',
            1,
        )
    )
    workbook.loaded_theme = official_theme.encode("utf-8")
    # Excel converts stored column-width units with the workbook's Normal font.
    # Match the official template's 等线 11pt base style so the visible widths
    # remain identical when opened or printed, even though every populated cell
    # has its own explicit font.
    workbook._named_styles[0].name = "常规"
    workbook._named_styles[0].font = Font(
        name="等线",
        size=11,
        family=2,
        scheme="minor",
    )
    worksheet = workbook.active
    worksheet.title = FINAL_GRADE_TRANSCRIPT_LAYOUT["sheet_name"]
    worksheet.sheet_format.baseColWidth = None
    worksheet.sheet_format.defaultRowHeight = 14
    worksheet.sheet_format.dyDescent = 0.3
    for column_letter, width in zip(
        "ABCDEFG",
        FINAL_GRADE_TRANSCRIPT_LAYOUT["column_widths"],
    ):
        worksheet.column_dimensions[column_letter].width = width
    worksheet.row_dimensions[1].height = FINAL_GRADE_TRANSCRIPT_LAYOUT["row_heights"]["header"]

    indexed_8 = Color(indexed=8)
    indexed_10 = Color(indexed=10)
    indexed_22 = Color(indexed=22)
    thin = Side(style="thin", color=Color(auto=True))
    border = Border(left=thin, right=thin, top=thin, bottom=thin, diagonal=Side())
    center = Alignment(horizontal="center", vertical="center")
    unlocked = Protection(locked=False)
    header_fill = PatternFill(patternType="solid", fgColor=indexed_22, bgColor=indexed_22)
    body_fill = PatternFill(patternType=None, bgColor=indexed_22)
    main_header_font = Font(
        name="宋体",
        size=23,
        bold=True,
        color=indexed_8,
        charset=134,
        family=3,
    )
    score_header_font = Font(
        name="黑体",
        size=15,
        bold=True,
        color=indexed_10,
        charset=134,
        family=3,
    )
    body_font = Font(
        name="宋体",
        size=13,
        bold=True,
        color=indexed_8,
        charset=134,
        family=3,
    )

    for column, header in enumerate(FINAL_GRADE_TRANSCRIPT_HEADERS, start=1):
        cell = worksheet.cell(1, column, header)
        cell.font = score_header_font if column in {5, 6} else main_header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
        cell.protection = unlocked
        if column == 5:
            cell.number_format = "0.00_ "

    worksheet["E1"].comment = Comment(
        "该分项或者阶段成绩录入级制为【百分制】,请输入 0 至 100 之间的数值!",
        "None",
    )
    worksheet["F1"].comment = Comment(
        "该分项或者阶段成绩录入级制为【百分制】,请输入 0 至 100 之间的数值!",
        "None",
    )
    worksheet["G1"].comment = Comment(
        "请从违纪,免训,免训*,取消考试资格,作弊,缺考选择相应备注!",
        "None",
    )

    for index, student in enumerate(students, start=1):
        row = index + 1
        worksheet.row_dimensions[row].height = FINAL_GRADE_TRANSCRIPT_LAYOUT["row_heights"]["student"]
        values = [
            student.get("index") or index,
            student.get("class_name") or fields.get("class_name") or "",
            student.get("student_number") or "",
            student.get("student_name") or "",
            student.get("ordinary_score") if student.get("ordinary_score") not in (None, "") else "",
            student.get("final_score") if student.get("final_score") not in (None, "") else "",
            student.get("remark") or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row, column, value)
            cell.font = body_font
            cell.fill = body_fill
            cell.border = border
            cell.alignment = center
            cell.protection = unlocked
            if column == 5:
                cell.number_format = "0.00_ "
            else:
                cell.number_format = "General"

    last_row = max(2, len(students) + 1)
    score_validation = DataValidation(
        type="decimal",
        formula1="0",
        formula2="100",
        allow_blank=True,
        showDropDown=None,
        showInputMessage=None,
        showErrorMessage=True,
        errorTitle="警告",
        error="请输入 0 至 100 之间的数值!",
        promptTitle="提示",
    )
    worksheet.add_data_validation(score_validation)
    score_validation.add(f"E2:F{last_row}")
    remark_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(FINAL_GRADE_TRANSCRIPT_REMARKS)}"',
        allow_blank=True,
        showDropDown=None,
        showInputMessage=None,
        showErrorMessage=True,
        errorTitle="警告",
        promptTitle="提示",
    )
    worksheet.add_data_validation(remark_validation)
    remark_validation.add(f"G2:G{last_row}")

    margins = FINAL_GRADE_TRANSCRIPT_LAYOUT["margins_in"]
    worksheet.page_margins.left = margins["left"]
    worksheet.page_margins.right = margins["right"]
    worksheet.page_margins.top = margins["top"]
    worksheet.page_margins.bottom = margins["bottom"]
    worksheet.page_margins.header = margins["header"]
    worksheet.page_margins.footer = margins["footer"]
    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_final_grade_transcript_export_filename(fields: dict[str, Any]) -> str:
    course = _filename_part(fields.get("course_name")) or "课程"
    class_name = _filename_part(fields.get("class_name")) or "班级"
    teacher = _filename_part(fields.get("teacher_name"))
    suffix = f"[{teacher}]" if teacher else ""
    return f"{course}学生成绩录入模板{suffix}-{class_name}.xlsx"


def normalize_final_grade_academic_year(value: Any) -> str:
    return _academic_year_token(value)


def normalize_final_grade_semester(value: Any) -> str:
    return _semester_label(value)


def build_final_grade_transcript_readiness(
    conn,
    *,
    class_offering_id: int,
    teacher_id: int,
) -> dict[str, Any]:
    context = _load_context(
        conn,
        class_offering_id=int(class_offering_id),
        teacher_id=int(teacher_id),
    )
    roster_item, roster_students = _load_exam_roster(conn, class_offering_id=int(class_offering_id), teacher_id=int(teacher_id))
    if not roster_item:
        return {
            "ready": False,
            "status": "roster_missing",
            "message": "尚未取得教务系统考试名单，请重新同步后再生成。",
            "classroom": context,
            "roster": {"ready": False, "student_count": 0},
            "sources": _empty_sources(class_offering_id),
        }

    ordinary_record = _select_matching_record(
        conn,
        teacher_id=int(teacher_id),
        class_offering_id=int(class_offering_id),
        document_type=ORDINARY_GRADE_RECORD_TYPE,
        context=context,
        roster_students=roster_students,
    )
    exam_record = _select_matching_record(
        conn,
        teacher_id=int(teacher_id),
        class_offering_id=int(class_offering_id),
        document_type=EXAM_GRADE_RECORD_TYPE,
        context=context,
        roster_students=roster_students,
    )
    ordinary_source = _source_match_summary(
        ordinary_record,
        roster_students=roster_students,
        document_type=ORDINARY_GRADE_RECORD_TYPE,
        class_offering_id=int(class_offering_id),
    )
    exam_source = _source_match_summary(
        exam_record,
        roster_students=roster_students,
        document_type=EXAM_GRADE_RECORD_TYPE,
        class_offering_id=int(class_offering_id),
    )
    duplicate_roster_numbers = _duplicates(
        [_text(student.get("student_number")) for student in roster_students]
    )
    roster_signature = _roster_signature(roster_item, roster_students)
    roster_ready = bool(roster_students) and not duplicate_roster_numbers
    ready = roster_ready and ordinary_source["ready"] and exam_source["ready"]
    if duplicate_roster_numbers:
        message = f"教务考试名单存在重复学号：{'、'.join(duplicate_roster_numbers[:5])}，已阻止生成。"
        status = "roster_invalid"
    elif not roster_students:
        message = "教务考试名单为空，已阻止生成。"
        status = "roster_empty"
    elif not ordinary_source["record_found"] and not exam_source["record_found"]:
        message = "尚未找到同课堂、同学年学期的平时成绩表和考核登分表。"
        status = "sources_missing"
    elif not ordinary_source["ready"]:
        message = ordinary_source["message"]
        status = "ordinary_source_incomplete"
    elif not exam_source["ready"]:
        message = exam_source["message"]
        status = "exam_source_incomplete"
    else:
        message = f"名单及两份上游成绩已逐人核对完成，可生成 {len(roster_students)} 人期末成绩单。"
        status = "ready"
    return {
        "ready": ready,
        "status": status,
        "message": message,
        "classroom": context,
        "roster": {
            "ready": roster_ready,
            "item_id": int(roster_item["id"]),
            "student_count": len(roster_students),
            "synced_at": roster_item["synced_at"] or "",
            "course_name": roster_item["course_name"] or "",
            "exam_course_key": roster_item["exam_course_key"] or "",
            "signature": roster_signature,
            "duplicate_student_numbers": duplicate_roster_numbers,
            "students": [
                {
                    "row_order": student.get("row_order"),
                    "student_number": student.get("student_number"),
                    "student_name": student.get("student_name"),
                    "class_name": student.get("admin_class_name") or context.get("class_name") or "",
                }
                for student in roster_students
            ],
            "preview": [
                {
                    "row_order": student.get("row_order"),
                    "student_number": student.get("student_number"),
                    "student_name": student.get("student_name"),
                    "class_name": student.get("admin_class_name") or context.get("class_name") or "",
                }
                for student in roster_students[:8]
            ],
        },
        "sources": {
            "ordinary_grade_record": ordinary_source,
            "exam_grade_record": exam_source,
        },
    }


def build_final_grade_transcript_payload(
    conn,
    *,
    class_offering_id: int,
    teacher_id: int,
    expected_roster_synced_at: str = "",
    expected_roster_signature: str = "",
    expected_ordinary_record_id: int | None = None,
    expected_exam_record_id: int | None = None,
) -> dict[str, Any]:
    readiness = build_final_grade_transcript_readiness(
        conn,
        class_offering_id=int(class_offering_id),
        teacher_id=int(teacher_id),
    )
    if not readiness.get("ready"):
        raise HTTPException(409, readiness.get("message") or "期末成绩单来源尚未准备完成。")
    roster = _as_dict(readiness.get("roster"))
    sources = _as_dict(readiness.get("sources"))
    ordinary_summary = _as_dict(sources.get("ordinary_grade_record"))
    exam_summary = _as_dict(sources.get("exam_grade_record"))
    if expected_roster_synced_at and expected_roster_synced_at != str(roster.get("synced_at") or ""):
        raise HTTPException(409, "考试名单在确认后发生了变化，请返回生成窗口重新核对。")
    if expected_roster_signature and expected_roster_signature != str(roster.get("signature") or ""):
        raise HTTPException(409, "教务考试名单内容在确认后发生了变化，请重新同步并核对完整名单。")
    if expected_ordinary_record_id and int(ordinary_summary.get("record_id") or 0) != int(expected_ordinary_record_id):
        raise HTTPException(409, "平时成绩表在确认后发生了变化，请重新核对后生成。")
    if expected_exam_record_id and int(exam_summary.get("record_id") or 0) != int(expected_exam_record_id):
        raise HTTPException(409, "考核登分表在确认后发生了变化，请重新核对后生成。")

    roster_item, roster_students = _load_exam_roster(
        conn,
        class_offering_id=int(class_offering_id),
        teacher_id=int(teacher_id),
    )
    if not roster_item or str(roster_item["synced_at"] or "") != str(roster.get("synced_at") or ""):
        raise HTTPException(409, "考试名单在生成期间发生了变化，请返回生成窗口重新核对。")
    if _roster_signature(roster_item, roster_students) != str(roster.get("signature") or ""):
        raise HTTPException(409, "教务考试名单内容在生成期间发生了变化，请重新同步并核对。")
    ordinary_record = _load_record(conn, int(ordinary_summary["record_id"]))
    exam_record = _load_record(conn, int(exam_summary["record_id"]))
    if str(ordinary_record["updated_at"] or "") != str(ordinary_summary.get("updated_at") or ""):
        raise HTTPException(409, "平时成绩表在生成期间发生了变化，请重新核对后生成。")
    if str(exam_record["updated_at"] or "") != str(exam_summary.get("updated_at") or ""):
        raise HTTPException(409, "考核登分表在生成期间发生了变化，请重新核对后生成。")
    ordinary_students = _students_by_number(ordinary_record, ORDINARY_GRADE_RECORD_TYPE)
    exam_students = _students_by_number(exam_record, EXAM_GRADE_RECORD_TYPE)
    context = _as_dict(readiness.get("classroom"))
    rows: list[dict[str, Any]] = []
    for index, roster_student in enumerate(roster_students, start=1):
        number = _text(roster_student.get("student_number"))
        ordinary = ordinary_students.get(number)
        exam = exam_students.get(number)
        if not ordinary or not exam:
            raise HTTPException(409, "学生成绩来源在生成期间发生了变化，请重新核对后生成。")
        rows.append(
            {
                "index": index,
                "class_name": _text(roster_student.get("admin_class_name")) or context.get("class_name") or "",
                "student_number": number,
                "student_name": _text(roster_student.get("student_name")),
                "ordinary_score": ordinary["score"],
                "final_score": exam["score"],
                "remark": "",
                "source_match": {
                    "ordinary_grade_record_id": int(ordinary_summary["record_id"]),
                    "ordinary_source_name": ordinary.get("student_name") or "",
                    "exam_grade_record_id": int(exam_summary["record_id"]),
                    "exam_source_name": exam.get("student_name") or "",
                    "roster_row_order": roster_student.get("row_order"),
                },
            }
        )
    unique_classes = _dedupe([_text(row.get("class_name")) for row in rows if _text(row.get("class_name"))])
    fields = {
        **context,
        "title": FINAL_GRADE_TRANSCRIPT_LABEL,
        "class_name": "、".join(unique_classes) or context.get("class_name") or "",
        "class_size": len(rows),
    }
    lineage = {
        "exam_roster": {
            "item_id": int(roster_item["id"]),
            "synced_at": roster_item["synced_at"] or "",
            "student_count": len(rows),
            "order_field": "row_order",
            "exam_course_key": roster_item["exam_course_key"] or "",
        },
        "ordinary_grade_record": {
            "record_id": int(ordinary_summary["record_id"]),
            "updated_at": ordinary_summary.get("updated_at") or "",
            "matched_student_count": len(rows),
        },
        "exam_grade_record": {
            "record_id": int(exam_summary["record_id"]),
            "updated_at": exam_summary.get("updated_at") or "",
            "matched_student_count": len(rows),
        },
    }
    payload = normalize_final_grade_transcript_payload(
        metadata=fields,
        export_payload={
            "fields": fields,
            "structured": {
                "students": rows,
                "source_lineage": lineage,
                "warnings": [],
            },
        },
    )
    payload["content_markdown"] = _build_content_markdown(fields, rows)
    return payload


def _empty_sources(class_offering_id: int) -> dict[str, Any]:
    return {
        "ordinary_grade_record": {
            "ready": False,
            "record_found": False,
            "label": "平时成绩表",
            "message": "请先生成同课堂、同学年学期的平时成绩表。",
            "generate_url": _source_generate_url(ORDINARY_GRADE_RECORD_TYPE, class_offering_id),
        },
        "exam_grade_record": {
            "ready": False,
            "record_found": False,
            "label": "考核登分表",
            "message": "请先生成同课堂、同学年学期的考核登分表。",
            "generate_url": _source_generate_url(EXAM_GRADE_RECORD_TYPE, class_offering_id),
        },
    }


def _select_matching_record(
    conn,
    *,
    teacher_id: int,
    class_offering_id: int,
    document_type: str,
    context: dict[str, Any],
    roster_students: list[dict[str, Any]],
):
    rows = conn.execute(
        """
        SELECT r.*
        FROM material_ai_import_records r
        WHERE r.teacher_id = ?
          AND r.document_group = 'final_material'
          AND r.document_type = ?
          AND r.parse_status = 'completed'
        ORDER BY r.updated_at DESC, r.id DESC
        LIMIT 100
        """,
        (int(teacher_id), document_type),
    ).fetchall()
    candidates: list[tuple[tuple[Any, ...], Any]] = []
    for row in rows:
        payload = _record_export_payload(row)
        fields = _as_dict(payload.get("fields"))
        if not _same_context(fields, context):
            continue
        if not _record_matches_offering_scope(
            conn,
            row,
            fields=fields,
            class_offering_id=int(class_offering_id),
        ):
            continue
        summary = _source_match_summary(
            row,
            roster_students=roster_students,
            document_type=document_type,
            class_offering_id=int(class_offering_id),
        )
        issue_count = (
            int(summary.get("missing_count") or 0)
            + int(summary.get("conflict_count") or 0)
            + int(summary.get("duplicate_count") or 0)
        )
        rank = (
            1 if summary.get("ready") else 0,
            int(summary.get("matched_count") or 0),
            -issue_count,
            str(row["updated_at"] or ""),
            int(row["id"]),
        )
        candidates.append((rank, row))
    return max(candidates, key=lambda item: item[0])[1] if candidates else None


def _record_matches_offering_scope(
    conn,
    record,
    *,
    fields: dict[str, Any],
    class_offering_id: int,
) -> bool:
    field_offering_id = _positive_int(fields.get("class_offering_id"), 0)
    if field_offering_id:
        return field_offering_id == int(class_offering_id)
    material_ids = [
        int(value)
        for value in (
            record["package_material_id"],
            record["parsed_material_id"],
            record["source_material_id"],
        )
        if _positive_int(value, 0)
    ]
    if not material_ids:
        return False
    placeholders = ",".join("?" for _ in material_ids)
    assignment = conn.execute(
        f"""
        SELECT 1
        FROM course_material_assignments
        WHERE class_offering_id = ?
          AND material_id IN ({placeholders})
        LIMIT 1
        """,
        (int(class_offering_id), *material_ids),
    ).fetchone()
    return bool(assignment)


def _same_context(fields: dict[str, Any], context: dict[str, Any]) -> bool:
    comparisons = (
        ("academic_year", _academic_year_token),
        ("semester", _semester_token),
        ("course_name", _identity_text),
        ("class_name", _class_identity_text),
    )
    for key, normalizer in comparisons:
        expected = normalizer(context.get(key))
        actual = normalizer(fields.get(key))
        if not expected or not actual or expected != actual:
            return False
    field_offering_id = _positive_int(fields.get("class_offering_id"), 0)
    expected_offering_id = _positive_int(context.get("class_offering_id"), 0)
    return not field_offering_id or field_offering_id == expected_offering_id


def _source_match_summary(
    record,
    *,
    roster_students: list[dict[str, Any]],
    document_type: str,
    class_offering_id: int,
) -> dict[str, Any]:
    label = "平时成绩表" if document_type == ORDINARY_GRADE_RECORD_TYPE else "考核登分表"
    generate_url = _source_generate_url(document_type, class_offering_id)
    if record is None:
        return {
            "ready": False,
            "record_found": False,
            "label": label,
            "message": f"未找到与当前课堂、课程、班级、学年和学期完全一致的{label}。",
            "generate_url": generate_url,
            "matched_count": 0,
            "missing_count": len(roster_students),
            "conflict_count": 0,
        }
    students = _students_by_number(record, document_type)
    assignment_targets = _record_assignment_targets(record, document_type)
    raw_students = _record_raw_students(record)
    roster_numbers = {_text(student.get("student_number")) for student in roster_students}
    missing: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []
    for roster_student in roster_students:
        number = _text(roster_student.get("student_number"))
        roster_name = _text(roster_student.get("student_name"))
        jump_targets = _student_jump_targets(
            assignment_targets,
            raw_students.get(number),
            document_type=document_type,
            student_number=number,
        )
        source = students.get(number)
        if source and source.get("duplicate"):
            duplicates.append(
                {
                    "student_number": number,
                    "student_name": roster_name,
                    "jump_targets": jump_targets,
                }
            )
            continue
        if not source or source.get("score") in (None, ""):
            missing.append(
                {
                    "student_number": number,
                    "student_name": roster_name,
                    "jump_targets": jump_targets,
                }
            )
            continue
        if _identity_text(source.get("student_name")) != _identity_text(roster_name):
            conflicts.append(
                {
                    "student_number": number,
                    "roster_name": roster_name,
                    "source_name": _text(source.get("student_name")),
                    "jump_targets": jump_targets,
                }
            )
    extra = [
        {"student_number": number, "student_name": _text(source.get("student_name"))}
        for number, source in students.items()
        if number not in roster_numbers
    ]
    matched_count = len(roster_students) - len(missing) - len(conflicts) - len(duplicates)
    ready = not missing and not conflicts and not duplicates and matched_count == len(roster_students)
    if duplicates:
        message = f"{label}有 {len(duplicates)} 个重复学号，无法确定唯一成绩，已阻止生成。"
    elif conflicts:
        message = f"{label}有 {len(conflicts)} 名学生学号相同但姓名不一致，已阻止自动填入。"
    elif missing:
        message = f"{label}缺少 {len(missing)} 名学生的可用成绩，已阻止生成。"
    else:
        message = f"{label}已按学号和姓名核对 {matched_count} 人。"
    return {
        "ready": ready,
        "record_found": True,
        "record_id": int(record["id"]),
        "label": label,
        "title": record["document_type_label"] or label,
        "updated_at": record["updated_at"] or "",
        "message": message,
        "generate_url": generate_url,
        "matched_count": matched_count,
        "missing_count": len(missing),
        "conflict_count": len(conflicts),
        "duplicate_count": len(duplicates),
        "extra_count": len(extra),
        "missing_students": missing[:12],
        "conflicts": conflicts[:12],
        "duplicate_students": duplicates[:12],
        "extra_students": extra[:12],
    }


def _students_by_number(record, document_type: str) -> dict[str, dict[str, Any]]:
    if record is None:
        return {}
    payload = _record_export_payload(record)
    structured = _as_dict(payload.get("structured"))
    result: dict[str, dict[str, Any]] = {}
    duplicates: set[str] = set()
    for student in structured.get("students") if isinstance(structured.get("students"), list) else []:
        if not isinstance(student, dict):
            continue
        number = _text(student.get("student_number"))
        if not number:
            continue
        if document_type == ORDINARY_GRADE_RECORD_TYPE:
            calculated = _as_dict(student.get("calculated_scores"))
            score = calculated.get("ordinary_score")
            if score in (None, ""):
                score = calculate_ordinary_grade_score(
                    _float_or_zero(student.get("attendance_raw_score")),
                    [
                        _float_or_zero(value)
                        for value in (
                            student.get("homework_scores")
                            if isinstance(student.get("homework_scores"), list)
                            else []
                        )
                    ],
                    _float_or_zero(student.get("assessment_score")),
                )
        else:
            score = student.get("total_score")
            if score in (None, ""):
                section_scores = student.get("section_scores") if isinstance(student.get("section_scores"), list) else []
                if section_scores and all(value not in (None, "") for value in section_scores):
                    score = sum(_float_or_zero(value) for value in section_scores)
        normalized_score = _score_or_blank(score, row_number=0, label="成绩")
        if number in result:
            duplicates.add(number)
        result[number] = {
            "student_name": _text(student.get("student_name")),
            "score": normalized_score,
        }
    for number in duplicates:
        result[number]["score"] = ""
        result[number]["duplicate"] = True
    return result


def _record_assignment_targets(record, document_type: str) -> list[dict[str, Any]]:
    """Source assignments behind a grade record, so blocked students can be
    traced back to the exact homework/exam page where their score is missing."""
    if record is None:
        return []
    structured = _as_dict(_record_export_payload(record).get("structured"))
    if document_type == EXAM_GRADE_RECORD_TYPE:
        source_exam = _as_dict(structured.get("source_exam"))
        assignment_id = _positive_int(source_exam.get("assignment_id"), 0)
        if not assignment_id:
            return []
        return [
            {
                "assignment_id": assignment_id,
                "title": _text(source_exam.get("assignment_title")) or "课程考试",
                "kind": "exam",
                "position": None,
            }
        ]
    source = _as_dict(structured.get("source_assignments"))
    homework_ids = (
        source.get("homework_assignment_ids")
        if isinstance(source.get("homework_assignment_ids"), list)
        else []
    )
    homework_meta = (
        source.get("homework_assignments")
        if isinstance(source.get("homework_assignments"), list)
        else []
    )
    targets: list[dict[str, Any]] = []
    for position, raw_id in enumerate(homework_ids):
        assignment_id = _positive_int(raw_id, 0)
        if not assignment_id:
            continue
        meta = (
            homework_meta[position]
            if position < len(homework_meta) and isinstance(homework_meta[position], dict)
            else {}
        )
        targets.append(
            {
                "assignment_id": assignment_id,
                "title": _text(meta.get("title")) or f"作业 {assignment_id}",
                "kind": "homework",
                "position": position,
            }
        )
    assessment_id = _positive_int(source.get("assessment_assignment_id"), 0)
    if assessment_id:
        assessment_meta = _as_dict(source.get("assessment_assignment"))
        targets.append(
            {
                "assignment_id": assessment_id,
                "title": _text(assessment_meta.get("title")) or f"测评 {assessment_id}",
                "kind": "assessment",
                "position": None,
            }
        )
    return targets


def _record_raw_students(record) -> dict[str, dict[str, Any]]:
    if record is None:
        return {}
    structured = _as_dict(_record_export_payload(record).get("structured"))
    result: dict[str, dict[str, Any]] = {}
    for student in structured.get("students") if isinstance(structured.get("students"), list) else []:
        if not isinstance(student, dict):
            continue
        number = _text(student.get("student_number"))
        if number and number not in result:
            result[number] = student
    return result


def _student_jump_targets(
    assignment_targets: list[dict[str, Any]],
    raw_student: dict[str, Any] | None,
    *,
    document_type: str,
    student_number: str,
) -> list[dict[str, Any]]:
    if not assignment_targets or not student_number:
        return []
    selected = assignment_targets
    if document_type == ORDINARY_GRADE_RECORD_TYPE and isinstance(raw_student, dict):
        homework_scores = (
            raw_student.get("homework_scores")
            if isinstance(raw_student.get("homework_scores"), list)
            else []
        )
        blank_targets = []
        for target in assignment_targets:
            if target.get("kind") == "homework":
                position = target.get("position")
                value = (
                    homework_scores[position]
                    if isinstance(position, int) and position < len(homework_scores)
                    else None
                )
                if value in (None, ""):
                    blank_targets.append(target)
            elif target.get("kind") == "assessment" and raw_student.get("assessment_score") in (None, ""):
                blank_targets.append(target)
        if blank_targets:
            selected = blank_targets
    return [
        {
            "assignment_id": target["assignment_id"],
            "title": target["title"],
            "kind": target.get("kind") or "",
            "url": f"/assignment/{target['assignment_id']}?locate={quote(student_number)}",
        }
        for target in selected
    ]


def _load_context(conn, *, class_offering_id: int, teacher_id: int) -> dict[str, Any]:
    if get_configured_db_engine() == "postgres":
        course_columns = {
            row["column_name"]
            for row in conn.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = current_schema() AND table_name = 'courses'
                """
            ).fetchall()
        }
    else:
        course_columns = {row["name"] for row in conn.execute("PRAGMA table_info(courses)").fetchall()}
    course_code_expr = "co.academic_course_code" if "academic_course_code" in course_columns else "''"
    row = conn.execute(
        f"""
        SELECT o.id AS class_offering_id,
               o.semester_id,
               o.semester,
               o.teacher_id,
               o.course_id,
               o.class_id,
               co.name AS course_name,
               {course_code_expr} AS course_code,
               co.school_name AS course_school,
               co.college AS course_college,
               co.department AS course_department,
               cl.name AS class_name,
               cl.school_name AS class_school,
               cl.college AS class_college,
               cl.department AS class_department,
               t.name AS teacher_name,
               t.school_name AS teacher_school,
               t.college AS teacher_college,
               t.department AS teacher_department
        FROM class_offerings o
        JOIN courses co ON co.id = o.course_id
        JOIN classes cl ON cl.id = o.class_id
        JOIN teachers t ON t.id = o.teacher_id
        WHERE o.id = ? AND o.teacher_id = ?
        LIMIT 1
        """,
        (int(class_offering_id), int(teacher_id)),
    ).fetchone()
    if not row:
        raise HTTPException(404, "课堂不存在或您无权生成期末成绩单。")
    data = dict(row)
    academic = conn.execute(
        """
        SELECT course_nature, academic_year, academic_year_name, academic_term, academic_term_name
        FROM teacher_academic_course_sync_items
        WHERE teacher_id = ?
          AND (? IS NULL OR semester_id = ? OR semester_id IS NULL)
          AND (course_id = ? OR TRIM(course_name) = TRIM(?))
        ORDER BY synced_at DESC, id DESC
        LIMIT 1
        """,
        (
            int(teacher_id),
            data.get("semester_id"),
            data.get("semester_id"),
            int(data["course_id"]),
            data.get("course_name") or "",
        ),
    ).fetchone()
    academic_data = dict(academic) if academic else {}
    period_text = data.get("semester") or ""
    return _compact_dict(
        {
            "class_offering_id": int(class_offering_id),
            "course_id": int(data["course_id"]),
            "class_id": int(data["class_id"]),
            "school": data.get("course_school") or data.get("class_school") or data.get("teacher_school") or "广西外国语学院",
            "college": data.get("course_college") or data.get("class_college") or data.get("teacher_college") or "",
            "department": data.get("course_department") or data.get("class_department") or data.get("teacher_department") or "",
            "course_name": data.get("course_name") or "",
            "course_code": data.get("course_code") or "",
            "course_nature": academic_data.get("course_nature") or "",
            "class_name": data.get("class_name") or "",
            "teacher_name": data.get("teacher_name") or "",
            "academic_year": _academic_year_token(academic_data.get("academic_year_name"))
            or _academic_year_token(academic_data.get("academic_year"))
            or _academic_year_token(period_text),
            "semester": _semester_label(academic_data.get("academic_term_name"))
            or _semester_label(academic_data.get("academic_term"))
            or _semester_label(period_text),
        }
    )


def _load_exam_roster(conn, *, class_offering_id: int, teacher_id: int):
    item = conn.execute(
        """
        SELECT *
        FROM teacher_academic_exam_roster_items
        WHERE teacher_id = ?
          AND class_offering_id = ?
          AND sync_status = 'active'
        ORDER BY synced_at DESC, id DESC
        LIMIT 1
        """,
        (int(teacher_id), int(class_offering_id)),
    ).fetchone()
    if not item:
        return None, []
    rows = conn.execute(
        """
        SELECT *
        FROM teacher_academic_exam_roster_students
        WHERE exam_roster_item_id = ?
        ORDER BY row_order ASC, id ASC
        """,
        (int(item["id"]),),
    ).fetchall()
    return item, [dict(row) for row in rows]


def _load_record(conn, record_id: int):
    row = conn.execute(
        "SELECT * FROM material_ai_import_records WHERE id = ? AND parse_status = 'completed'",
        (int(record_id),),
    ).fetchone()
    if not row:
        raise HTTPException(409, "上游成绩材料已不存在或状态发生变化，请重新核对。")
    return row


def _record_export_payload(record) -> dict[str, Any]:
    try:
        value = json.loads(record["export_payload_json"] or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        value = {}
    return value if isinstance(value, dict) else {}


def _locate_transcript_worksheet(
    workbook: Any,
) -> tuple[Any | None, int | None, dict[str, int], int]:
    candidates: list[tuple[Any, int, dict[str, int]]] = []
    for worksheet in workbook.worksheets:
        header_row, columns = _locate_header(worksheet)
        if header_row:
            candidates.append((worksheet, header_row, columns))
    if not candidates:
        return None, None, {}, 0
    worksheet, header_row, columns = candidates[0]
    return worksheet, header_row, columns, len(candidates)


def _locate_header(worksheet) -> tuple[int | None, dict[str, int]]:
    aliases = {
        "index": {"序号"},
        "class_name": {"班级"},
        "student_number": {"学号"},
        "student_name": {"姓名"},
        "ordinary_score": {"平时(必填)", "平时（必填）", "平时"},
        "final_score": {"期末(必填)", "期末（必填）", "期末"},
        "remark": {"备注"},
    }
    for row in range(1, min(worksheet.max_row, 20) + 1):
        columns: dict[str, int] = {}
        for column in range(1, min(worksheet.max_column, 30) + 1):
            value = _identity_text(worksheet.cell(row, column).value)
            for key, candidates in aliases.items():
                if value in {_identity_text(item) for item in candidates}:
                    columns[key] = column
        if set(columns) == set(aliases):
            return row, columns
    return None, {}


def _roster_signature(roster_item, roster_students: list[dict[str, Any]]) -> str:
    payload = {
        "exam_course_key": _text(roster_item["exam_course_key"] if roster_item else ""),
        "students": [
            {
                "row_order": _positive_int(student.get("row_order"), index),
                "student_number": _text(student.get("student_number")),
                "student_name": _text(student.get("student_name")),
                "class_name": _text(student.get("admin_class_name")),
            }
            for index, student in enumerate(roster_students, start=1)
        ],
    }
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _normalize_students(values: Any) -> list[dict[str, Any]]:
    if not isinstance(values, list):
        return []
    result: list[dict[str, Any]] = []
    for index, item in enumerate(values, start=1):
        if not isinstance(item, dict):
            continue
        number = _text(item.get("student_number"))
        name = _text(item.get("student_name"))
        if not number and not name:
            continue
        result.append(
            {
                **item,
                "index": _positive_int(item.get("index"), index),
                "class_name": _text(item.get("class_name")),
                "student_number": number,
                "student_name": name,
                "ordinary_score": _score_or_blank(item.get("ordinary_score"), row_number=0, label="平时"),
                "final_score": _score_or_blank(item.get("final_score"), row_number=0, label="期末"),
                "remark": _text(item.get("remark")),
            }
        )
    return result


def _students_from_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    for table in tables:
        rows = table.get("rows") if isinstance(table, dict) else None
        if not isinstance(rows, list) or len(rows) < 2:
            continue
        headers = [_identity_text(value) for value in rows[0]]
        if "学号" not in headers or "姓名" not in headers:
            continue
        result = []
        for index, row in enumerate(rows[1:], start=1):
            if not isinstance(row, list):
                continue
            values = {header: row[position] if position < len(row) else "" for position, header in enumerate(headers)}
            result.append(
                {
                    "index": values.get("序号") or index,
                    "class_name": values.get("班级") or "",
                    "student_number": values.get("学号") or "",
                    "student_name": values.get("姓名") or "",
                    "ordinary_score": values.get("平时(必填)") or values.get("平时（必填）") or "",
                    "final_score": values.get("期末(必填)") or values.get("期末（必填）") or "",
                    "remark": values.get("备注") or "",
                }
            )
        return result
    return []


def _table_from_students(students: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [list(FINAL_GRADE_TRANSCRIPT_HEADERS)]
    for student in students:
        rows.append(
            [
                student.get("index") or "",
                student.get("class_name") or "",
                student.get("student_number") or "",
                student.get("student_name") or "",
                student.get("ordinary_score") if student.get("ordinary_score") not in (None, "") else "",
                student.get("final_score") if student.get("final_score") not in (None, "") else "",
                student.get("remark") or "",
            ]
        )
    return {"title": FINAL_GRADE_TRANSCRIPT_LABEL, "rows": rows}


def _build_content_markdown(fields: dict[str, Any], students: list[dict[str, Any]]) -> str:
    lines = [
        f"# {FINAL_GRADE_TRANSCRIPT_LABEL}",
        "",
        f"- 学校：{fields.get('school') or ''}",
        f"- 学院：{fields.get('college') or ''}",
        f"- 系部：{fields.get('department') or ''}",
        f"- 课程：{fields.get('course_name') or ''}",
        f"- 课程属性：{fields.get('course_nature') or ''}",
        f"- 班级：{fields.get('class_name') or ''}",
        f"- 学年学期：{fields.get('academic_year') or ''} {fields.get('semester') or ''}",
        f"- 学生人数：{len(students)}",
        "",
        "| 序号 | 班级 | 学号 | 姓名 | 平时(必填) | 期末(必填) | 备注 |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for student in students:
        lines.append(
            "| {index} | {class_name} | {student_number} | {student_name} | {ordinary} | {final} | {remark} |".format(
                index=student.get("index") or "",
                class_name=student.get("class_name") or "",
                student_number=student.get("student_number") or "",
                student_name=student.get("student_name") or "",
                ordinary=student.get("ordinary_score") if student.get("ordinary_score") not in (None, "") else "",
                final=student.get("final_score") if student.get("final_score") not in (None, "") else "",
                remark=student.get("remark") or "",
            )
        )
    return "\n".join(lines)


def _source_generate_url(document_type: str, class_offering_id: int) -> str:
    path = (
        "/manage/teaching/ordinary-grade-records"
        if document_type == ORDINARY_GRADE_RECORD_TYPE
        else "/manage/teaching/exam-grade-records"
    )
    return f"{path}?open=classroom-generate&class_offering_id={int(class_offering_id)}"


def _teacher_from_filename(value: str) -> str:
    match = re.search(r"\[([^\]]+)]", Path(value or "").stem)
    return _text(match.group(1)) if match else ""


def _course_from_filename(value: str) -> str:
    stem = re.sub(r"\[[^\]]+]", "", Path(value or "").stem)
    stem = re.sub(r"(学生)?成绩录入模板.*$", "", stem)
    return _text(stem)


def _score_or_blank(value: Any, *, row_number: int, label: str) -> float | int | str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        location = f"第 {row_number} 行" if row_number else ""
        raise HTTPException(422, f"{location}{label}成绩“{value}”不是有效数值。")
    if number < 0 or number > 100:
        location = f"第 {row_number} 行" if row_number else ""
        raise HTTPException(422, f"{location}{label}成绩 {number:g} 超出 0 至 100。")
    return int(number) if number.is_integer() else number


def _float_or_zero(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _academic_year_token(value: Any) -> str:
    text = _text(value)
    match = re.search(r"(20\d{2})\s*[-—至/]\s*(20\d{2})", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    match = re.search(r"(20\d{2})", text)
    if match:
        start = int(match.group(1))
        return f"{start}-{start + 1}"
    return ""


def _semester_token(value: Any) -> str:
    raw = _text(value)
    if re.search(r"(?:^|[-_/])(?:12|2)\s*$", raw):
        return "2"
    if re.search(r"(?:^|[-_/])(?:3|1)\s*$", raw):
        return "1"
    text = _identity_text(raw)
    if text in {"12", "2"} or re.search(r"(第二|第2|二学期|2学期)", text):
        return "2"
    if text in {"3", "1"} or re.search(r"(第一|第1|一学期|1学期)", text):
        return "1"
    return ""


def _semester_label(value: Any) -> str:
    token = _semester_token(value)
    return "第二学期" if token == "2" else ("第一学期" if token == "1" else "")


def _duplicates(values: list[str]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for value in values:
        if not value:
            continue
        if value in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(value)
    return duplicates


def _filename_part(value: Any) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "-", _text(value)).strip(" .-")


def _identity_text(value: Any) -> str:
    return re.sub(r"[\s·•._\-—()（）\[\]【】]+", "", _text(value)).lower()


def _class_identity_text(value: Any) -> str:
    return re.sub(r"班$", "", _identity_text(value))


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _positive_int(value: Any, default: int = 0) -> int:
    try:
        number = int(float(value))
    except (TypeError, ValueError):
        return default
    return number if number > 0 else default


def _is_blank(value: Any) -> bool:
    return value in (None, "", [], {})


def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _compact_dict(value: dict[str, Any]) -> dict[str, Any]:
    return {key: item for key, item in value.items() if not _is_blank(item)}


def _string_list(value: Any) -> list[str]:
    return [str(item).strip() for item in value] if isinstance(value, list) else []


def _dedupe(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        item = str(value or "").strip()
        if item and item not in result:
            result.append(item)
    return result
